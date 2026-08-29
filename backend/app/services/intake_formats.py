"""Turn whatever a client sends into something the extractor can read.

Clients do not only send PDFs. Real intake folders contain Word documents,
Excel workbooks - including the pre-2007 .xls format some carriers still
produce - plain text, scans, and Outlook emails with the Schedule A sitting
inside as an attachment.

Two of those need work before extraction:

* An email is a wrapper. Sending the email itself to the extractor would
  extract the subject line and the signature block, not the Schedule A. The
  attachment inside is the actual document.
* Legacy .xls is a different binary format from .xlsx and is not among the
  file types the extractor accepts, so it is converted first.

Everything else is passed through untouched. When a conversion is not
possible the original file is returned with a note explaining why, so the
filing still reaches a human instead of disappearing.
"""
from __future__ import annotations

import email
import logging
import os
import re
from dataclasses import dataclass
from email import policy
from html.parser import HTMLParser
from io import BytesIO
from zipfile import BadZipFile, ZipFile

logger = logging.getLogger(__name__)

# File types the extraction engine (EyeLevel/GroundX) can read directly.
EXTRACTABLE_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".xlsx",
    ".csv",
    ".txt",
    ".png",
    ".jpg",
    ".jpeg",
    ".tif",
    ".tiff",
}
# Types we accept at intake and convert or unwrap on the way through.
CONVERTIBLE_EXTENSIONS = {".xls", ".msg", ".eml"}
# Everything a client may drop into a filing folder that we are willing to
# take. Anything outside this set is not a document we can file.
SUPPORTED_INTAKE_EXTENSIONS = EXTRACTABLE_EXTENSIONS | CONVERTIBLE_EXTENSIONS

EMAIL_EXTENSIONS = {".msg", ".eml"}
_SCHEDULE_A_NAME = re.compile(r"sched\w*\s*a\b|schedulea", re.IGNORECASE)
_EMAIL_FIELD_HINT = re.compile(
    r"\b(?:ein|naic|legal\s+name|policy|contract|premium|commission|fees?|persons?\s+covered|"
    r"employee\s+lives?|coverage\s+period|plan\s+year|indirect\s+compensation)\b",
    re.IGNORECASE,
)
_BRANDING_IMAGE_NAME = re.compile(
    r"^(?:image\d+|logo|signature|banner|spacer|facebook|linkedin|twitter)(?:[-_ ].*)?\.(?:png|jpe?g|gif|tiff?)$",
    re.IGNORECASE,
)


@dataclass
class IntakeDocument:
    """The document as the extractor should see it."""

    file_name: str
    file_bytes: bytes
    original_file_name: str | None = None
    conversion: str | None = None
    note: str | None = None

    @property
    def converted(self) -> bool:
        return bool(self.original_file_name and self.original_file_name != self.file_name)


def file_extension(file_name: str) -> str:
    return os.path.splitext(str(file_name or "").lower())[1]


def is_supported_intake_file(file_name: str) -> bool:
    return file_extension(file_name) in SUPPORTED_INTAKE_EXTENSIONS


def normalize_intake_document(file_name: str, file_bytes: bytes) -> IntakeDocument:
    """Return the file the extractor should actually receive."""
    return normalize_intake_documents(file_name, file_bytes)[0]


def normalize_intake_documents(file_name: str, file_bytes: bytes) -> list[IntakeDocument]:
    """Return every useful document represented by one intake file.

    Most files produce one document. An email may produce its substantive
    body plus one or more real attachments; keeping both prevents a signature
    logo from replacing filing values written directly in the message.
    """
    extension = file_extension(file_name)
    detected_extension = detect_actual_extension(file_bytes)
    if detected_extension and detected_extension != extension:
        corrected_name = f"{os.path.splitext(file_name)[0]}{detected_extension}"
        return [
            IntakeDocument(
                file_name=corrected_name,
                file_bytes=file_bytes,
                original_file_name=file_name,
                conversion=f"File content detected as {detected_extension.lstrip('.').upper()} instead of {extension.lstrip('.').upper() or 'unknown'}",
            )
        ]

    if extension in EMAIL_EXTENSIONS:
        return _unwrap_email_documents(file_name, file_bytes, extension)
    if extension == ".xls":
        return [_convert_xls(file_name, file_bytes)]
    return [IntakeDocument(file_name=file_name, file_bytes=file_bytes)]


def detect_actual_extension(file_bytes: bytes) -> str | None:
    """Identify supported document containers from their bytes, not their name.

    ShareFile occasionally contains a Word or Excel package named ``.pdf``.
    Passing that ZIP container to a PDF parser produces an empty extraction,
    so route recognizable formats before provider or parser selection.
    """
    data = bytes(file_bytes or b"")
    if data.startswith(b"%PDF-"):
        return ".pdf"
    if data.startswith((b"\x89PNG\r\n\x1a\n",)):
        return ".png"
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith((b"II*\x00", b"MM\x00*")):
        return ".tiff"
    if not data.startswith(b"PK\x03\x04"):
        return None
    try:
        with ZipFile(BytesIO(data)) as archive:
            names = {name.lower() for name in archive.namelist()}
    except (BadZipFile, OSError):
        return None
    if "word/document.xml" in names:
        return ".docx"
    if "xl/workbook.xml" in names:
        return ".xlsx"
    return None


# --- email ------------------------------------------------------------------


def _unwrap_email(file_name: str, file_bytes: bytes, extension: str) -> IntakeDocument:
    return _unwrap_email_documents(file_name, file_bytes, extension)[0]


def _unwrap_email_documents(file_name: str, file_bytes: bytes, extension: str) -> list[IntakeDocument]:
    try:
        if extension == ".eml":
            attachments = _read_eml_attachments(file_bytes)
            body = _read_eml_body(file_bytes)
        else:
            attachments = _read_msg_attachments(file_bytes)
            body = _read_msg_body(file_bytes)
    except Exception as exc:  # a malformed email must not break intake
        logger.warning("Could not read attachments from %s: %s", file_name, exc)
        return [
            IntakeDocument(
                file_name=file_name,
                file_bytes=file_bytes,
                note=f"Email could not be opened ({type(exc).__name__}); needs manual handling.",
            )
        ]

    meaningful_body = _clean_email_body(body)
    useful_attachments = _useful_email_attachments(attachments)
    documents = [
        IntakeDocument(
            file_name=name,
            file_bytes=data,
            original_file_name=file_name,
            conversion=f"Attachment taken from email {file_name}",
        )
        for name, data in useful_attachments
    ]

    if meaningful_body:
        stem = os.path.splitext(_clean_name(file_name))[0]
        documents.append(
            IntakeDocument(
                file_name=f"{stem} email body.txt",
                file_bytes=meaningful_body.encode("utf-8"),
                original_file_name=file_name,
                conversion=f"Relevant message body taken from email {file_name}",
            )
        )

    if documents:
        return documents
    return [
        IntakeDocument(
            file_name=file_name,
            file_bytes=file_bytes,
            note="Email has no relevant body and no attachment that can be extracted; needs manual handling.",
        )
    ]


def _read_eml_attachments(file_bytes: bytes) -> list[tuple[str, bytes]]:
    message = email.message_from_bytes(file_bytes, policy=policy.default)
    attachments: list[tuple[str, bytes]] = []
    for part in message.walk():
        if part.is_multipart():
            continue
        name = part.get_filename()
        if not name:
            continue
        payload = part.get_payload(decode=True)
        if payload:
            attachments.append((_clean_name(name), payload))
    return attachments


def _read_eml_body(file_bytes: bytes) -> str:
    message = email.message_from_bytes(file_bytes, policy=policy.default)
    body = message.get_body(preferencelist=("plain", "html"))
    if body is None:
        return ""
    try:
        content = body.get_content()
    except Exception:
        payload = body.get_payload(decode=True) or b""
        content = payload.decode(body.get_content_charset() or "utf-8", errors="ignore")
    return _html_to_text(content) if body.get_content_type() == "text/html" else str(content)


def _read_msg_attachments(file_bytes: bytes) -> list[tuple[str, bytes]]:
    """Read attachments out of an Outlook .msg file.

    A .msg is an OLE compound file: each attachment is its own storage, with
    the bytes in the 3701 property stream and the file name in 3707 (long) or
    3704 (short).
    """
    import olefile

    attachments: list[tuple[str, bytes]] = []
    with olefile.OleFileIO(BytesIO(file_bytes)) as ole:
        entries = ole.listdir(streams=True, storages=True)
        storages = {
            entry[0]
            for entry in entries
            if entry and entry[0].startswith("__attach_version1.0")
        }
        for storage in sorted(storages):
            data = _read_msg_stream(ole, [storage, "__substg1.0_37010102"])
            if not data:
                continue
            name = (
                _read_msg_text(ole, [storage, "__substg1.0_3707001F"])
                or _read_msg_text(ole, [storage, "__substg1.0_3707001E"])
                or _read_msg_text(ole, [storage, "__substg1.0_3704001F"])
                or _read_msg_text(ole, [storage, "__substg1.0_3704001E"])
                or "attachment"
            )
            attachments.append((_clean_name(name), data))
    return attachments


def _read_msg_body(file_bytes: bytes) -> str:
    """Read the plain or HTML message body from an Outlook MSG container."""
    import olefile

    with olefile.OleFileIO(BytesIO(file_bytes)) as ole:
        plain = (
            _read_msg_text(ole, ["__substg1.0_1000001F"])
            or _read_msg_text(ole, ["__substg1.0_1000001E"])
        )
        if plain:
            return plain
        html = _read_msg_stream(ole, ["__substg1.0_10130102"])
        if not html:
            return ""
        decoded = html.decode("utf-8", errors="ignore")
        if "<html" not in decoded.lower():
            decoded = html.decode("cp1252", errors="ignore")
        return _html_to_text(decoded)


def _read_msg_stream(ole, path: list[str]) -> bytes | None:
    try:
        if not ole.exists("/".join(path)):
            return None
        with ole.openstream(path) as stream:
            return stream.read()
    except Exception:
        return None


def _read_msg_text(ole, path: list[str]) -> str | None:
    raw = _read_msg_stream(ole, path)
    if not raw:
        return None
    encoding = "utf-16-le" if path[-1].endswith("001F") else "cp1252"
    try:
        return raw.decode(encoding, errors="ignore").strip("\x00").strip() or None
    except Exception:
        return None


def _best_attachment(attachments: list[tuple[str, bytes]]) -> tuple[str, bytes] | None:
    """Pick the attachment most likely to be the Schedule A.

    Signatures and logos ride along in most emails, so prefer a file the
    extractor can read, then a Schedule A-looking name, then the largest.
    """
    usable = [
        (name, data)
        for name, data in attachments
        if file_extension(name) in SUPPORTED_INTAKE_EXTENSIONS and data
    ]
    if not usable:
        return None
    named = [item for item in usable if _SCHEDULE_A_NAME.search(item[0])]
    pool = named or usable
    return max(pool, key=lambda item: len(item[1]))


def _useful_email_attachments(
    attachments: list[tuple[str, bytes]],
) -> list[tuple[str, bytes]]:
    usable = [
        (_clean_name(name), data)
        for name, data in attachments
        if file_extension(name) in SUPPORTED_INTAKE_EXTENSIONS
        and data
        and not _BRANDING_IMAGE_NAME.match(_clean_name(name))
    ]
    return sorted(
        usable,
        key=lambda item: (
            0 if _SCHEDULE_A_NAME.search(item[0]) else 1,
            0 if file_extension(item[0]) not in {".png", ".jpg", ".jpeg", ".tif", ".tiff"} else 1,
            -len(item[1]),
        ),
    )


class _VisibleHTMLText(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts: list[str] = []
        self.hidden_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag in {"style", "script", "head"}:
            self.hidden_depth += 1
        elif not self.hidden_depth and tag in {"br", "p", "div", "tr", "li"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"style", "script", "head"} and self.hidden_depth:
            self.hidden_depth -= 1
        elif not self.hidden_depth and tag in {"p", "div", "tr", "li", "table"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self.hidden_depth:
            self.parts.append(data)


def _html_to_text(value: str) -> str:
    parser = _VisibleHTMLText()
    try:
        parser.feed(str(value or ""))
        return "".join(parser.parts)
    except Exception:
        return re.sub(r"<[^>]+>", " ", str(value or ""))


def _clean_email_body(value: str) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    # Values in a reply belong to the current response. Quoted history can
    # contain a different plan or year and must not be treated as current data.
    text = re.split(
        r"\n\s*(?:-{2,}\s*Original Message\s*-{2,}|From:\s|On .+? wrote:\s*)",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    text = "\n".join(line for line in lines if line).strip()
    if len(text) < 40 or not _EMAIL_FIELD_HINT.search(text):
        return ""
    return text


def _clean_name(name: str) -> str:
    cleaned = os.path.basename(str(name or "").replace("\\", "/")).strip()
    return cleaned or "attachment"


# --- legacy Excel -----------------------------------------------------------


def _convert_xls(file_name: str, file_bytes: bytes) -> IntakeDocument:
    """Convert a pre-2007 .xls workbook to .xlsx.

    Carriers still send the old binary format, which the extractor does not
    accept. The cell grid is what matters for extraction, so the values are
    copied sheet by sheet into a modern workbook.
    """
    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError as exc:
        logger.warning("Cannot convert %s: %s", file_name, exc)
        return IntakeDocument(
            file_name=file_name,
            file_bytes=file_bytes,
            note="Legacy .xls support is not installed; needs manual handling.",
        )

    try:
        book = xlrd.open_workbook(file_contents=file_bytes)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet in book.sheets():
            target = workbook.create_sheet(title=(sheet.name or "Sheet")[:31])
            for row_index in range(sheet.nrows):
                for column_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, column_index)
                    if value == "":
                        continue
                    target.cell(row=row_index + 1, column=column_index + 1, value=value)
        if not workbook.sheetnames:
            workbook.create_sheet(title="Sheet1")
        buffer = BytesIO()
        workbook.save(buffer)
    except Exception as exc:
        logger.warning("Could not convert %s to xlsx: %s", file_name, exc)
        return IntakeDocument(
            file_name=file_name,
            file_bytes=file_bytes,
            note=f"Legacy .xls could not be converted ({type(exc).__name__}); needs manual handling.",
        )

    converted_name = f"{os.path.splitext(file_name)[0]}.xlsx"
    return IntakeDocument(
        file_name=converted_name,
        file_bytes=buffer.getvalue(),
        original_file_name=file_name,
        conversion=f"Converted from legacy Excel format ({file_name})",
    )
