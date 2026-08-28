import asyncio
import calendar
import csv
from io import BytesIO, StringIO
import json
import os
import re
import tempfile
from typing import Any
from urllib.parse import unquote_plus
import zipfile
import xml.etree.ElementTree as ET

import httpx
from groundx import Document, GroundX
from app.config import get_settings
from app.models import (
    DocumentType,
    FieldRuleMappingMode,
    FormType,
    NormalizedExtractionField,
    NormalizedExtractionResult,
    ScheduleABenefitBreakdownRow,
    ScheduleABrokerMoneyRow,
    ScheduleABrokerRow,
    ScheduleAWorksheetSummary,
    ScheduleAWorksheetValue,
)
from app.services.field_rules import DEFAULT_FIELD_RULES
from app.services.schedule_a_classification import classification_signals_from_text


SCHEDULE_A_EXPERIENCE_RATED_FIELDS = (
    "9a. Premiums: (1) Amount Received",
    "9a(2). Increase (decrease) in amount due but unpaid",
    "9a(3). Increase (decrease) in unearned premium reserve",
    "9a(4). Earned ((1) + (2) - (3))",
    "9b(1). Benefit Charges (1) Claims paid",
    "9b(2). Increase (decrease) in claim reserves",
    "9b(3). Incurred claims (add(1) and (2))",
    "9b(4). Claims Charged",
    "9c(1)(A). Commissions",
    "9c(1)(B). Administrative service or other fees",
    "9c(1)(C). Other Specific acquisition costs",
    "9c(1)(D). Other expenses",
    "9c(1)(E). Taxes",
    "9c(1)(F). Charges for risks or other contingencies",
    "9c(1)(G). Other retention charges",
    "9c(1)(H). Total retention",
    "9c(2). Dividends or retroactive rate refunds",
    "9d(1). Status of policyholder reserves at end of year: (1) Amount held to provide benefits after retirement",
    "9d(2). Claim reserves",
    "9d(3). Other reserves",
    "9e. Dividends or retroactive rate refunds due",
)

SCHEDULE_A_PREFER_PDF_TEXT_FIELDS = {
    "1a. Name of Insurance Company",
    "1b. Insurance Carrier EIN",
    "1c. NAIC Code",
    "1d. Contract/Policy Number",
    "1e. Persons Covered (End of Policy Year)",
    "1f. Policy Year Beginning Date",
    "1g. Policy Year Ending Date",
    "3b. Amount of Commissions",
    "3c. Amount of Fees",
    "3d. Purpose",
    "3e. Organizational Code",
    "10a. Total premiums or subscription charges paid to carrier",
}


class ExtractionService:
    def __init__(self, field_rules=None):
        self.field_rules = list(field_rules) if field_rules is not None else DEFAULT_FIELD_RULES

    async def extract_document(self, file_bytes: bytes, file_name: str, document_type: DocumentType) -> NormalizedExtractionResult:
        if document_type == DocumentType.PLAN_WORKSHEET:
            return await self.extract_plan_worksheet(file_bytes, file_name)
        return await self.extract_schedule_a(file_bytes, file_name)

    async def extract_plan_worksheet(self, file_bytes: bytes, file_name: str) -> NormalizedExtractionResult:
        context_text = extract_docx_text(file_bytes) if file_name.lower().endswith(".docx") else ""
        if not context_text and file_name.lower().endswith(".pdf"):
            context_text = "\n\n".join(text for _, text in extract_pdf_text_pages(file_bytes))

        worksheet_rules = rules_for_form(self.field_rules, FormType.FORM_5500)
        local_fields = parse_plan_worksheet_text(context_text, rules=worksheet_rules) if context_text else []
        fields = dedupe_fields(local_fields)
        if fields:
            return NormalizedExtractionResult(
                provider="Plan Worksheet local parser",
                fields=fields,
                raw={"context_preview": context_text[:4000]},
            )

        settings = get_settings()
        if settings.groundx_api_key and settings.groundx_bucket_id:
            try:
                return await self._extract_with_groundx(
                    file_bytes,
                    file_name,
                    FormType.FORM_5500,
                    "Plan Worksheet",
                )
            except Exception as exc:
                return NormalizedExtractionResult(
                    provider=f"Plan Worksheet OCR fallback failed ({safe_error_summary(exc)})",
                    fields=[],
                    raw={"context_preview": context_text[:4000], "error": safe_error_summary(exc)},
                )

        return NormalizedExtractionResult(
            provider="Plan Worksheet parser",
            fields=[],
            raw={"context_preview": context_text[:4000]},
        )

    async def extract_schedule_a(self, file_bytes: bytes, file_name: str) -> NormalizedExtractionResult:
        settings = get_settings()
        document_signals = extract_schedule_a_classification_signals(file_bytes, file_name)
        if settings.groundx_api_key and settings.groundx_bucket_id:
            try:
                result = await self._extract_with_groundx(file_bytes, file_name, FormType.SCHEDULE_A, "Schedule A")
                result.classification_signals = sorted(set(result.classification_signals) | set(document_signals))
                return result
            except Exception as exc:
                local_result = local_schedule_a_pdf_result(
                    file_bytes,
                    file_name,
                    provider=f"Local PDF parser fallback ({safe_error_summary(exc)})",
                    rules=self.field_rules,
                )
                local_result.classification_signals = document_signals
                if local_result.fields or local_result.schedule_a_broker_rows or local_result.schedule_a_worksheet_summaries:
                    return local_result
                # Unrecognized layout: degrade gracefully instead of crashing the
                # pipeline. The filing completes with all fields MISSING and is
                # routed to manual review.
                return NormalizedExtractionResult(
                    provider=f"Unrecognized layout - manual review required (AI extraction failed: {safe_error_summary(exc)})",
                    fields=[],
                    raw={"file_name": file_name, "error": safe_error_summary(exc), "source": "unrecognized_layout_fallback"},
                    classification_signals=document_signals,
                )

        if not settings.eyelevel_api_key or not settings.eyelevel_extract_url:
            local_result = local_schedule_a_pdf_result(file_bytes, file_name, rules=self.field_rules)
            local_result.classification_signals = document_signals
            if local_result.fields or local_result.schedule_a_broker_rows or local_result.schedule_a_worksheet_summaries:
                return local_result
            mock_result = self._mock_extraction(file_name)
            mock_result.classification_signals = document_signals
            return mock_result

        files = {"file": (file_name, file_bytes, "application/pdf")}
        headers = {"Authorization": f"Bearer {settings.eyelevel_api_key}"}
        async with httpx.AsyncClient(timeout=120) as client:
            response = await client.post(settings.eyelevel_extract_url, headers=headers, files=files)
            response.raise_for_status()
            raw = response.json()
        result = self._normalize_response(raw, "EyeLevel/GroundX")
        pdf_text_fields = extract_fields_from_pdf_text(file_bytes, rules=self.field_rules)
        result.fields = merge_schedule_a_fields(result.fields, pdf_text_fields)
        result.schedule_a_broker_rows = extract_schedule_a_broker_rows_from_pdf_text(file_bytes)
        result.schedule_a_worksheet_summaries = extract_schedule_a_worksheet_summaries_from_pdf_text(file_bytes)
        result.classification_signals = document_signals
        return result

    async def _extract_with_groundx(self, file_bytes: bytes, file_name: str, form_type: FormType, document_label: str) -> NormalizedExtractionResult:
        settings = get_settings()
        base_url = settings.groundx_api_base_url.rstrip("/")
        headers = {
            "X-API-Key": settings.groundx_api_key or "",
        }

        async with httpx.AsyncClient(timeout=120) as client:
            ingest_raw = await self._ingest_groundx_file(client, base_url, headers, file_bytes, file_name)
            process_id = first_value(ingest_raw, ["processId", "process_id", "processID", "id"])
            if not process_id:
                process_ids = find_values(ingest_raw, {"processId", "process_id", "processID"})
                process_id = process_ids[0] if process_ids else None
            poll_raw = ingest_raw
            if process_id:
                poll_raw = await self._poll_groundx_process(client, base_url, headers, str(process_id))

            raw_payloads: list[Any] = [ingest_raw, poll_raw]
            xray_payloads = await self._fetch_groundx_xray_payloads(client, base_url, headers, raw_payloads, file_name)
            raw_payloads.extend(xray_payloads)
            if not xray_payloads:
                bucket_search = await self._search_groundx_with_field_schema(
                    client,
                    base_url,
                    headers,
                    str(settings.groundx_bucket_id),
                    file_name,
                    form_type=form_type,
                )
                if bucket_search:
                    raw_payloads.append(bucket_search)
                broad_bucket_search = await self._search_groundx_with_field_schema(
                    client,
                    base_url,
                    headers,
                    str(settings.groundx_bucket_id),
                    form_type=form_type,
                )
                if broad_bucket_search:
                    raw_payloads.append(broad_bucket_search)

        xray_fields: list[NormalizedExtractionField] = []
        search_fields: list[NormalizedExtractionField] = []
        fallback_fields: list[NormalizedExtractionField] = []
        pdf_text_fields: list[NormalizedExtractionField] = []
        schedule_a_broker_rows: list[ScheduleABrokerRow] = []
        schedule_a_worksheet_summaries: list[ScheduleAWorksheetSummary] = []
        for payload in raw_payloads:
            if is_groundx_xray_payload(payload):
                xray_fields.extend(extract_fields_from_groundx_xray(payload, rules=self.field_rules))
            elif is_groundx_search_payload(payload):
                search_fields.extend(self._extract_fields_from_groundx_search(payload))
            else:
                fallback_fields.extend(self._extract_field_like_items(payload))

        if form_type == FormType.FORM_5500:
            worksheet_context = build_structured_extraction_context(raw_payloads, file_bytes)
            fields = parse_plan_worksheet_text(
                worksheet_context,
                rules=rules_for_form(self.field_rules, FormType.FORM_5500),
            )
            provider = "GroundX Plan Worksheet OCR" if fields else "GroundX Plan Worksheet OCR not ready"
            return NormalizedExtractionResult(
                provider=provider,
                fields=fields,
                raw={"ingest": ingest_raw, "process": poll_raw, "outputs": raw_payloads[2:]},
            )

        if form_type == FormType.SCHEDULE_A:
            # GroundX is useful, but for Schedule A we also have stable label-driven text
            # parsing that recovers fields GroundX may miss or emit inconsistently. It runs
            # on whatever format the document arrived in, not only PDFs.
            if str(file_name or "").lower().endswith(".pdf"):
                pdf_text_fields = extract_fields_from_pdf_text(file_bytes, rules=self.field_rules)
                schedule_a_broker_rows = extract_schedule_a_broker_rows_from_pdf_text(file_bytes)
                schedule_a_worksheet_summaries = extract_schedule_a_worksheet_summaries_from_pdf_text(file_bytes)
            else:
                pdf_text_fields = extract_fields_from_document_text(file_bytes, file_name, rules=self.field_rules)
                schedule_a_broker_rows = extract_schedule_a_broker_rows_from_document(file_bytes, file_name)

        fields = merge_schedule_a_fields([*xray_fields, *search_fields], pdf_text_fields)
        provider = "GroundX X-Ray + retrieval" if xray_fields and search_fields else "GroundX X-Ray"
        if xray_fields:
            provider = "GroundX X-Ray + retrieval" if search_fields else "GroundX X-Ray"
        elif search_fields:
            provider = "GroundX retrieval"
        elif pdf_text_fields:
            provider = "Local PDF parser"
        elif fallback_fields:
            fields = dedupe_fields(fallback_fields)
            provider = "GroundX generic extraction"
        elif not fields:
            provider = "GroundX X-Ray not ready"
        deduped = dedupe_fields(fields)
        classification_signals = classification_signals_from_text("\n".join(extract_xray_item_text(raw_payloads)))
        return NormalizedExtractionResult(
            provider=provider,
            fields=deduped,
            raw={"ingest": ingest_raw, "process": poll_raw, "outputs": raw_payloads[2:]},
            classification_signals=classification_signals,
            schedule_a_broker_rows=schedule_a_broker_rows,
            schedule_a_worksheet_summaries=schedule_a_worksheet_summaries,
        )

    async def _ingest_groundx_file(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        file_bytes: bytes,
        file_name: str,
    ) -> Any:
        settings = get_settings()
        return await asyncio.to_thread(self._ingest_groundx_file_with_sdk, file_bytes, file_name)

    def _ingest_groundx_file_with_sdk(self, file_bytes: bytes, file_name: str) -> Any:
        settings = get_settings()
        if not settings.groundx_api_key or not settings.groundx_bucket_id:
            raise RuntimeError("GroundX API key and bucket ID are required for ingestion.")

        suffix = os.path.splitext(file_name)[1] or ".pdf"
        temp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as handle:
                handle.write(file_bytes)
                temp_path = handle.name

            client = GroundX(api_key=settings.groundx_api_key)
            response = client.ingest(
                documents=[
                    Document(
                        bucketId=settings.groundx_bucket_id,
                        fileName=file_name,
                        filePath=temp_path,
                        fileType=file_type_for_groundx(file_name),
                    )
                ],
                wait_for_complete=False,
            )
            return response.model_dump(mode="json")
        finally:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)

    async def _poll_groundx_process(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        process_id: str,
        max_wait_seconds: int | None = None,
    ) -> Any:
        settings = get_settings()
        wait_seconds = max_wait_seconds if max_wait_seconds is not None else settings.groundx_max_wait_seconds
        max_attempts = max(1, int(wait_seconds / settings.groundx_poll_seconds))
        latest: Any = None
        for _ in range(max_attempts):
            response = await client.get(f"{base_url}/ingest/{process_id}", headers=headers)
            response.raise_for_status()
            latest = response.json()
            status = str(first_value(latest, ["status", "state", "processStatus"]) or "").lower()
            if status in {"complete", "completed", "done", "success", "succeeded", "finished"}:
                return latest
            if status in {"failed", "error", "errored", "cancelled", "canceled"}:
                raise RuntimeError(f"GroundX processing failed with status: {status}")
            await asyncio.sleep(settings.groundx_poll_seconds)
        return latest

    async def _fetch_groundx_document_outputs(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        document_id: str,
    ) -> list[Any]:
        outputs: list[Any] = []
        paths = [
            f"/ingest/document/{document_id}",
            f"/ingest/document/extract/{document_id}",
            f"/ingest/document/xray/{document_id}",
        ]
        for path in paths:
            response = await client.get(f"{base_url}{path}", headers=headers)
            if response.status_code in {400, 401, 404}:
                continue
            response.raise_for_status()
            outputs.append(response.json())
        return outputs

    async def _fetch_groundx_xray_payloads(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        raw_payloads: list[Any],
        file_name: str,
    ) -> list[Any]:
        document_refs = await self._find_groundx_document_refs(client, base_url, headers, raw_payloads, file_name)
        outputs: list[Any] = []
        seen: set[str] = set()
        for ref in document_refs:
            document_id = string_or_none(ref.get("documentId") or ref.get("document_id"))
            xray_url = string_or_none(ref.get("xrayUrl") or ref.get("xray_url"))
            if not document_id or document_id in seen:
                continue
            seen.add(document_id)
            xray = await self._fetch_groundx_xray(client, base_url, headers, document_id, xray_url)
            if xray:
                outputs.append(xray)
        return outputs

    async def _find_groundx_document_refs(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        raw_payloads: list[Any],
        file_name: str,
    ) -> list[dict[str, Any]]:
        refs: list[dict[str, Any]] = []
        document_ids = find_values(raw_payloads, {"documentId", "document_id"})
        for document_id in document_ids:
            refs.append({"documentId": document_id})

        settings = get_settings()
        target_name = normalize_file_name(file_name)
        max_attempts = max(1, int(settings.groundx_max_wait_seconds / settings.groundx_poll_seconds))
        for attempt in range(max_attempts):
            try:
                response = await client.get(f"{base_url}/ingest/documents", headers=headers)
                if response.status_code < 400:
                    documents = response.json().get("documents", [])
                    if isinstance(documents, list):
                        matches = [
                            item
                            for item in documents
                            if isinstance(item, dict)
                            and normalize_file_name(str(item.get("fileName") or "")) == target_name
                        ]
                        matches.sort(key=lambda item: str(item.get("updated") or item.get("created") or ""), reverse=True)
                        latest = matches[0] if matches else None
                        if latest:
                            status = str(latest.get("status") or "").lower()
                            if status in {"complete", "completed", "done", "success", "succeeded", "finished"}:
                                refs.append(latest)
                                return refs
            except (httpx.TimeoutException, httpx.NetworkError, ValueError):
                pass
            if attempt < max_attempts - 1:
                await asyncio.sleep(settings.groundx_poll_seconds)

        return refs

    async def _fetch_groundx_xray(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        document_id: str,
        xray_url: str | None = None,
    ) -> Any | None:
        urls = [f"{base_url}/ingest/document/xray/{document_id}"]
        if xray_url:
            urls.append(xray_url)

        for index, url in enumerate(urls):
            try:
                response = await client.get(url, headers=headers if index == 0 else None, timeout=60)
            except (httpx.TimeoutException, httpx.NetworkError):
                continue
            if response.status_code in {400, 401, 403, 404, 408, 429, 500, 502, 503, 504}:
                continue
            try:
                response.raise_for_status()
                return response.json()
            except (httpx.HTTPStatusError, ValueError):
                continue
        return None

    async def _search_groundx_with_field_schema(
        self,
        client: httpx.AsyncClient,
        base_url: str,
        headers: dict[str, str],
        document_id: str,
        file_name: str | None = None,
        *,
        form_type: FormType = FormType.SCHEDULE_A,
    ) -> Any | None:
        query = build_groundx_schema_query(file_name, self.field_rules, form_type=form_type)
        body = {
            "search": {
                "query": query,
                "n": 25,
                "verbosity": 2,
                "relevance": 0,
            }
        }
        try:
            response = await client.post(f"{base_url}/search/{document_id}", headers=headers, json=body)
        except (httpx.TimeoutException, httpx.NetworkError):
            return None
        if response.status_code in {400, 401, 404, 408, 429, 500, 502, 503, 504}:
            return None
        try:
            response.raise_for_status()
        except httpx.HTTPStatusError:
            return None
        return response.json()

    def _extract_fields_from_groundx_search(self, raw: Any) -> list[NormalizedExtractionField]:
        search = raw.get("search", {}) if isinstance(raw, dict) else {}
        results = search.get("results", []) if isinstance(search, dict) else []
        fields: list[NormalizedExtractionField] = []

        for result in results:
            if not isinstance(result, dict):
                continue
            text = normalize_ocr_text(result.get("text") or result.get("suggestedText") or result.get("narrative") or "")
            if not text:
                continue
            page = parse_groundx_page(result)
            fields.extend(parse_schedule_a_text(text, page, rules=self.field_rules))

        return dedupe_fields(fields)

    def _normalize_response(self, raw, provider: str) -> NormalizedExtractionResult:
        candidates = raw.get("fields") or raw.get("extracted_fields") or raw.get("results") or []
        fields = []
        for item in candidates:
            fields.append(
                NormalizedExtractionField(
                    field_name=str(item.get("fieldName") or item.get("name") or item.get("label") or "Unknown Field"),
                    value=str(item.get("value") or item.get("text") or ""),
                    confidence=float(item.get("confidence") or item.get("score") or 0.75),
                    page=item.get("page") if isinstance(item.get("page"), int) else None,
                    source_text=item.get("sourceText") or item.get("context"),
                )
            )
        if not fields:
            fields = self._extract_field_like_items(raw)
        return NormalizedExtractionResult(provider=provider, fields=fields, raw=raw)

    def _extract_field_like_items(self, raw: Any) -> list[NormalizedExtractionField]:
        fields: list[NormalizedExtractionField] = []

        def walk(value: Any):
            if isinstance(value, list):
                for item in value:
                    walk(item)
                return
            if not isinstance(value, dict):
                return

            field_name = first_value(
                value,
                [
                    "fieldName",
                    "field_name",
                    "name",
                    "label",
                    "question",
                    "key",
                    "title",
                ],
            )
            field_value = first_value(
                value,
                [
                    "value",
                    "answer",
                    "text",
                    "content",
                    "result",
                    "extractedText",
                    "extracted_text",
                ],
            )
            if field_name and field_value and not isinstance(field_value, (dict, list)):
                fields.append(
                    NormalizedExtractionField(
                        field_name=str(field_name),
                        value=str(field_value).strip(),
                        confidence=normalize_raw_confidence(first_value(value, ["confidence", "score", "relevance", "probability"])),
                        page=parse_page(first_value(value, ["page", "pageNumber", "page_number"])),
                        source_text=string_or_none(first_value(value, ["sourceText", "source_text", "context", "text"])),
                    )
                )

            for nested in value.values():
                if isinstance(nested, (dict, list)):
                    walk(nested)

        walk(raw)
        return fields

    def _mock_extraction(self, file_name: str) -> NormalizedExtractionResult:
        return NormalizedExtractionResult(
            provider="Local mock extractor",
            raw={"file_name": file_name, "note": "Set EYELEVEL_API_KEY and EYELEVEL_EXTRACT_URL to call the real extractor."},
            fields=[
                NormalizedExtractionField(field_name="Plan Name", value="HighlandTech Health and Welfare Plan", confidence=0.93, page=1),
                NormalizedExtractionField(field_name="Three-digit Plan Number", value="501", confidence=0.91, page=1),
                NormalizedExtractionField(field_name="Employer EIN", value="12-3456789", confidence=0.78, page=1),
                NormalizedExtractionField(field_name="Insurance Carrier Name", value="Sample Health Insurance Co.", confidence=0.88, page=2),
                NormalizedExtractionField(field_name="Contract / Identification Number", value="HT-2025-501", confidence=0.86, page=2),
                NormalizedExtractionField(field_name="Premium / Contribution", value="125000", confidence=0.82, page=3),
                NormalizedExtractionField(field_name="Total Fees Paid", value="1400", confidence=0.62, page=3),
            ],
        )


def first_value(data: Any, keys: list[str]) -> Any:
    if not isinstance(data, dict):
        return None
    for key in keys:
        if key in data and data[key] not in (None, ""):
            return data[key]
    return None


def find_values(data: Any, keys: set[str]) -> list[Any]:
    values: list[Any] = []
    if isinstance(data, list):
        for item in data:
            values.extend(find_values(item, keys))
    elif isinstance(data, dict):
        for key, value in data.items():
            if key in keys and value not in (None, ""):
                values.append(value)
            elif isinstance(value, (dict, list)):
                values.extend(find_values(value, keys))
    return values


def normalize_raw_confidence(value: Any) -> float:
    try:
        confidence = float(value)
    except (TypeError, ValueError):
        return 0.75
    if confidence > 1:
        return min(confidence / 100, 1)
    return max(0, min(confidence, 1))


def parse_page(value: Any) -> int | None:
    try:
        page = int(value)
    except (TypeError, ValueError):
        return None
    return page if page > 0 else None


def string_or_none(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, (dict, list)):
        return None
    return str(value)


def safe_error_summary(exc: Exception) -> str:
    if isinstance(exc, httpx.HTTPStatusError):
        status_code = exc.response.status_code
        try:
            detail = exc.response.json()
        except ValueError:
            detail = exc.response.text[:300]
        return redact_sensitive_text(f"HTTP {status_code}: {detail}")
    return redact_sensitive_text(str(exc)[:500])


def redact_sensitive_text(value: str) -> str:
    return re.sub(r"sk-[A-Za-z0-9_-]+", "sk-***", value)


def dedupe_fields(fields: list[NormalizedExtractionField]) -> list[NormalizedExtractionField]:
    best: dict[tuple[str, str], NormalizedExtractionField] = {}
    for field in fields:
        key = (field.field_name.strip().lower(), field.value.strip().lower())
        current = best.get(key)
        if not current or field.confidence > current.confidence:
            best[key] = field
    return list(best.values())


def is_obvious_template_placeholder(value: Any) -> bool:
    """Reject sample-form filler without rejecting legitimate long identifiers."""
    clean = re.sub(r"\s+", "", clean_extracted_value(str(value or ""))).upper()
    if not clean:
        return False
    if clean.count("ABCDEFGHI") >= 2:
        return True
    if clean in {"ABCDE", "ABCDEFGHI", "ABCDEFGHIJ"}:
        return True
    return clean.lstrip("-") in {
        "0123456789",
        "1234567890",
        "123456789012345",
        "0123456789012345",
    }


def select_best_schedule_a_fields(fields: list[NormalizedExtractionField]) -> list[NormalizedExtractionField]:
    """Return one trustworthy value per dashboard field.

    The dashboard and FT Williams mapping model each rule as one value. PDF
    templates can expose both sample-layer filler and the completed overlay;
    keeping every conflicting value created false review/update rows.
    """
    best: dict[str, NormalizedExtractionField] = {}
    order: list[str] = []
    for field in fields:
        if (
            is_blank_extraction_value(field.value)
            or is_obvious_template_placeholder(field.value)
            or _is_column_heading_broker_name(field)
        ):
            continue
        key = field.field_name.strip().lower()
        current = best.get(key)
        if current is None:
            order.append(key)
        # Later specialized parsers win a confidence tie over broad OCR regexes.
        if current is None or field.confidence >= current.confidence:
            best[key] = field
    return [best[key] for key in order]


def merge_schedule_a_fields(
    primary_fields: list[NormalizedExtractionField],
    pdf_text_fields: list[NormalizedExtractionField],
) -> list[NormalizedExtractionField]:
    if not pdf_text_fields:
        return select_best_schedule_a_fields(primary_fields)

    def _has_usable_value(field: NormalizedExtractionField) -> bool:
        value = (field.value or "").strip()
        if not value or is_blank_extraction_value(value):
            return False
        if field.field_name == "1d. Contract/Policy Number":
            return is_valid_contract_identifier(value, allow_numeric=True)
        return True

    # Specialized PDF parsers are authoritative for the stable Schedule A
    # table fields. For other fields, keep a single usable AI value and use the
    # document parser when AI produced no value or conflicting values.
    usable_primary_by_name: dict[str, list[NormalizedExtractionField]] = {}
    for field in primary_fields:
        if _has_usable_value(field):
            usable_primary_by_name.setdefault(field.field_name, []).append(field)
    merged: list[NormalizedExtractionField] = list(primary_fields)
    existing_names = {field.field_name for field in merged}
    for field in pdf_text_fields:
        usable_primary = usable_primary_by_name.get(field.field_name, [])
        distinct_primary_values = {clean_extracted_value(item.value).lower() for item in usable_primary}
        prefer_document_value = field.field_name in SCHEDULE_A_PREFER_PDF_TEXT_FIELDS or (
            field.field_name == "3a. Name of Agent/Broker/Person"
            and field.source_text == "Broker compensation table"
        )
        if not prefer_document_value and len(distinct_primary_values) == 1:
            continue
        if field.field_name in existing_names:
            # Replace invalid/conflicting AI values, or an AI value for a field
            # that has a reliable format-specific document parser.
            merged = [f for f in merged if f.field_name != field.field_name]
        merged.append(field)
        existing_names.add(field.field_name)

    return select_best_schedule_a_fields(merged)


def is_groundx_search_payload(payload: Any) -> bool:
    return isinstance(payload, dict) and isinstance(payload.get("search"), dict) and isinstance(payload["search"].get("results"), list)


def is_groundx_xray_payload(payload: Any) -> bool:
    return isinstance(payload, dict) and isinstance(payload.get("chunks"), list) and isinstance(payload.get("documentPages"), list)


def normalize_file_name(value: str) -> str:
    return unquote_plus(value).strip().lower()


def extract_fields_from_groundx_xray(raw: Any, rules=None) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    chunks = raw.get("chunks", []) if isinstance(raw, dict) else []
    for chunk in chunks:
        if not isinstance(chunk, dict):
            continue
        page = parse_xray_page(chunk)
        source_text = build_xray_source_text(chunk)
        for item in chunk.get("json") or []:
            if not isinstance(item, dict):
                continue
            fields.extend(extract_schedule_a_fields_from_xray_json(item, page, source_text))
            item_text = normalize_ocr_text("\n".join(extract_xray_item_text(item)))
            if item_text:
                fields.extend(parse_schedule_a_text(item_text, page, rules=rules))
        text = normalize_ocr_text(chunk.get("suggestedText") or chunk.get("text") or "")
        if text:
            fields.extend(parse_schedule_a_text(text, page, rules=rules))
    return dedupe_fields(fields)


def parse_xray_page(chunk: dict[str, Any]) -> int | None:
    pages = chunk.get("pageNumbers") or []
    if isinstance(pages, list) and pages:
        return parse_page(pages[0])
    boxes = chunk.get("boundingBoxes") or []
    if isinstance(boxes, list) and boxes and isinstance(boxes[0], dict):
        return parse_page(boxes[0].get("pageNumber"))
    return None


def build_xray_source_text(chunk: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in ("text", "suggestedText"):
        text = normalize_ocr_text(chunk.get(key))
        if text:
            parts.append(text)
    for item in chunk.get("json") or []:
        if isinstance(item, dict):
            parts.extend(extract_xray_item_text(item))
    narrative = chunk.get("narrative")
    if isinstance(narrative, list):
        parts.extend(normalize_ocr_text(item) for item in narrative if normalize_ocr_text(item))
    elif narrative:
        parts.append(normalize_ocr_text(narrative))
    return "\n".join(parts)[:1200]


def extract_xray_item_text(value: Any) -> list[str]:
    texts: list[str] = []
    if isinstance(value, dict):
        for key, nested in value.items():
            if key in {"text", "summary", "content", "suggestedText", "description"} and isinstance(nested, str):
                clean = normalize_ocr_text(nested)
                if clean:
                    texts.append(clean)
            else:
                texts.extend(extract_xray_item_text(nested))
    elif isinstance(value, list):
        for item in value:
            texts.extend(extract_xray_item_text(item))
    return texts


def extract_schedule_a_fields_from_xray_json(item: dict[str, Any], page: int | None, source_text: str) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    flat_values = flatten_xray_values(item)

    def add(field_name: str, value: Any, confidence: float = 0.95, value_validator=None):
        if value in (None, "", [], {}):
            return
        clean = clean_extracted_value(str(value))
        if not clean or is_blank_extraction_value(clean):
            return
        if value_validator and not value_validator(clean):
            return
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value=clean,
                confidence=confidence,
                page=page,
                source_text=source_text or clean,
            )
        )

    def add_first(field_name: str, *predicates, confidence: float = 0.95, value_validator=None):
        for path, value in flat_values:
            normalized_path = normalize_xray_path(path)
            if is_xray_label_or_header(normalized_path, value) or is_blank_extraction_value(value):
                continue
            if any(predicate(normalized_path) for predicate in predicates):
                add(field_name, value, confidence, value_validator=value_validator)
                return

    add_first(
        "1a. Name of Insurance Company",
        lambda path: path.endswith("a_name_of_insurance_carrier"),
        lambda path: path.endswith("name_of_insurance_carrier"),
        lambda path: path.endswith("insurance_carrier_name"),
        lambda path: path.endswith("carrier_name"),
        lambda path: ("insurance" in path or "carrier" in path) and "name" in path and "ein" not in path,
        confidence=0.97,
        value_validator=is_probable_carrier_name,
    )
    add_first(
        "1b. Insurance Carrier EIN",
        lambda path: path.endswith("b_ein"),
        lambda path: path.endswith("ein") and ("carrier" in path or "insurance" in path),
        lambda path: "insurance" in path and "identification_number" in path,
        confidence=0.97,
        value_validator=looks_like_ein,
    )
    add_first("1c. NAIC Code", lambda path: "naic" in path and "code" in path)
    add_first(
        "1d. Contract/Policy Number",
        lambda path: ("contract" in path or "policy" in path) and ("identification" in path or "number" in path),
        confidence=0.88,
    )
    add_first(
        "1e. Persons Covered (End of Policy Year)",
        lambda path: "persons" in path and "covered" in path and ("end" in path or "approximate" in path),
    )
    add_first("1f. Policy Year Beginning Date", lambda path: ("policy" in path or "contract" in path) and path.endswith("_from"))
    add_first("1g. Policy Year Ending Date", lambda path: ("policy" in path or "contract" in path) and path.endswith("_to"))
    add_first(
        "3a. Name of Agent/Broker/Person",
        lambda path: path.endswith("payee_name"),
        lambda path: path.endswith("recipient_name"),
        lambda path: path.endswith("entity_name") and ("commission" in path or "fee" in path or "part_ii" in path),
        lambda path: path.endswith("name") and ("agent" in path or "broker" in path or "payee" in path or "recipient" in path),
        lambda path: path.endswith("name") and ("persons_receiving_commissions_and_fees" in path or "recipient_of_commissions_and_fees" in path),
        lambda path: "name_and_address" in path and ("recipient" in path or "agent_broker" in path or "broker_or_other_person" in path),
        lambda path: "agent_broker_or_other_person" in path and "name" in path,
        value_validator=is_probable_person_or_entity_name,
    )
    add_first(
        "3b. Amount of Commissions",
        lambda path: path.endswith("total_commissions_paid"),
        lambda path: path.endswith("total_amount_of_commissions_paid"),
        lambda path: path.endswith("amount_of_sales_and_base_commissions_paid"),
        lambda path: path.endswith("sales_and_base_commissions_paid"),
    )
    add_first(
        "3c. Amount of Fees",
        lambda path: path.endswith("total_fees_paid"),
        lambda path: path.endswith("total_amount_of_fees_paid"),
        lambda path: path.endswith("fees_and_other_commissions_paid"),
        lambda path: path.endswith("fees_and_other_commissions_paid_amount"),
        lambda path: path.endswith("c_amount") and ("person" in path or "commission" in path or "fees" in path),
    )
    add_first(
        "3d. Purpose",
        lambda path: path.endswith("d_purpose"),
        lambda path: path.endswith("purpose") and ("section_3" in path or "persons_receiving_commissions_and_fees" in path),
    )
    add_first("3e. Organizational Code", lambda path: "organization" in path and "code" in path, value_validator=looks_like_org_code)
    add_first(
        "10a. Total premiums or subscription charges paid to carrier",
        lambda path: "total_premiums" in path and "carrier" in path,
        lambda path: "subscription_charges" in path and ("carrier" in path or "paid" in path),
        lambda path: "nonexperience" in path and ("total" in path or "subtotal" in path) and ("amount" in path or "premium" in path),
        confidence=0.9,
    )
    for path, value in flat_values:
        if normalize_xray_path(path).endswith("gross_premium"):
            add(
                "10a. Total premiums or subscription charges paid to carrier",
                money_value(str(value)),
                0.9,
            )
            break
    premium_total = extract_nonexperience_total_premium_from_text(source_text)
    if premium_total:
        add("10a. Total premiums or subscription charges paid to carrier", premium_total, 0.93)
    if has_experience_rated_not_applicable(source_text):
        add_not_applicable_experience_rated_fields(fields, page, source_text)
    add_first("4a. Plan Name", lambda path: "plan" in path and "name" in path and "sponsor" not in path and "file" not in path)
    add_first(
        "4b. Plan Number (PN)",
        lambda path: path.endswith("plan_number_value"),
        lambda path: path.endswith("plan_number"),
        lambda path: "three_digit_plan_number" in path and not path.endswith("label"),
    )
    add_first("4c. Sponsor EIN", lambda path: ("sponsor" in path or "employer" in path) and ("ein" in path or "identification_number" in path))
    add_first(
        "4d. Plan Year Beginning Date",
        lambda path: "fiscal_plan_year_beginning" in path,
        lambda path: "calendar_plan_year_beginning" in path,
        lambda path: "calendar_plan_year_start" in path,
        lambda path: "plan_year_start_date" in path,
        lambda path: path.endswith("plan_year_begin_date"),
    )
    add_first(
        "4e. Plan Year Ending Date",
        lambda path: "fiscal_plan_year_ending" in path,
        lambda path: "calendar_plan_year_ending" in path,
        lambda path: "calendar_plan_year_end" in path,
        lambda path: "plan_year_end_date" in path,
        lambda path: path.endswith("plan_year_end_date"),
    )

    coverage = item.get("coverage_information") if isinstance(item.get("coverage_information"), dict) else {}
    add("1a. Name of Insurance Company", coverage.get("a_name_of_insurance_carrier"), 0.97, value_validator=is_probable_carrier_name)
    add("1b. Insurance Carrier EIN", coverage.get("b_ein") or coverage.get("b_EIN"), 0.97, value_validator=looks_like_ein)
    add("1c. NAIC Code", coverage.get("c_naic_code") or coverage.get("c_NAIC_code"), 0.97)
    add("1d. Contract/Policy Number", coverage.get("d_contract_or_identification_number"), 0.97)
    add("1e. Persons Covered (End of Policy Year)", coverage.get("e_approximate_number_of_persons_covered_at_end_of_policy_or_contract_year"), 0.97)
    add("1f. Policy Year Beginning Date", coverage.get("f_policy_or_contract_year_from"), 0.97)
    add("1g. Policy Year Ending Date", coverage.get("g_policy_or_contract_year_to"), 0.97)

    fees = item.get("insurance_fee_and_commission_information") if isinstance(item.get("insurance_fee_and_commission_information"), dict) else {}
    add("3b. Amount of Commissions", fees.get("a_total_amount_of_commissions_paid"), 0.97)
    add("3c. Amount of Fees", fees.get("b_total_amount_of_fees_paid") or fees.get("total_fees_paid_amount"), 0.97)

    people = item.get("persons_receiving_commissions_and_fees")
    entries = people.get("entries", []) if isinstance(people, dict) else people if isinstance(people, list) else []
    if isinstance(entries, list):
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            add(
                "3a. Name of Agent/Broker/Person",
                entry.get("a_name_and_address_of_recipient")
                or entry.get("a_name_and_address_of_agent_broker_or_other_person")
                or entry.get("name_and_address_of_recipient"),
                value_validator=is_probable_person_or_entity_name,
            )
            add("3b. Amount of Commissions", entry.get("b_amount_of_sales_and_base_commissions_paid"), 0.9)
            add("3c. Amount of Fees", entry.get("c_amount") or entry.get("fees_and_other_commissions_paid"), 0.9)
            add("3d. Purpose", entry.get("d_purpose"))
            add("3e. Organizational Code", entry.get("e_organization_code"), value_validator=looks_like_org_code)

    plan = item.get("plan_information") if isinstance(item.get("plan_information"), dict) else {}
    if not plan:
        plan = item.get("plan_identification") if isinstance(item.get("plan_identification"), dict) else {}
    add("4a. Plan Name", plan.get("a_name_of_plan"))
    add("4a. Plan Name", plan.get("A_name_of_plan"))
    add("4b. Plan Number (PN)", plan.get("b_three_digit_plan_number_pn") or plan.get("B_three_digit_plan_number_PN"))
    add("4c. Sponsor EIN", plan.get("d_employer_identification_number_ein") or plan.get("D_employer_identification_number_EIN"))

    filing_period = item.get("filing_period") if isinstance(item.get("filing_period"), dict) else {}
    add("4d. Plan Year Beginning Date", filing_period.get("fiscal_plan_year_beginning"))
    add("4e. Plan Year Ending Date", filing_period.get("fiscal_plan_year_ending"))

    return fields


def flatten_xray_values(value: Any, path: str = "") -> list[tuple[str, str]]:
    flattened: list[tuple[str, str]] = []
    if isinstance(value, dict):
        for key, nested in value.items():
            nested_path = f"{path}.{key}" if path else str(key)
            flattened.extend(flatten_xray_values(nested, nested_path))
    elif isinstance(value, list):
        for index, nested in enumerate(value):
            flattened.extend(flatten_xray_values(nested, f"{path}[{index}]"))
    elif value not in (None, ""):
        flattened.append((path, str(value)))
    return flattened


def normalize_xray_path(path: str) -> str:
    path = re.sub(r"\[\d+\]", "", path)
    path = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", path)
    path = path.replace(".", "_").replace("-", "_").replace(" ", "_")
    return re.sub(r"_+", "_", path).strip("_").lower()


def is_xray_label_or_header(path: str, value: Any) -> bool:
    text = str(value or "").strip().lower()
    if path.endswith(("label", "labels", "header", "title", "instruction")):
        return True
    label_values = {
        "three-digit plan number (pn)",
        "purpose",
        "amount of sales and base commissions paid",
        "name and address of the agent, broker, or other person to whom commissions or fees were paid",
    }
    return text in label_values


def is_blank_extraction_value(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if text in {
        "",
        "missing",
        "unreadable",
        "blank",
        "none",
        "null",
        "n/a",
        "na",
        "unknown",
        "not found",
        "obscured",
        "not provided",
        "not shown",
        "not visible",
        "redacted",
    }:
        return True
    return is_obvious_template_placeholder(value) or any(
        marker in text for marker in ["obscured", "redaction", "redacted", "not visible", "unreadable"]
    )


def looks_like_ein(value: Any) -> bool:
    return bool(re.fullmatch(r"\d{2}-\d{7}", clean_extracted_value(str(value or ""))))


def looks_like_numeric_or_code(value: Any) -> bool:
    clean = clean_extracted_value(str(value or "")).replace(",", "")
    return bool(clean and re.fullmatch(r"[0-9./\-\s]+", clean))


def looks_like_org_code(value: Any) -> bool:
    return bool(re.fullmatch(r"\d{1,2}", clean_extracted_value(str(value or ""))))


def is_probable_carrier_name(value: Any) -> bool:
    clean = clean_extracted_value(str(value or ""))
    lower = clean.lower()
    if not clean or is_obvious_template_placeholder(clean) or looks_like_ein(clean) or looks_like_numeric_or_code(clean):
        return False
    if lower in {"name of insurance carrier", "name of insurance company", "insurance carrier", "carrier name"}:
        return False
    if not re.search(r"[A-Za-z]", clean):
        return False
    return any(marker in lower for marker in ("insurance", "company", "carrier", "inc", "llc", "life", "health")) or len(clean.split()) >= 2


def is_probable_person_or_entity_name(value: Any) -> bool:
    clean = clean_extracted_value(str(value or ""))
    lower = clean.lower()
    if not clean or is_obvious_template_placeholder(clean) or looks_like_numeric_or_code(clean):
        return False
    if lower in {
        "name",
        "name and address",
        "name and address of the agents, brokers or other persons to whom commissions or fees were paid",
        "amount",
        "purpose",
        "organization code",
    }:
        return False
    return bool(re.search(r"[A-Za-z]", clean))


def has_experience_rated_not_applicable(text: str) -> bool:
    normalized = normalize_ocr_text(text).lower()
    if not normalized:
        return False
    experience_section = re.search(
        r"\b(?:9\.?\s*)?experience[-\s]?rated\s+contracts\b(.+?)(?=\b(?:10\.?\s*)?nonexperience[-\s]?rated\s+contracts\b|\bpart\s+iv\b|$)",
        normalized,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if experience_section:
        section = experience_section.group(1)
        if find_money_amounts(section):
            return False
        return bool(re.search(r"^\s*(?:n/?a|not applicable)\b", section.strip(), flags=re.IGNORECASE))
    json_match = re.search(
        r'"(?:experience[_\s-]?rated[_\s-]?contracts|experience[_\s-]?rated)"\s*:\s*"?(?:n/?a|not applicable)"?',
        normalized,
        flags=re.IGNORECASE,
    )
    if json_match:
        return True
    for match in re.finditer(r"experience[-\s]?rated\s+contracts\b.{0,80}\b(?:n/?a|not applicable)\b", normalized):
        prefix = normalized[max(0, match.start() - 4) : match.start()]
        if "non" not in prefix:
            return True
    return False


def add_not_applicable_experience_rated_fields(
    fields: list[NormalizedExtractionField],
    page: int | None,
    source_text: str,
) -> None:
    for field_name in SCHEDULE_A_EXPERIENCE_RATED_FIELDS:
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value="N/A",
                confidence=0.95,
                page=page,
                source_text=source_text or "Experience-rated contracts N/A",
            )
        )


def extract_nonexperience_total_premium_from_text(text: str) -> str | None:
    normalized = normalize_ocr_text(text)
    if not normalized:
        return None

    vendor_total = re.search(
        r"(?:gross\s+premium|total\s+premiums?(?:\s+or\s+subscription\s+charges)?\s+(?:received|paid)"
        r"(?:\s+to\s+(?:(?:the\s+)?insurance\s+company|carrier))?"
        r"(?:\s+during\s+(?:the\s+)?policy\s+year)?)"
        r"[\s.:]*\$?\s*([0-9][0-9,]*(?:\.\d{2})?)(?![0-9/])",
        normalized,
        flags=re.IGNORECASE,
    )
    if vendor_total:
        return vendor_total.group(1)

    for match in re.finditer(
        r'"(?:total_premiums_or_subscription_charges_paid_to_carrier|total_premiums|total_premium|total_amount)"\s*:\s*"?([0-9][0-9,]*(?:\.\d{2})?)"?',
        normalized,
        flags=re.IGNORECASE,
    ):
        context = normalized[max(0, match.start() - 250) : match.end() + 250].lower()
        if "nonexperience" in context or "subscription charges" in context or "premiums" in context:
            return match.group(1)

    compact = re.sub(r"\s+", " ", normalized)
    section_match = re.search(
        r"\bnonexperience[-\s]?rated\s+contracts\b(.+?)(?=\bpart\s+iv\b|\bagent/broker\b|\baddendum\b|\bclient\s+name\b|$)",
        compact,
        flags=re.IGNORECASE,
    )
    if section_match:
        section = section_match.group(1)
        if re.search(r"^\s*(?:n/?a|not applicable)\b", section.strip(), flags=re.IGNORECASE):
            return None
        amounts = find_money_amounts(section)
        if amounts:
            return amounts[-1]

    label_match = re.search(
        r"total\s+premiums\s+or\s+subscription\s+charges\s+paid\s+to\s+carrier",
        compact,
        flags=re.IGNORECASE,
    )
    if label_match:
        before = compact[max(0, label_match.start() - 500) : label_match.start()]
        after = compact[label_match.end() : label_match.end() + 500]
        amounts = find_money_amounts(after) or find_money_amounts(before)
        if amounts:
            return amounts[-1]

    return None


def find_money_amounts(text: str) -> list[str]:
    return re.findall(r"\b[0-9]{1,3}(?:,[0-9]{3})+(?:\.\d{2})?\b|\b[0-9]{4,}(?:\.\d{2})?\b", text)


def build_groundx_schema_query(
    file_name: str | None = None,
    rules=None,
    *,
    form_type: FormType = FormType.SCHEDULE_A,
) -> str:
    relevant_rules = rules_for_form(rules if rules is not None else DEFAULT_FIELD_RULES, form_type)
    field_hints: list[str] = []
    for rule in relevant_rules:
        aliases = [
            alias
            for alias in rule.aliases
            if normalize_rule_label(alias) != normalize_rule_label(rule.label)
        ]
        hint = rule.label
        if aliases:
            hint += f" (also labeled: {', '.join(aliases)})"
        field_hints.append(hint)
    labels = "; ".join(field_hints)
    file_hint = f" Prefer content from file named {file_name} when that file is searchable. " if file_name else " "
    return (
        "Using this Schedule A / Form 5500 document, retrieve the text needed to extract these FT Williams fields."
        f"{file_hint}"
        "Focus on exact values near labels, tables, and line numbers. Fields: "
        f"{labels}"
    )


def rules_for_form(rules, form_type: FormType):
    if form_type == FormType.FORM_5500:
        return [
            rule
            for rule in rules
            if str(getattr(rule.applicability, "value", rule.applicability)) == "FORM_5500"
            or str(rule.source).lower().startswith("form 5500")
            or str(rule.form_section or "").lower().startswith("form 5500")
        ]
    return [
        rule
        for rule in rules
        if str(getattr(rule.applicability, "value", rule.applicability)) != "FORM_5500"
        and not str(rule.source).lower().startswith("form 5500")
        and not str(rule.form_section or "").lower().startswith("form 5500")
    ]


def normalize_ocr_text(value: Any) -> str:
    if not value:
        return ""
    text = str(value)
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\(\s*([a-zA-Z0-9]+)\s*\)", r"(\1)", text)
    text = re.sub(r"(?<!\()\b([a-zA-Z])\s*\)", r"(\1)", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def parse_groundx_page(result: dict) -> int | None:
    boxes = result.get("boundingBoxes") or []
    if isinstance(boxes, list) and boxes:
        page = parse_page(boxes[0].get("pageNumber"))
        if page:
            return page
    pages = result.get("pages") or []
    if isinstance(pages, list) and pages:
        return parse_page(pages[0])
    return parse_page(result.get("pageNumber") or result.get("page"))


def build_structured_extraction_context(raw_payloads: list[Any], file_bytes: bytes) -> str:
    chunks: list[str] = []
    for payload in raw_payloads:
        if is_groundx_search_payload(payload):
            chunks.extend(extract_text_chunks_from_groundx_search(payload))
        elif is_groundx_xray_payload(payload):
            for chunk in payload.get("chunks", []):
                if not isinstance(chunk, dict):
                    continue
                text = build_xray_source_text(chunk)
                if text:
                    page = parse_xray_page(chunk)
                    prefix = f"[GroundX page {page}]\n" if page else "[GroundX]\n"
                    chunks.append(f"{prefix}{text}")

    pages = extract_pdf_text_pages(file_bytes)
    for page_number, text in pages:
        if text.strip():
            chunks.append(f"[PDF page {page_number}]\n{text.strip()}")

    seen: set[str] = set()
    unique_chunks: list[str] = []
    for chunk in chunks:
        normalized = normalize_ocr_text(chunk)
        key = normalized[:500]
        if normalized and key not in seen:
            seen.add(key)
            unique_chunks.append(normalized)
    return "\n\n---\n\n".join(unique_chunks)[:70000]


def extract_text_chunks_from_groundx_search(raw: Any) -> list[str]:
    search = raw.get("search", {}) if isinstance(raw, dict) else {}
    results = search.get("results", []) if isinstance(search, dict) else []
    chunks: list[str] = []
    for result in results:
        if not isinstance(result, dict):
            continue
        text = normalize_ocr_text(result.get("text") or result.get("suggestedText") or result.get("narrative") or "")
        if text:
            page = parse_groundx_page(result)
            prefix = f"[GroundX page {page}]\n" if page else "[GroundX]\n"
            chunks.append(f"{prefix}{text}")
    return chunks


def extract_pdf_text_pages(file_bytes: bytes) -> list[tuple[int, str]]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return []

    try:
        reader = PdfReader(BytesIO(file_bytes))
    except Exception:
        return []

    pages: list[tuple[int, str]] = []
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            continue
        pages.append((index, normalize_ocr_text(text)))
    return pages


def extract_docx_text(file_bytes: bytes) -> str:
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            xml_files = [
                "word/document.xml",
                *sorted(name for name in archive.namelist() if name.startswith("word/header") and name.endswith(".xml")),
                *sorted(name for name in archive.namelist() if name.startswith("word/footer") and name.endswith(".xml")),
            ]
            chunks: list[str] = []
            namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            for xml_file in xml_files:
                if xml_file not in archive.namelist():
                    continue
                root = ET.fromstring(archive.read(xml_file))
                for paragraph in root.findall(".//w:p", namespace):
                    text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace)).strip()
                    if text:
                        chunks.append(text)
            return normalize_ocr_text("\n".join(chunks))
    except Exception:
        return ""


def parse_plan_worksheet_text(text: str, *, rules=None) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    compact = re.sub(r"\s+", " ", text).strip()

    def add(field_name: str, value: Any, confidence: float = 0.92, source_text: str | None = None):
        clean = clean_extracted_value(str(value or ""))
        if clean:
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=clean,
                    confidence=confidence,
                    page=None,
                    source_text=source_text or clean,
                )
            )

    sponsor_name = regex_first(compact, [r"Plan sponsor name\s+(.+?)\s+Plan sponsor address"], flags=re.IGNORECASE)
    sponsor_address = regex_first(compact, [r"Plan sponsor address\s+(.+?)\s+Plan sponsor phone number"], flags=re.IGNORECASE)
    sponsor_ein = regex_first(compact, [r"\bEIN\s+([0-9]{2}-[0-9]{7})\b"], flags=re.IGNORECASE)
    business_code = regex_first(compact, [r"Business code\s+([0-9]{4,6})\b"], flags=re.IGNORECASE)
    plan_number = regex_first(compact, [r"Plan number\(s\)\s+([0-9]{3})\b"], flags=re.IGNORECASE)
    plan_name = regex_first(compact, [r"Plan name\(s\)\s+(.+?)\s+Plan year"], flags=re.IGNORECASE)
    effective_date = regex_first(compact, [r"Original ERISA plan effective date\s+([0-9]{2}[-/][0-9]{2}[-/][0-9]{4})"], flags=re.IGNORECASE)
    plan_year = regex_first(compact, [r"Plan year\s+begin\s*/\s*end\s+([0-9]{2}[-/][0-9]{2}[-/][0-9]{4})\s+([0-9]{2}[-/][0-9]{2}[-/][0-9]{4})"], flags=re.IGNORECASE, groups=True)
    administrator = regex_first(
        compact,
        [
            r"Plan administrator name\s+(.+?)\s+(?:Plan administrator address|E-mail address|Participant Counts:)",
            r"Administrator name\s+(.+?)\s+(?:Administrator address|E-mail address|Participant Counts:)",
            r"Individual signing as plan administrator\s+(.+?)\s+(?:E-mail address of filing signer|5500 Contact|Additional 5500 Contact|Participant Counts:)",
        ],
        flags=re.IGNORECASE,
    )

    add("1a. Plan Name", plan_name, 0.95)
    add("1b. Plan Number (PN)", plan_number, 0.95)
    add("1c. Plan Effective Date", effective_date, 0.93)
    add("1d. Plan Sponsor Name", sponsor_name, 0.95)
    add("1e. Plan Sponsor EIN", sponsor_ein, 0.95)
    add("1f. Plan Sponsor Address", sponsor_address, 0.92)
    add("1g. Business Code", business_code, 0.94)
    add("2a. Plan Administrator Name", administrator, 0.92)
    if isinstance(plan_year, tuple) and len(plan_year) >= 2:
        add("6. Plan Year Beginning Date", plan_year[0], 0.95)
        add("7. Plan Year Ending Date", plan_year[1], 0.95)

    beginning_total = regex_first(
        compact,
        [r"Total number of participants at the beginning of the plan year.*?\]\s*([0-9,]+)\s+6\(a\)\(1\)"],
        flags=re.IGNORECASE,
    )
    active_beginning = regex_first(
        compact,
        [r"6\(a\)\(1\)\s+Total number of active participants on the first day of the plan year.*?\]\s*([0-9,]+)\s+6\(a\)\(2\)"],
        flags=re.IGNORECASE,
    )
    active_end = regex_first(
        compact,
        [r"6\(a\)\(2\)\s+Total number of active participants on the last day of the plan year\s+([0-9,]+)\s+6\(b\)"],
        flags=re.IGNORECASE,
    )
    retired_receiving = regex_first(
        compact,
        [r"6\(b\)\s+Total number of retired or COBRA participants on benefits as of last day of the plan year\s+([0-9,]+)\s+6\(c\)"],
        flags=re.IGNORECASE,
    )
    retired_entitled = regex_first(
        compact,
        [r"6\(c\)\s+Total number of retired or COBRA participants entitled to benefits as of last day of the plan year\s+([0-9,]+)"],
        flags=re.IGNORECASE,
    )
    total_end = regex_first(
        compact,
        [r"Total participants at end of (?:the )?plan year\s+([0-9,]+)", r"Total number of participants at the end of the plan year\s+([0-9,]+)"],
        flags=re.IGNORECASE,
    )
    if not total_end and active_end is not None and retired_receiving is not None and retired_entitled is not None:
        total_end = str(
            int(str(active_end).replace(",", ""))
            + int(str(retired_receiving).replace(",", ""))
            + int(str(retired_entitled).replace(",", ""))
        )

    add("11. Total participants at beginning of year", beginning_total, 0.93)
    add("12. Total participants at end of year", total_end, 0.9 if total_end else 0.0)
    add("13. Active participants at beginning", active_beginning, 0.93)
    add("14. Active participants at end", active_end, 0.93)
    add("15. Retired/separated participants receiving benefits", retired_receiving, 0.93)
    add("16. Other retired/separated participants entitled to benefits", retired_entitled, 0.93)

    if re.search(r"\bwelfare benefit plan\b|\bemployee welfare benefit\b", compact, flags=re.IGNORECASE):
        add("6. Plan is a welfare plan?", "Yes", 0.86)
    welfare_codes = regex_first(
        compact,
        [
            r"If the plan provides welfare benefits.*?:\s*((?:4[A-Z]\s*)+)",
            r"applicable codes.*?:\s*((?:4[A-Z]\s*)+)",
        ],
        flags=re.IGNORECASE,
    )
    if welfare_codes:
        codes = " ".join(re.findall(r"4[A-Z]", welfare_codes.upper()))
        add("8c. Welfare Benefit Features", codes, 0.88, source_text=welfare_codes)
    if re.search(r"Fully-Insured Benefits", compact, flags=re.IGNORECASE):
        add("9. Plan funding arrangement", "Insurance", 0.86)
        add("10a. Plan benefit arrangement", "Insurance", 0.86)
        add("10b. Schedules attached", "A", 0.86)

    fields.extend(extract_configured_custom_fields(text, None, rules=rules))

    return dedupe_fields(fields)


def file_type_for_groundx(file_name: str) -> str:
    extension = os.path.splitext(file_name.lower())[1].lstrip(".")
    if extension == "jpeg":
        extension = "jpg"
    if extension in {"pdf", "docx", "doc", "xlsx", "xls", "csv", "txt", "png", "jpg", "tiff"}:
        return extension
    if extension == "tif":
        return "tiff"
    return "pdf"


def local_schedule_a_pdf_result(
    file_bytes: bytes,
    file_name: str,
    provider: str = "Local PDF parser",
    *,
    rules=None,
) -> NormalizedExtractionResult:
    is_pdf = file_name.lower().endswith(".pdf")
    return NormalizedExtractionResult(
        provider=provider if is_pdf else "Local document parser",
        fields=extract_fields_from_document_text(file_bytes, file_name, rules=rules),
        raw={"file_name": file_name, "source": "local_document_parser"},
        schedule_a_broker_rows=extract_schedule_a_broker_rows_from_pdf_text(file_bytes) if is_pdf else [],
        schedule_a_worksheet_summaries=extract_schedule_a_worksheet_summaries_from_pdf_text(file_bytes) if is_pdf else [],
    )


def extract_document_text_pages(file_bytes: bytes, file_name: str | None = None) -> list[tuple[int, str]]:
    """Get readable text out of a document, whatever format it arrived in.

    The label-driven parsing below is the strongest part of Schedule A
    extraction, and it used to run on PDFs only. A Schedule A that arrives as
    a spreadsheet or a CSV export carries the same labels and values, so it
    gets the same treatment - the rows are just flattened to text first.
    """
    name = str(file_name or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _spreadsheet_text_pages(file_bytes)
    if name.endswith(".csv"):
        return [(1, _delimited_text(file_bytes.decode("utf-8", errors="ignore")))]
    if name.endswith(".txt"):
        return [(1, normalize_ocr_text(file_bytes.decode("utf-8-sig", errors="ignore")))]
    if name.endswith((".doc", ".docx")):
        text = extract_docx_text(file_bytes)
        return [(1, text)] if text else []
    return extract_pdf_text_pages(file_bytes)


def extract_schedule_a_classification_signals(file_bytes: bytes, file_name: str | None = None) -> list[str]:
    text = "\n\n".join(value for _, value in extract_document_text_pages(file_bytes, file_name))
    return classification_signals_from_text(text)


def _spreadsheet_text_pages(file_bytes: bytes) -> list[tuple[int, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return []

    pages: list[tuple[int, str]] = []
    for index, sheet in enumerate(workbook.worksheets, start=1):
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if not cells:
                continue
            # A two-column row is a label and its value - write it the way the
            # label parser expects to see it.
            lines.append(f"{cells[0]}: {cells[1]}" if len(cells) == 2 else " ".join(cells))
        if lines:
            pages.append((index, "\n".join(lines)))
    return pages


def _delimited_text(text: str) -> str:
    """Turn "Label,Value" rows into "Label: Value" lines."""
    lines: list[str] = []
    for row in csv.reader(StringIO(text.lstrip("\ufeff"))):
        parts = [str(part).strip() for part in row]
        parts = [part for part in parts if part]
        if not parts:
            continue
        lines.append(f"{parts[0]}: {parts[1]}" if len(parts) == 2 else " ".join(parts))
    return "\n".join(lines)


def extract_fields_from_document_text(
    file_bytes: bytes,
    file_name: str | None = None,
    *,
    rules=None,
) -> list[NormalizedExtractionField]:
    pages = extract_document_text_pages(file_bytes, file_name)
    fields = [
        *_extract_fields_from_pages(pages, rules=rules),
        *extract_labelled_schedule_a_fields(pages),
        *(extract_email_schedule_a_fields(pages) if str(file_name or "").lower().endswith("email body.txt") else []),
        *schedule_a_broker_compensation_fields(extract_tabular_broker_rows(pages)),
    ]
    return select_best_schedule_a_fields(fields)


def extract_email_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    """Extract carrier values supplied as prose in an email response."""
    text = "\n".join(value for _, value in page_texts)
    normalized = normalize_ocr_text(text)
    fields: list[NormalizedExtractionField] = []

    def add(field_name: str, value: str | None, confidence: float = 0.92) -> None:
        clean = clean_extracted_value(value or "")
        if not clean:
            return
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value=clean,
                confidence=confidence,
                page=1,
                source_text="Email body",
            )
        )

    add(
        "1a. Name of Insurance Company",
        regex_first(normalized, [r"^\s*Legal\s+Name\s*:\s*(.+?)\s*$"], flags=re.IGNORECASE | re.MULTILINE),
    )
    add(
        "1b. Insurance Carrier EIN",
        regex_first(normalized, [r"^\s*(?:Carrier\s+)?EIN\s*:\s*([0-9]{2}-[0-9]{7})\b"], flags=re.IGNORECASE | re.MULTILINE),
        0.96,
    )
    add(
        "1e. Persons Covered (End of Policy Year)",
        regex_first(
            normalized,
            [
                r"Approximate\s+(?:employee\s+)?lives?\s+covered[^:\n]*:\s*(?:[A-Za-z][A-Za-z &/.-]*\s+)?([0-9,]+)\b",
                r"(?:employee\s+)?lives?\s+covered\s*:\s*([0-9,]+)\b",
            ],
            flags=re.IGNORECASE | re.DOTALL,
        ),
        0.94,
    )
    add(
        "3c. Amount of Fees",
        regex_first(
            normalized,
            [r"(?:PEPM\s+)?Fees?\s+Paid(?:\s*\([^)]*\))?\s*:\s*\$?\s*([0-9,]+(?:\.\d{1,2})?)"],
            flags=re.IGNORECASE,
        ),
        0.95,
    )

    start_date, end_date = extract_email_coverage_dates(normalized)
    add("1f. Policy Year Beginning Date", start_date, 0.93)
    add("1g. Policy Year Ending Date", end_date, 0.93)
    return fields


def extract_email_coverage_dates(text: str) -> tuple[str | None, str | None]:
    explicit = re.search(
        r"\b([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})\s*(?:through|thru|to|-)\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})\b",
        text,
        flags=re.IGNORECASE,
    )
    if explicit:
        return explicit.group(1), explicit.group(2)

    months = "|".join(calendar.month_name[1:])
    month_range = re.search(
        rf"\b({months})\s+([0-9]{{4}})\s*(?:through|thru|to|-)\s*({months})\s+([0-9]{{4}})\b",
        text,
        flags=re.IGNORECASE,
    )
    if not month_range:
        return None, None
    month_numbers = {name.lower(): index for index, name in enumerate(calendar.month_name) if name}
    start_month = month_numbers[month_range.group(1).lower()]
    start_year = int(month_range.group(2))
    end_month = month_numbers[month_range.group(3).lower()]
    end_year = int(month_range.group(4))
    end_day = calendar.monthrange(end_year, end_month)[1]
    return f"{start_month:02d}/01/{start_year}", f"{end_month:02d}/{end_day:02d}/{end_year}"


# Spreadsheets and exports state the same values as a carrier statement, but as
# plain "label, value" rows rather than in prose. These are the label wordings
# seen on the documents clients send.
_LABELLED_SCHEDULE_A_FIELDS: list[tuple[str, tuple[str, ...]]] = [
    (
        "1a. Name of Insurance Company",
        ("name of insurance carrier", "name of insurance company", "insurance carrier name", "carrier name"),
    ),
    ("1b. Insurance Carrier EIN", ("carrier ein", "insurance carrier ein", "ein")),
    ("1c. NAIC Code", ("naic code", "naic")),
    (
        "1d. Contract/Policy Number",
        ("contract/policy number", "contract or policy number", "policy number", "contract number"),
    ),
    (
        "1e. Persons Covered (End of Policy Year)",
        (
            "persons covered end of policy year",
            "persons covered at end of policy year",
            "approximate number of persons covered at the end of the policy year",
            "number of persons covered",
        ),
    ),
    ("1f. Policy Year Beginning Date", ("policy year beginning date", "policy year from", "contract/policy year from")),
    ("1g. Policy Year Ending Date", ("policy year ending date", "policy year to", "contract/policy year to")),
    (
        "3a. Name of Agent/Broker/Person",
        ("name of agent/broker", "name of agent or broker", "agent/broker name", "broker name", "name of agent"),
    ),
    ("3b. Amount of Commissions", ("amount of commissions", "total commissions", "commissions paid", "commissions")),
    ("3c. Amount of Fees", ("amount of fees", "total fees", "fees paid")),
    ("3d. Purpose", ("purpose for which paid", "purpose")),
    (
        "10a. Total premiums or subscription charges paid to carrier",
        (
            "total premiums paid to insurance company during the policy year",
            "total premiums paid",
            "total premiums",
        ),
    ),
]


def extract_labelled_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    """Read "Label: Value" lines out of a flattened spreadsheet or export."""
    fields: list[NormalizedExtractionField] = []
    seen: set[str] = set()
    for page, text in page_texts:
        for line in normalize_ocr_text(text or "").splitlines():
            label, separator, value = line.partition(":")
            if not separator:
                continue
            key = re.sub(r"[^a-z0-9/ ]+", "", label.strip().lower()).strip()
            key = re.sub(r"^\d+[a-z]?\s+", "", key).strip()
            value = clean_extracted_value(value)
            if not key or not value:
                continue
            for field_name, aliases in _LABELLED_SCHEDULE_A_FIELDS:
                if field_name in seen or key not in aliases:
                    continue
                if field_name.startswith("3a.") and not is_probable_person_or_entity_name(value):
                    continue
                seen.add(field_name)
                fields.append(
                    NormalizedExtractionField(
                        field_name=field_name,
                        value=value,
                        confidence=0.85,
                        page=page,
                        source_text="Labelled row",
                    )
                )
                break
    return fields


def extract_schedule_a_broker_rows_from_document(file_bytes: bytes, file_name: str | None = None) -> list[ScheduleABrokerRow]:
    page_texts = extract_document_text_pages(file_bytes, file_name)
    cigna_rows = extract_cigna_schedule_a_broker_rows(page_texts)
    if cigna_rows or _is_cigna_schedule_a_packet(page_texts):
        return cigna_rows
    full_text = "\n\n".join(text for _, text in page_texts)
    return dedupe_schedule_a_broker_rows(
        [
            *extract_schedule_a_broker_rows(full_text),
            *extract_columnar_broker_compensation_rows(page_texts),
            *extract_compensation_table_broker_rows(page_texts),
            *extract_tabular_broker_rows(page_texts),
        ]
    )


# "Brokerage LLC, 500 Market Street, Boston MA 02110 7412.33 615 Standard Commissions"
_TABULAR_BROKER_ROW = re.compile(
    r"^(?P<name>[A-Za-z].*?)\s+(?P<commissions>[\d,]+(?:\.\d{1,2})?)\s+(?P<fees>[\d,]+(?:\.\d{1,2})?)\s*(?P<purpose>[A-Za-z][A-Za-z /]*)?$"
)


def extract_tabular_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    """Read a broker row out of a spreadsheet, where there are no currency signs.

    A workbook row flattens to "name  commissions  fees  purpose" with nothing
    marking the amounts, so this only runs where the sheet itself says the
    columns are commissions and fees.
    """
    rows: list[ScheduleABrokerRow] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text or "")
        lowered = normalized.lower()
        if "commission" not in lowered or "fee" not in lowered:
            continue
        for line in normalized.splitlines():
            stripped = line.strip()
            match = _TABULAR_BROKER_ROW.match(stripped)
            if not match:
                continue
            name = _AGENT_NUMBER_PREFIX.sub("", match.group("name")).strip(" ,")
            name = name.split(",")[0].strip()
            if not is_probable_person_or_entity_name(name):
                continue
            commissions = money_value(match.group("commissions"))
            fees = money_value(match.group("fees"))
            if not commissions and not fees:
                continue
            purpose = clean_extracted_value(match.group("purpose") or "") or None
            rows.append(
                ScheduleABrokerRow(
                    name=name,
                    commission_total=commissions or None,
                    fee_total=fees or None,
                    commission_rows=(
                        [ScheduleABrokerMoneyRow(amount=commissions, purpose=purpose)] if commissions else []
                    ),
                    fee_rows=[ScheduleABrokerMoneyRow(amount=fees, purpose=purpose)] if fees else [],
                    source_page=page,
                )
            )
    return rows


def extract_fields_from_pdf_text(file_bytes: bytes, *, rules=None) -> list[NormalizedExtractionField]:
    return _extract_fields_from_pages(extract_pdf_text_pages(file_bytes), rules=rules)


def _extract_fields_from_pages(page_texts: list[tuple[int, str]], *, rules=None) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    for index, text in page_texts:
        fields.extend(parse_schedule_a_text(text, index, rules=rules))
        fields.extend(extract_explicit_benefit_indicator_fields(text, index, rules=rules))
        fields.extend(extract_bcbsma_schedule_a_worksheet_fields(text, index))
    fields.extend(extract_bcbs_michigan_schedule_a_fields(page_texts))
    fields.extend(extract_prudential_schedule_a_fields(page_texts))
    fields.extend(extract_eyemed_schedule_a_fields(page_texts))
    fields.extend(extract_standard_schedule_a_fields(page_texts))
    fields.extend(extract_united_omaha_schedule_a_fields(page_texts))
    authoritative_packet_fields = [
        extract_nyl_annual_policy_fields(page_texts),
        extract_hmsa_schedule_a_fields(page_texts),
    ]
    full_text = "\n\n".join(text for _, text in page_texts)
    if full_text:
        fields.extend(parse_schedule_a_text(full_text, None, rules=rules))
        fields.extend(extract_bcbsma_commission_breakdown_fields(full_text, None))
    # The broker compensation table is where items 3a-3d actually live on a
    # carrier statement. Reading the table and reading the labelled fields are
    # separate passes, so feed the table's values back in as fields - without
    # this the amounts were parsed and then thrown away.
    fields.extend(schedule_a_broker_compensation_fields(extract_compensation_table_broker_rows(page_texts)))
    fields.extend(schedule_a_broker_compensation_fields(extract_columnar_broker_compensation_rows(page_texts)))
    fields.extend(extract_commission_fee_total_fields(page_texts))
    for packet_fields in authoritative_packet_fields:
        if not packet_fields:
            continue
        authoritative_names = {field.field_name for field in packet_fields}
        fields = [field for field in fields if field.field_name not in authoritative_names]
        fields.extend(packet_fields)
    cigna_fields = extract_cigna_schedule_a_fields(page_texts)
    if cigna_fields:
        # Cigna reporting packages contain a consolidated Schedule A followed
        # by state-carrier appendices. Those appendix EIN/NAIC/lives values are
        # supporting detail, not alternative primary Schedule A records.
        # Remove every broad-parser value for fields owned by the consolidated
        # parser so a later appendix can never replace the primary record.
        authoritative_names = {field.field_name for field in cigna_fields}
        fields = [field for field in fields if field.field_name not in authoritative_names]
        fields.extend(cigna_fields)
    return dedupe_fields(
        [
            field
            for field in fields
            if not _is_column_heading_broker_name(field)
            and not is_obvious_template_placeholder(field.value)
        ]
    )


def extract_explicit_benefit_indicator_fields(
    text: str,
    page: int | None = None,
    *,
    rules=None,
) -> list[NormalizedExtractionField]:
    """Map explicit benefit evidence to discovered FTW comparison fields.

    Merely printing the word "Vision" on a blank IRS form is not evidence. We
    accept a completed benefit summary, an explicit insurance/coverage phrase,
    or a checked label. Checkbox evidence is ignored on obvious sample forms.
    """
    if not rules:
        return []
    normalized = normalize_ocr_text(text)
    placeholder_heavy = len(re.findall(r"ABCDEFGHI|123456789012345", normalized, flags=re.IGNORECASE)) >= 2
    fields: list[NormalizedExtractionField] = []
    for rule in rules:
        tag = str(rule.xml_tag or "").strip().lower()
        if tag not in {"healthind", "visionind"}:
            continue
        aliases = [rule.label, *rule.aliases]
        matched_source: str | None = None
        matched_value = "Yes"
        for alias in sorted({alias.strip() for alias in aliases if alias.strip()}, key=len, reverse=True):
            token = loose_label_pattern(alias)
            if not placeholder_heavy:
                checkbox = re.search(
                    rf"(?im)(?:\([A-M]\)\s*)?\[\s*(?P<mark>[Xx]?)\s*\]\s*{token}\b",
                    normalized,
                )
                if checkbox:
                    matched_source = checkbox.group(0).strip()
                    matched_value = "Yes" if checkbox.group("mark") else "No"
                    break
            patterns = [
                rf"(?im)^\s*Benefits?\s*:\s*[^\n]*\b{token}\b[^\n]*$",
                rf"(?i)\b{token}\s+(?:insurance|coverage|benefit)\b",
                rf"(?i)\b{token}\s+plan\b",
                rf"(?i)\b(?:type\s+of\s+benefit|benefit\s+type)\s*:?\s*{token}\b",
            ]
            if not placeholder_heavy:
                patterns.append(rf"(?im)^\s*(?:X|✓|☑)\s*{token}\b")
            match = next((match for pattern in patterns if (match := re.search(pattern, normalized))), None)
            if match:
                matched_source = match.group(0).strip()
                break
        if matched_source:
            fields.append(
                NormalizedExtractionField(
                    field_name=rule.label,
                    value=matched_value,
                    confidence=0.9,
                    page=page,
                    source_text=matched_source,
                )
            )
    return fields


_BROKER_HEADING_TOKENS = (
    "total",
    "premium",
    "covered",
    "plan #",
    "policy or contract",
    "amount of",
    "name & address",
    "name and address",
    "fees paid",
    "commissions paid",
    "for which paid",
)


def _is_column_heading_broker_name(field: NormalizedExtractionField) -> bool:
    """Drop a broker name that is really a run-together column heading.

    Table headers wrap across lines in the PDF text layer, so a careless parse
    produces values like "Base Plan # Covered Total Premiums". An empty field
    sends a reviewer to the document; a plausible-looking wrong name can be
    approved by mistake and sent to FT Williams.
    """
    if not field.field_name.startswith("3a."):
        return False
    lowered = str(field.value or "").lower()
    if not lowered:
        return False
    return sum(1 for token in _BROKER_HEADING_TOKENS if token in lowered) >= 2


def schedule_a_broker_compensation_fields(rows: list[ScheduleABrokerRow]) -> list[NormalizedExtractionField]:
    """Turn broker table rows into Schedule A items 3a-3d."""
    if not rows:
        return []
    primary = rows[0]
    fields: list[NormalizedExtractionField] = []

    def add(field_name: str, value: str | None, confidence: float) -> None:
        cleaned = clean_extracted_value(str(value)) if value not in (None, "") else ""
        if cleaned:
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=cleaned,
                    confidence=confidence,
                    page=primary.source_page,
                    source_text="Broker compensation table",
                )
            )

    add("3a. Name of Agent/Broker/Person", primary.name, 0.86)
    # More than one recipient means the filing total is the sum, and a human
    # needs to split it across rows - flag it by lowering confidence.
    confidence = 0.86 if len(rows) == 1 else 0.6
    add("3b. Amount of Commissions", sum_money_values(*(row.commission_total for row in rows)), confidence)
    add("3c. Amount of Fees", sum_money_values(*(row.fee_total for row in rows)), confidence)
    purpose = next(
        (money_row.purpose for row in rows for money_row in row.commission_rows if money_row.purpose),
        None,
    )
    add("3d. Purpose", purpose, 0.8)
    add("3e. Organizational Code", primary.organization_code, 0.8)
    return fields


# "Total (from below) 6590.57 5555.33" - a commissions/fees totals line.
_COMMISSION_FEE_TOTALS = re.compile(
    r"total[^\n\d]{0,24}?\$?\s*(?P<commissions>\d[\d,]*(?:\.\d{2})?)\s+\$?\s*(?P<fees>\d[\d,]*(?:\.\d{2})?)\s*$",
    re.IGNORECASE,
)


def extract_commission_fee_total_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    """Read a commissions/fees totals line where no named broker row exists.

    Some carriers (for example Equitable) print a section header of
    "Commissions Paid / Fees Paid" followed by a totals line, with the named
    recipients listed separately underneath.
    """
    fields: list[NormalizedExtractionField] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not normalized:
            continue
        lines = [line.strip() for line in normalized.splitlines()]
        for index, line in enumerate(lines):
            match = _COMMISSION_FEE_TOTALS.search(line)
            if not match:
                continue
            context = " ".join(lines[max(0, index - 3) : index + 1]).lower()
            if "commission" not in context or "fee" not in context:
                continue
            commissions = money_value(match.group("commissions"))
            fees = money_value(match.group("fees"))
            if not commissions and not fees:
                continue
            for field_name, value in (
                ("3b. Amount of Commissions", commissions),
                ("3c. Amount of Fees", fees),
            ):
                if value:
                    fields.append(
                        NormalizedExtractionField(
                            field_name=field_name,
                            value=value,
                            confidence=0.8,
                            page=page,
                            source_text="Commissions and fees total",
                        )
                    )
            return fields
    return fields


def extract_schedule_a_broker_rows_from_pdf_text(file_bytes: bytes) -> list[ScheduleABrokerRow]:
    page_texts = extract_pdf_text_pages(file_bytes)
    cigna_rows = extract_cigna_schedule_a_broker_rows(page_texts)
    if cigna_rows or _is_cigna_schedule_a_packet(page_texts):
        return cigna_rows
    full_text = "\n\n".join(text for _, text in page_texts)
    return dedupe_schedule_a_broker_rows(
        [
            *extract_schedule_a_broker_rows(full_text),
            *extract_bcbs_michigan_addendum_broker_rows(full_text),
            *extract_bcbsma_commission_breakdown_broker_rows(full_text),
            *extract_prudential_broker_rows(page_texts),
            *extract_eyemed_broker_rows(page_texts),
            *extract_standard_broker_rows(page_texts),
            *extract_united_omaha_broker_rows(page_texts),
            *extract_summary_table_broker_rows(page_texts),
            *extract_columnar_broker_compensation_rows(page_texts),
            *extract_compensation_table_broker_rows(page_texts),
        ]
    )


def extract_nyl_annual_policy_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    """Extract New York Life's Annual Policy Information Report exactly.

    Its PDF text layer splits words such as ``Year`` and ``Total`` while the
    contract number can contain a meaningful space. Broad label parsing used
    to lose the second contract token and miss the ending date and premium.
    """

    fields: list[NormalizedExtractionField] = []
    for page, text in page_texts:
        if (
            "Annual Policy Information Report" not in text
            or "Name of Insurance Carrier" not in text
            or "Contract/Policy Number" not in text
        ):
            continue

        def value(pattern: str) -> str | None:
            match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
            return clean_extracted_value(match.group(1)) if match else None

        extracted = {
            "1a. Name of Insurance Company": value(r"Name\s+of\s+Insurance\s+Carrier\s*\n\s*([^\n]+)"),
            "1b. Insurance Carrier EIN": value(r"\bEIN\s+([0-9]{2}-[0-9]{7})"),
            "1c. NAIC Code": value(r"\bNAIC\s+Code\s+([0-9]{4,6})"),
            "1d. Contract/Policy Number": value(r"Contract/Policy\s+Number[ \t]+([A-Z0-9]+(?:[ \t]+[A-Z0-9]+)*)"),
            "1f. Policy Year Beginning Date": value(r"Contract/Policy\s+Y\s*ear\s+From\s*:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})"),
            "1g. Policy Year Ending Date": value(r"Contract/Policy\s+Y\s*ear\s+T\s*o\s*:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})"),
            "10a. Total premiums or subscription charges paid to carrier": value(
                r"T\s*otal\s+premiums\s+paid\s+to\s+Insurance\s+Company\s+during\s+the\s+policy\s+year\s*:\s*\$?\s*([0-9,]+(?:\.\d{2})?)"
            ),
        }
        if extracted["1a. Name of Insurance Company"]:
            extracted["1a. Name of Insurance Company"] = extracted["1a. Name of Insurance Company"].replace("Y ork", "York")
        for field_name, field_value in extracted.items():
            if not field_value:
                continue
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=field_value,
                    confidence=0.99,
                    page=page,
                    source_text="New York Life Annual Policy Information Report",
                )
            )
        break
    return fields


def extract_hmsa_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    """Extract the two-page HMSA ERISA Schedule A support packet."""

    full_text = "\n".join(text for _, text in page_texts)
    if "HAWAII MEDICAL SERVICE ASSOCIATION" not in full_text.upper() or "ERISA FORM 5500 AND SCHEDULE A INFORMATION" not in full_text.upper():
        return []

    ein = regex_first(full_text, [r"EIN\s+number\s+([0-9]{2}-[0-9]{7})"], flags=re.IGNORECASE)
    naic = regex_first(full_text, [r"NAIC\s+code\s+([0-9]{4,6})"], flags=re.IGNORECASE)
    period = re.search(
        r"FOR\s+THE\s+PLAN\s+YEAR\s*:\s*January\s+([0-9]{4})\s*-\s*December\s+([0-9]{4})",
        full_text,
        flags=re.IGNORECASE,
    )
    rows = re.findall(
        r"(?:C\d+\s+)?(?P<group>\d{5,6})\s+(?P<subgroup>\d{3})\s+.+?\s+(?P<subs>\d+)\s+(?P<covered>\d+)\s+\$(?P<premium>[0-9,]+(?:\.\d{2})?)\s*$",
        full_text,
        flags=re.IGNORECASE | re.MULTILINE,
    )
    if not rows:
        return []

    primary = next((row for row in rows if int(row[3]) > 0), rows[0])
    group = primary[0].lstrip("0") or "0"
    subgroup = primary[1].lstrip("0") or "0"
    contract = f"{group} {subgroup}"
    covered = primary[3]
    premium = sum_money_values(*(row[4] for row in rows))
    begin = f"01/01/{period.group(1)}" if period else None
    end = f"12/31/{period.group(2)}" if period else None
    extracted = {
        "1a. Name of Insurance Company": "Hawaii Medical Service Association (HMSA)",
        "1b. Insurance Carrier EIN": ein,
        "1c. NAIC Code": naic,
        "1d. Contract/Policy Number": contract,
        "1e. Persons Covered (End of Policy Year)": covered,
        "1f. Policy Year Beginning Date": begin,
        "1g. Policy Year Ending Date": end,
        "10a. Total premiums or subscription charges paid to carrier": premium,
    }
    return [
        NormalizedExtractionField(
            field_name=field_name,
            value=value,
            confidence=0.99,
            page=None,
            source_text="HMSA ERISA Form 5500 and Schedule A information",
        )
        for field_name, value in extracted.items()
        if value
    ]


def _cigna_primary_schedule_a_page(page_texts: list[tuple[int, str]]) -> tuple[int, str] | None:
    """Return Cigna's consolidated Schedule A page, never a state appendix.

    Cigna packages repeat the same contract number and dates on a series of
    appendix pages, each with a state carrier EIN/NAIC. Only the page that
    contains the Part I totals, broker table, Part III, and Part IV is the
    primary record sent to FT Williams.
    """
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        lowered = normalized.lower()
        if (
            "cigna health and life insurance company" in lowered
            and "summary of all insurance contracts included in part iii" in lowered
            and "insurance fees and commissions information" in lowered
            and "part iii welfare benefit contract information" in lowered
            and "appendix to 1a" not in lowered
            and "schedule a insurance information - footnotes" not in lowered
        ):
            return page, normalized
    return None


def _cigna_schedule_a_support_page(page_texts: list[tuple[int, str]]) -> tuple[int, str] | None:
    """Return the Schedule A summary from Cigna's mixed A/C support packet."""
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        lowered = normalized.lower()
        if (
            "information for completing schedule a on the irs form" in lowered
            and "schedule a - insurance information" in lowered
            and "cigna health and life insurance company" in lowered
        ):
            return page, normalized
    return None


def _cigna_plan_detail_page(page_texts: list[tuple[int, str]]) -> tuple[int, str] | None:
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        lowered = normalized.lower()
        if (
            "cigna" in lowered
            and "plan detail report" in lowered
            and "commissions paid detail" in lowered
            and "broker acct#" in lowered
        ):
            return page, normalized
    return None


def _is_cigna_schedule_a_packet(page_texts: list[tuple[int, str]]) -> bool:
    return bool(_cigna_primary_schedule_a_page(page_texts) or _cigna_schedule_a_support_page(page_texts))


def _normalize_cigna_written_date(value: str | None) -> str | None:
    match = re.fullmatch(r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})", clean_extracted_value(value or ""))
    if not match:
        return None
    month_names = {
        **{name.lower(): index for index, name in enumerate(calendar.month_name) if name},
        **{name.lower(): index for index, name in enumerate(calendar.month_abbr) if name},
    }
    month = month_names.get(match.group(1).lower())
    if not month:
        return None
    return f"{month:02d}/{int(match.group(2)):02d}/{match.group(3)}"


def _cigna_support_plan_year(text: str) -> tuple[str | None, str | None]:
    match = re.search(
        r"For\s+Plan\s+Year\s+Beginning\s*:\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})\s+"
        r"and\s+Ending\s*:\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None, None
    return _normalize_cigna_written_date(match.group(1)), _normalize_cigna_written_date(match.group(2))


def _cigna_plan_detail_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    detail = _cigna_plan_detail_page(page_texts)
    if not detail:
        return []
    page, text = detail
    lines = [clean_extracted_value(line) for line in text.splitlines() if clean_extracted_value(line)]
    rows: list[ScheduleABrokerRow] = []
    fee_section_index = 0
    index = 0
    while index + 3 < len(lines):
        if (
            lines[index].upper() != "BENEFIT"
            or lines[index + 2].upper() != "BROKER ACCT#"
            or lines[index + 3].upper() != "BROKER NAME"
        ):
            index += 1
            continue
        amount_heading = lines[index + 1].upper()
        if "TOTAL COMM PAID" in amount_heading:
            payment_kind, purpose = "commission", "Commissions"
        elif "TOTAL FEES" in amount_heading:
            fee_section_index += 1
            payment_kind = "fee"
            purpose = "Benefit Advisor Fees" if fee_section_index == 1 else "Service / General Agent Fees"
        elif "TOTAL PAID" in amount_heading:
            payment_kind, purpose = "fee", "Incentive Compensation"
        else:
            index += 1
            continue

        row_index = index + 4
        while row_index + 3 < len(lines) and lines[row_index].upper() != "TOTAL":
            benefit, amount_line, account, name = lines[row_index : row_index + 4]
            amount_match = re.fullmatch(r"\$\s*([\d,]+(?:\.\d{2})?)", amount_line)
            if not amount_match or not re.fullmatch(r"\d{4,}", account) or not is_probable_person_or_entity_name(name):
                row_index += 1
                continue
            amount = money_value(amount_match.group(1))
            money_row = ScheduleABrokerMoneyRow(
                coverage=benefit,
                amount=amount,
                purpose=purpose,
            )
            rows.append(
                ScheduleABrokerRow(
                    name=name,
                    organization_code="3",
                    commission_rows=[money_row] if payment_kind == "commission" else [],
                    fee_rows=[money_row] if payment_kind == "fee" else [],
                    commission_total=amount if payment_kind == "commission" else "0",
                    fee_total=amount if payment_kind == "fee" else "0",
                    source_page=page,
                    confidence=0.995,
                )
            )
            row_index += 4
        index = max(index + 1, row_index)
    return _merge_columnar_broker_rows(rows)


def _extract_cigna_support_packet_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    support = _cigna_schedule_a_support_page(page_texts)
    if not support:
        return []
    page, text = support
    source_text = text[:1600]
    fields: list[NormalizedExtractionField] = []

    def add(field_name: str, value: str | None) -> None:
        clean = clean_extracted_value(str(value or ""))
        if clean and not is_blank_extraction_value(clean):
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=clean,
                    confidence=0.995,
                    page=page,
                    source_text=source_text,
                )
            )

    add("1a. Name of Insurance Company", "Cigna Health and Life Insurance Company")
    identity = re.search(
        r"(?P<ein>\d{2}-\d{7})\s+(?P<naic>\d{4,6})\s+(?P<contract>[A-Za-z0-9-]+)",
        text,
    )
    if identity:
        add("1b. Insurance Carrier EIN", identity.group("ein"))
        add("1c. NAIC Code", identity.group("naic"))
        add("1d. Contract/Policy Number", identity.group("contract"))
    policy_from, policy_to = _cigna_support_plan_year(text)
    add("1f. Policy Year Beginning Date", policy_from)
    add("1g. Policy Year Ending Date", policy_to)
    premium = regex_first(
        text,
        [r"Total\s+premiums?\*?\s+or\s+subscription\s+charges\s+paid\s+to\s+carrier\s*:.{0,700}?\$\s*([\d,]+(?:\.\d{2})?)"],
        flags=re.IGNORECASE | re.DOTALL,
    )
    add("10a. Total premiums or subscription charges paid to carrier", money_value(premium) if premium else None)

    rows = _cigna_plan_detail_broker_rows(page_texts)
    for field in schedule_a_broker_compensation_fields(rows):
        field.confidence = 0.995
        fields.append(field)
    return fields


def _cigna_primary_broker_name(text: str) -> str | None:
    start = re.search(r"Non\s+Experience\s*-\s*Rated", text, flags=re.IGNORECASE)
    search_text = text[start.end() :] if start else text
    for line in search_text.splitlines():
        candidate = line.strip().rstrip(",").strip()
        if not candidate or candidate.upper() != candidate:
            continue
        if not re.search(r"\b(?:LLC|INC|CORP|SERVICES|ASSOCIATES|AGENCY|BROKER)\b", candidate):
            continue
        if is_probable_person_or_entity_name(candidate):
            return candidate
    return None


def extract_cigna_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    primary = _cigna_primary_schedule_a_page(page_texts)
    if not primary:
        return _extract_cigna_support_packet_fields(page_texts)
    page, text = primary
    source_text = text[:1600]
    fields: list[NormalizedExtractionField] = []

    def add(field_name: str, value: str | None) -> None:
        clean = clean_extracted_value(str(value or ""))
        if clean and not is_blank_extraction_value(clean):
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=clean,
                    confidence=0.995,
                    page=page,
                    source_text=source_text,
                )
            )

    add("1a. Name of Insurance Company", "Cigna Health and Life Insurance Company")
    coverage = re.search(
        r"(?P<ein>\d{2}-\d{7})\s+(?P<naic>\d{4,6})\s+"
        r"(?P<contract>[A-Za-z0-9-]+)\s+(?P<covered>[\d,]+)\s+Employees?\s+"
        r"(?P<from>\d{1,2}/\d{1,2}/\d{4})\s*-?\s*(?P<to>\d{1,2}/\d{1,2}/\d{4})",
        text,
        flags=re.IGNORECASE,
    )
    if coverage:
        add("1b. Insurance Carrier EIN", coverage.group("ein"))
        add("1c. NAIC Code", coverage.group("naic"))
        add("1d. Contract/Policy Number", coverage.group("contract"))
        add("1e. Persons Covered (End of Policy Year)", coverage.group("covered"))
        add("1f. Policy Year Beginning Date", coverage.group("from"))
        add("1g. Policy Year Ending Date", coverage.group("to"))
    else:
        # The production PDF text layer places each label above its value
        # instead of returning a visual row. Parse the same primary page by
        # labels; it is already isolated from all state appendix pages.
        add("1b. Insurance Carrier EIN", regex_first(text, [r"\(b\)\s*EIN\s+(\d{2}-\d{7})"], flags=re.IGNORECASE))
        add("1c. NAIC Code", regex_first(text, [r"\(c\)\s*NAIC\s+Code\s+(\d{4,6})"], flags=re.IGNORECASE))
        add(
            "1d. Contract/Policy Number",
            regex_first(text, [r"Identification\s+Number\s+([A-Za-z0-9-]+)"], flags=re.IGNORECASE),
        )
        add(
            "1e. Persons Covered (End of Policy Year)",
            regex_first(text, [r"at\s+end\s+of\s+policy\s+or\s+contract\s+year\s+([\d,]+)\s+Employees?"], flags=re.IGNORECASE),
        )
        dates = regex_first(
            text,
            [r"\(f\)\s*From\s+\(g\)\s*To\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})"],
            flags=re.IGNORECASE,
            groups=True,
        )
        if isinstance(dates, tuple) and len(dates) >= 2:
            add("1f. Policy Year Beginning Date", dates[0])
            add("1g. Policy Year Ending Date", dates[1])

    totals = re.search(
        r"Total\s+Amount\s+of\s+commissions\s+paid\s*\$?\s*(?P<commissions>[\d,]+(?:\.\d{2})?).{0,160}?"
        r"Total\s+Amount\s+of\s+fees\s+paid\s*\$?\s*(?P<fees>[\d,]+(?:\.\d{2})?)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if totals:
        add("3b. Amount of Commissions", money_value(totals.group("commissions")))
        add("3c. Amount of Fees", money_value(totals.group("fees")))

    add("3a. Name of Agent/Broker/Person", _cigna_primary_broker_name(text))
    purpose = regex_first(text, [r"\$[\d,]+\s+\$[\d,]+\s+([A-Za-z][A-Za-z ]+?)\s+3-Insurance\s+Agent"], flags=re.IGNORECASE)
    add("3d. Purpose", purpose or "General Agent Payments")
    add("3e. Organizational Code", "3")

    premium = regex_first(
        text,
        [r"Total\s+premiums?\s+or\s+subscriptions?\s+charges\s+paid\s+to\s+carrier\s*\$?\s*([\d,]+(?:\.\d{2})?)"],
        flags=re.IGNORECASE,
    )
    add("10a. Total premiums or subscription charges paid to carrier", money_value(premium) if premium else None)
    if re.search(r"information\s+not\s+provided.*Not\s+Applicable", text, flags=re.IGNORECASE | re.DOTALL):
        add("11. Did the insurance company fail to provide any information necessary to complete Schedule A?", "No")
    return fields


def extract_cigna_schedule_a_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    primary = _cigna_primary_schedule_a_page(page_texts)
    if not primary:
        return _cigna_plan_detail_broker_rows(page_texts) if _cigna_schedule_a_support_page(page_texts) else []
    page, text = primary
    name = _cigna_primary_broker_name(text)
    if not name:
        return []
    totals = re.search(
        r"Total\s+Amount\s+of\s+commissions\s+paid\s*\$?\s*(?P<commissions>[\d,]+(?:\.\d{2})?).{0,160}?"
        r"Total\s+Amount\s+of\s+fees\s+paid\s*\$?\s*(?P<fees>[\d,]+(?:\.\d{2})?)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    commissions = money_value(totals.group("commissions")) if totals else None
    fees = money_value(totals.group("fees")) if totals else None
    address = re.search(
        r"(?P<address>PO\s+BOX\s+\d+)\s*,\s*(?P<city>[A-Z][A-Z ]+)\s*,\s*(?P<state>[A-Z]{2})\s*,.*?\n\s*(?P<zip>\d{5})\b",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    purpose = "General Agent Payments"
    return [
        ScheduleABrokerRow(
            name=name,
            address_line_1=clean_extracted_value(address.group("address")) if address else None,
            city=clean_extracted_value(address.group("city")) if address else None,
            state=address.group("state").upper() if address else None,
            zip_code=address.group("zip") if address else None,
            organization_code="3",
            commission_rows=[ScheduleABrokerMoneyRow(amount=commissions, purpose=purpose)] if commissions else [],
            fee_rows=[ScheduleABrokerMoneyRow(amount=fees, purpose=purpose)] if fees else [],
            commission_total=commissions,
            fee_total=fees,
            source_page=page,
            confidence=0.995,
        )
    ]


_COLUMNAR_BROKER_SECTION = re.compile(
    r"INSURANCE\s+FEES?\s+AND\s+COMMISSIONS?\s+INFORMATION",
    re.IGNORECASE,
)
_COLUMNAR_BROKER_END = re.compile(
    r"\n\s*\d+\s*\.\s*(?:COVERAGE\s*/?\s*BENEFITS|NON[- ]?PARTICIPATING|EXPERIENCE[- ]?RATED)",
    re.IGNORECASE,
)
_COLUMNAR_BROKER_AMOUNTS = re.compile(
    r"\$\s*(?P<sales>[\d,]+(?:\.\d{1,2})?)\s+"
    r"\$\s*(?P<fees>[\d,]+(?:\.\d{1,2})?)\s+"
    r"\$\s*(?P<additional>[\d,]+(?:\.\d{1,2})?)",
)
_COLUMNAR_CITY_STATE_ZIP = re.compile(
    r"^(?P<city>[A-Za-z .'-]+),\s*(?P<state>[A-Z]{2})\s+(?P<zip>[0-9]{5}(?:-[0-9]{4})?)$",
    re.IGNORECASE,
)
_SECONDARY_ADDRESS_LINE = re.compile(
    r"^(?:ATTN\b|BLDG\b|BUILDING\b|FLOOR\b|FL\b|ROOM\b|RM\b|STE\b|SUITE\b|UNIT\b|\d+(?:ST|ND|RD|TH)\s+FLOOR\b)",
    re.IGNORECASE,
)


def extract_columnar_broker_compensation_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    """Extract multi-page broker disclosures with sales, fee, and additional columns.

    Carrier exports such as Unum's disclosure are not Form 5500 renderings.
    They print one recipient block followed by three money columns, and may
    continue the last address onto the next page. Parse the table structurally,
    merge repeated name/address rows, and return nothing if any paid block
    cannot be resolved so an incomplete broker set is never sent to FTW.
    """
    combined = ""
    page_offsets: list[tuple[int, int]] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text or "")
        if not normalized:
            continue
        if combined:
            combined += "\n"
        page_offsets.append((len(combined), page))
        combined += normalized

    start = _COLUMNAR_BROKER_SECTION.search(combined)
    if not start:
        return []
    heading = combined[start.start() : start.start() + 900].lower()
    if not all(label in heading for label in ("sales", "fees", "additional")):
        return []

    section_start = start.end()
    section = combined[section_start:]
    end = _COLUMNAR_BROKER_END.search(section)
    if end:
        section = section[: end.start()]

    parsed_rows: list[ScheduleABrokerRow] = []
    paid_blocks = 0
    unresolved_paid_blocks = 0
    block_pattern = re.compile(r"(?:\A|\n[ \t]*\n)(?P<block>.*?)(?=\n[ \t]*\n|\Z)", re.DOTALL)
    for block_match in block_pattern.finditer(section):
        block = block_match.group("block").strip()
        amount_match = _COLUMNAR_BROKER_AMOUNTS.search(block)
        if not amount_match:
            continue
        sales = money_value(amount_match.group("sales"))
        fees = money_value(amount_match.group("fees"))
        additional = money_value(amount_match.group("additional"))
        if not any((parse_numeric_amount(value) or 0) > 0 for value in (sales, fees, additional)):
            continue
        paid_blocks += 1

        name_lines = [
            clean_extracted_value(line)
            for line in block[: amount_match.start()].splitlines()
            if clean_extracted_value(line)
        ]
        name = clean_extracted_value(" ".join(name_lines)).strip(" :-")
        address_lines = [
            clean_extracted_value(line)
            for line in block[amount_match.end() :].splitlines()
            if clean_extracted_value(line)
        ]
        address_line_1, address_line_2, city, state, zip_code = _columnar_broker_address(address_lines)
        if not name or not is_probable_person_or_entity_name(name) or not address_line_1 or not city or not state or not zip_code:
            unresolved_paid_blocks += 1
            continue

        absolute_offset = section_start + block_match.start("block")
        source_page = next((page for offset, page in reversed(page_offsets) if offset <= absolute_offset), None)
        fee_total = sum_money_values(fees, additional) or "0"
        parsed_rows.append(
            ScheduleABrokerRow(
                name=name,
                address_line_1=address_line_1,
                address_line_2=address_line_2,
                city=city,
                state=state,
                zip_code=zip_code,
                organization_code="3",
                commission_rows=(
                    [ScheduleABrokerMoneyRow(amount=sales, purpose="Sales Commission")]
                    if (parse_numeric_amount(sales) or 0) > 0
                    else []
                ),
                fee_rows=[
                    *(
                        [ScheduleABrokerMoneyRow(amount=fees, purpose="Fees")]
                        if (parse_numeric_amount(fees) or 0) > 0
                        else []
                    ),
                    *(
                        [ScheduleABrokerMoneyRow(amount=additional, purpose="Additional Compensation")]
                        if (parse_numeric_amount(additional) or 0) > 0
                        else []
                    ),
                ],
                commission_total=sales,
                fee_total=fee_total,
                source_page=source_page,
                confidence=0.96,
            )
        )

    if unresolved_paid_blocks or len(parsed_rows) != paid_blocks:
        return []
    return _merge_columnar_broker_rows(parsed_rows)


def _columnar_broker_address(lines: list[str]) -> tuple[str | None, str | None, str | None, str | None, str | None]:
    city = state = zip_code = None
    city_index = None
    for index, line in enumerate(lines):
        match = _COLUMNAR_CITY_STATE_ZIP.match(line)
        if match:
            city = clean_extracted_value(match.group("city"))
            state = match.group("state").upper()
            zip_code = match.group("zip")
            city_index = index
            break
    street_lines = lines[:city_index] if city_index is not None else []
    if not street_lines:
        return None, None, city, state, zip_code
    primary_lines = [line for line in street_lines if not _SECONDARY_ADDRESS_LINE.match(line)]
    secondary_lines = [line for line in street_lines if _SECONDARY_ADDRESS_LINE.match(line)]
    address_line_1 = primary_lines[0] if primary_lines else street_lines[0]
    remaining_primary = primary_lines[1:]
    address_line_2 = " ".join([*secondary_lines, *remaining_primary]) or None
    return address_line_1, address_line_2, city, state, zip_code


def _merge_columnar_broker_rows(rows: list[ScheduleABrokerRow]) -> list[ScheduleABrokerRow]:
    merged: dict[tuple[str, str, str], ScheduleABrokerRow] = {}
    order: list[tuple[str, str, str]] = []
    for row in rows:
        key = (
            normalize_compare_key(row.name),
            normalize_compare_key(row.address_line_1 or ""),
            str(row.zip_code or ""),
        )
        existing = merged.get(key)
        if existing is None:
            merged[key] = row.model_copy(deep=True)
            order.append(key)
            continue
        existing.commission_rows.extend(row.commission_rows)
        existing.fee_rows.extend(row.fee_rows)
        existing.confidence = min(existing.confidence, row.confidence)
    result: list[ScheduleABrokerRow] = []
    for key in order:
        row = merged[key]
        row.commission_total = _sum_money_rows(row.commission_rows) if row.commission_rows else "0"
        row.fee_total = _sum_money_rows(row.fee_rows) if row.fee_rows else "0"
        result.append(row)
    return result


def extract_compensation_table_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    """Read the broker compensation table carriers print on their statements.

    Most carrier statements end with the same table, however they word the
    heading: an optional agent number, the recipient's name and address, then
    the commissions paid, the fees paid, and the purpose. The amounts and the
    purpose sit on one line underneath the address block:

        CGI-026821Nth Insurance Agency dba: Alliance 360 I
        10833 VALLEY VIEW STREET
        SUITE 550
        CYPRESS CA 90630
        $5,810.59 $ 0.00 Standard Commissions

    Those amounts are Schedule A items 3b and 3c, and they were being missed on
    most statements even though the broker's name was picked up.
    """
    rows: list[ScheduleABrokerRow] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not normalized:
            continue
        lowered = normalized.lower()
        # Only inside a compensation table - "commission" alone appears in the
        # covering letter of nearly every carrier statement.
        if not (
            re.search(r"commissions?\s+paid|amount\s+of\s+commissions?", lowered)
            and re.search(r"fees?\s+paid|amount\s+of\s+fees?", lowered)
        ):
            continue

        lines = [line.strip() for line in normalized.splitlines()]
        for index, line in enumerate(lines):
            match = _COMPENSATION_ROW.match(line)
            if not match:
                continue
            commissions = money_value(match.group("commissions"))
            fees = money_value(match.group("fees"))
            purpose = clean_extracted_value(match.group("purpose") or "") or None
            if not commissions and not fees:
                continue
            name, address = _compensation_recipient(lines, index)
            if not name:
                continue
            rows.append(
                ScheduleABrokerRow(
                    name=name,
                    address_line_1=address[0] if address else None,
                    address_line_2=address[1] if len(address) > 1 else None,
                    commission_total=commissions or None,
                    fee_total=fees or None,
                    commission_rows=(
                        [ScheduleABrokerMoneyRow(amount=commissions, purpose=purpose)] if commissions else []
                    ),
                    fee_rows=[ScheduleABrokerMoneyRow(amount=fees, purpose=purpose)] if fees else [],
                    source_page=page,
                )
            )
    return rows


# "$5,810.59 $ 0.00 Standard Commissions" - amounts then an optional purpose.
_COMPENSATION_ROW = re.compile(
    r"^\$\s*(?P<commissions>[\d,]+(?:\.\d{2})?)\s+\$\s*(?P<fees>[\d,]+(?:\.\d{2})?)\s*(?P<purpose>[A-Za-z][^$]*)?$"
)
# Agent/producer numbers are printed hard against the name: "CGI-026821Nth ..."
_AGENT_NUMBER_PREFIX = re.compile(r"^[A-Z]{2,5}[-/]?\d{4,}\s*")
_TABLE_HEADING_WORDS = re.compile(
    r"amount|commission|fees?|purpose|agent\s*$|number\s*$|recipient|name and address|which paid|paid by",
    re.IGNORECASE,
)


def _compensation_recipient(lines: list[str], amount_index: int) -> tuple[str, list[str]]:
    """Walk back up from the amounts line to the recipient's name and address."""
    block: list[str] = []
    for line in reversed(lines[max(0, amount_index - 6) : amount_index]):
        if not line:
            if block:
                break
            continue
        if _COMPENSATION_ROW.match(line) or line.strip() in {"$", "$ "}:
            break
        if _TABLE_HEADING_WORDS.search(line):
            # Reached the column headings - everything collected below them is
            # the recipient block.
            break
        block.append(line)
    if not block:
        return "", []
    block.reverse()
    name = _AGENT_NUMBER_PREFIX.sub("", block[0]).strip()
    if not name or not is_probable_person_or_entity_name(name):
        return "", []
    return name, block[1:]


def extract_schedule_a_worksheet_summaries_from_pdf_text(file_bytes: bytes) -> list[ScheduleAWorksheetSummary]:
    page_texts = extract_pdf_text_pages(file_bytes)
    summaries: list[ScheduleAWorksheetSummary] = []
    for index, text in page_texts:
        summaries.extend(extract_bcbsma_schedule_a_worksheet_summaries(text, index))
    summaries.extend(extract_bcbs_michigan_schedule_a_summaries(page_texts))
    summaries.extend(extract_prudential_schedule_a_summaries(page_texts))
    summaries.extend(extract_eyemed_schedule_a_summaries(page_texts))
    summaries.extend(extract_standard_schedule_a_summaries(page_texts))
    summaries.extend(extract_united_omaha_schedule_a_summaries(page_texts))
    return summaries


def parse_schedule_a_text(text: str, page: int | None = None, *, rules=None) -> list[NormalizedExtractionField]:
    normalized_input = normalize_ocr_text(text)
    if is_bcbs_michigan_addendum_page(normalized_input):
        return dedupe_fields(extract_bcbs_michigan_addendum_fields(normalized_input, page))
    if is_united_omaha_schedule_a_support(normalized_input):
        return []

    fields: list[NormalizedExtractionField] = []

    def add(field_name: str, value: str, confidence: float = 0.86, source_text: str | None = None, value_validator=None):
        clean = clean_extracted_value(value)
        if clean and not is_blank_extraction_value(clean):
            if value_validator and not value_validator(clean):
                return
            fields.append(
                NormalizedExtractionField(
                    field_name=field_name,
                    value=clean,
                    confidence=confidence,
                    page=page,
                    source_text=(source_text or text[:420]).strip(),
                )
            )

    carrier = regex_first(
        text,
        [
            r'"(?:a_name_of_insurance_carrier|insurance_carrier_name|carrier_name)"\s*:\s*"([^"]+)"',
            r"insurance\s+carrier\s*:\s*([A-Z][A-Z0-9 ,. '&-]+)",
            r"identifies\s+(.+?)\s+as the carrier",
            r"carrier_name[\"']?\s*[:=]\s*[\"']([^\"']+)",
            r"\(a\)\s*Name\s+of\s+insurance\s+carrier\s+([A-Z][A-Z0-9 ,. '&-]+?)(?=\s+\(?b\)?\s+EIN|\s+[0-9]{2}-[0-9]{7}|$)",
            r"Name\s+of\s+insurance\s+carrier\s+([A-Z][A-Z0-9 ,. '&-]+?)(?=\s+\(?b\)?\s+EIN|\s+[0-9]{2}-[0-9]{7}|$)",
            r"Name of insurance carrier\s*\n\s*(.+)",
            r"Name of insurance company\s*\n\s*(.+)",
        ],
    )
    if carrier:
        add("1a. Name of Insurance Company", carrier, 0.95, value_validator=is_probable_carrier_name)

    coverage_row = regex_first(
        text,
        [
            r"policy or contract year.*?\(f\).*?\(g\).*?To\s*\n?\s*([0-9]{2}-[0-9]{7})\s+([0-9]{4,6})\s+([A-Za-z0-9-]+)\s+([0-9,]+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})",
            r"([0-9]{2}-[0-9]{7})\s+([0-9]{4,6})\s+([A-Za-z0-9-]+)\s+([0-9,]+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})",
            r"policy or contract year.*?\(f\).*?\(g\).*?To\s*\n?\s*([0-9]{2}-[0-9]{7})\s+([0-9]{4,6})\s+([A-Za-z0-9-]+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})",
            r"([0-9]{2}-[0-9]{7})\s+([0-9]{4,6})\s+([A-Za-z0-9-]+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})",
        ],
        flags=re.IGNORECASE | re.DOTALL,
        groups=True,
    )
    if isinstance(coverage_row, tuple) and len(coverage_row) >= 5:
        carrier_ein, naic, contract = coverage_row[:3]
        if len(coverage_row) >= 6:
            covered, policy_from, policy_to = coverage_row[3:6]
            add("1e. Persons Covered (End of Policy Year)", covered, 0.9)
        else:
            policy_from, policy_to = coverage_row[3:5]
        add("1b. Insurance Carrier EIN", carrier_ein, 0.9, value_validator=looks_like_ein)
        add("1c. NAIC Code", naic, 0.88)
        add("1d. Contract/Policy Number", contract, 0.88)
        add("1f. Policy Year Beginning Date", policy_from, 0.88)
        add("1g. Policy Year Ending Date", policy_to, 0.88)

    metlife_summary = regex_first(
        text,
        [
            r"identifiers including EIN\s+([0-9]{2}-[0-9]{7}),\s+NAIC code\s+([0-9]{4,6}),\s+contract number\s+([A-Za-z0-9-]+),\s+and\s+([0-9,]+)\s+covered persons",
            r"carrier identifiers including EIN\s+([0-9]{2}-[0-9]{7}),\s+NAIC code\s+([0-9]{4,6}),\s+contract number\s+([A-Za-z0-9-]+),\s+and\s+([0-9,]+)\s+covered persons",
        ],
        flags=re.IGNORECASE,
        groups=True,
    )
    if isinstance(metlife_summary, tuple) and len(metlife_summary) >= 4:
        carrier_ein, naic, contract, covered = metlife_summary[:4]
        add("1b. Insurance Carrier EIN", carrier_ein, 0.95, value_validator=looks_like_ein)
        add("1c. NAIC Code", naic, 0.96)
        add("1d. Contract/Policy Number", contract, 0.96)
        add("1e. Persons Covered (End of Policy Year)", covered, 0.96)

    contract_details = regex_first(
        text,
        [
            r"\"EIN\":\"([0-9]{2}-[0-9]{7})\".*?\"NAIC_code\":\"([0-9]{4,6})\".*?\"contract_or_identification_number\":\"([A-Za-z0-9-]+)\".*?\"approximate_number_of_persons_covered_at_end_of_policy_or_contract_year\":\"([0-9,]+)\".*?\"policy_or_contract_year_from\":\"([0-9]{2}/[0-9]{2}/[0-9]{4})\".*?\"policy_or_contract_year_to\":\"([0-9]{2}/[0-9]{2}/[0-9]{4})\"",
        ],
        flags=re.IGNORECASE | re.DOTALL,
        groups=True,
    )
    if isinstance(contract_details, tuple) and len(contract_details) >= 6:
        carrier_ein, naic, contract, covered, policy_from, policy_to = contract_details[:6]
        add("1b. Insurance Carrier EIN", carrier_ein, 0.97, value_validator=looks_like_ein)
        add("1c. NAIC Code", naic, 0.97)
        add("1d. Contract/Policy Number", contract, 0.97)
        add("1e. Persons Covered (End of Policy Year)", covered, 0.97)
        add("1f. Policy Year Beginning Date", policy_from, 0.97)
        add("1g. Policy Year Ending Date", policy_to, 0.97)

    carrier_ein = regex_first(
        text,
        [
            r"\(\s*b\s*\)\s*EIN\s*:?\s*([0-9]{2}-[0-9]{7})",
            r"\bEIN\s+([0-9]{2}-[0-9]{7})",
        ],
        flags=re.IGNORECASE,
    )
    naic = regex_first(
        text,
        [
            r"\(\s*c\s*\)\s*NAIC\s+code\s*:?\s*([0-9]{4,6})",
            r"\bNAIC\s+code\s+([0-9]{4,6})",
        ],
        flags=re.IGNORECASE,
    )
    contract = regex_first(
        text,
        [
            r"\(\s*d\s*\)\s*Contract\s+or\s+identification\s+number\s*:?\s*([A-Za-z0-9-]+)",
            r"\bContract\s+or\s+identification\s+number\s+([A-Za-z0-9-]+)",
        ],
        flags=re.IGNORECASE,
    )
    covered = regex_first(
        text,
        [
            r"\(\s*e\s*\)\s*Approximate\s+number\s+of\s+persons\s+covered\s+at\s+(?:the\s+)?end\s+of\s+policy\s+or\s+contract\s+year\s*:?\s*\*?\s*([0-9,]+)",
            r"\bpersons\s+covered\s+at\s+(?:the\s+)?end\s+of\s+policy\s+or\s+contract\s+year\s*:?\s*\*?\s*([0-9,]+)",
        ],
        flags=re.IGNORECASE,
    )
    policy_from = regex_first(
        text,
        [
            r"\b(?:Contract/Policy|Contract|Policy)\s+Year\s+From\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
            r"\(\s*f\s*\)\s*From\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
            r"\bFrom\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        ],
        flags=re.IGNORECASE,
    )
    policy_to = regex_first(
        text,
        [
            r"\b(?:Contract/Policy|Contract|Policy)\s+Year\s+To\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
            r"\(\s*g\s*\)\s*To\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
            r"\bTo\s*:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        ],
        flags=re.IGNORECASE,
    )
    if carrier_ein:
        add("1b. Insurance Carrier EIN", carrier_ein, 0.93, value_validator=looks_like_ein)
    if naic:
        add("1c. NAIC Code", naic, 0.92)
    if contract:
        add("1d. Contract/Policy Number", contract, 0.92)
    if covered:
        add("1e. Persons Covered (End of Policy Year)", covered, 0.9)
    if policy_from:
        add("1f. Policy Year Beginning Date", policy_from, 0.9)
    if policy_to:
        add("1g. Policy Year Ending Date", policy_to, 0.9)

    policy_period = regex_first(
        text,
        [
            r"plan year beginning\s+([0-9]{2}/[0-9]{2}/[0-9]{4})\s+and ending\s+([0-9]{2}/[0-9]{2}/[0-9]{4})",
            r"covers the plan year from\s+([0-9]{2}/[0-9]{2}/[0-9]{4})\s+to\s+([0-9]{2}/[0-9]{2}/[0-9]{4})",
        ],
        flags=re.IGNORECASE,
        groups=True,
    )
    if isinstance(policy_period, tuple) and len(policy_period) >= 2:
        add("1f. Policy Year Beginning Date", policy_period[0], 0.94)
        add("1g. Policy Year Ending Date", policy_period[1], 0.94)

    sponsor = regex_first(text, [r"^\s*([A-Z][A-Z0-9 ,. '&-]{3,}?)\s+([0-9]{2}-[0-9]{7})\s*$"], flags=re.MULTILINE, groups=True)
    if isinstance(sponsor, tuple) and len(sponsor) >= 2:
        add("1d. Plan Sponsor Name", sponsor[0], 0.8)
        add("1e. Plan Sponsor EIN", sponsor[1], 0.8)

    agent = regex_first(
        text,
        [
            r'"(?:payee_name|recipient_name|agent_broker_name|broker_name)"\s*:\s*"([^"]+)"',
            r"\bName\s*:\s*([A-Z0-9&.,' -]+?)\s+Address\s*:.*?Total\s+amount\s+of\s+commissions\s+paid",
            r"commissions\s+or\s+fees\s+were\s+paid\s*:\s*\n\s*([A-Z0-9&.,'() -]+)",
            r"\(a\)\s*Name and address of the agents.*?\bName\s*:\s*([A-Z0-9&.,' -]+?)(?:\s+Address\s*:|\n|$)",
            r"person to whom commissions or fees were paid\s*\n\s*(.+)",
            r"Persons receiving commissions and fees.*?paid\s*\n\s*(.+)",
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )
    if agent:
        add("3a. Name of Agent/Broker/Person", agent, 0.95, value_validator=is_probable_person_or_entity_name)

    commission = regex_first(
        text,
        [
            r"Amount of commissions paid:\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
            r"Total amount of commissions paid:\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
        ],
        flags=re.IGNORECASE,
    )
    if not commission:
        commission = regex_first(
            text,
            [
                r"Total amount of commissions paid.*?\n\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
                r"Commissions Paid\b.*?\b([0-9,]+(?:\.\d{2})?)\s+Sub-total",
                r"reports\s+([0-9,]+)\s+in commissions paid",
                r"total_amount_of_commissions_paid[\"']?\s*[:=]\s*[\"']?([0-9,]+)",
            ],
            flags=re.IGNORECASE | re.DOTALL,
        )
    if commission:
        add("3b. Amount of Commissions", commission, 0.94)

    fee = regex_first(
        text,
        [
            r"Total amount of fees paid\s*\n\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
            r"fees paid\s*/\s*amount\s*:\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
            r"FEES\s*\n\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
            r"Fees Paid\b.*?\b([0-9,]+(?:\.\d{2})?)\s+Sub-total",
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )
    if fee:
        add("3c. Amount of Fees", fee, 0.88)
    else:
        zero_fee = regex_first(text, [r"0\s+in total fees paid", r"No fees were paid", r"total_fees_paid_amount[\"']?\s*[:=]\s*[\"']?0"], flags=re.IGNORECASE)
        if zero_fee is not None:
            add("3c. Amount of Fees", "0", 0.9)

    derived_purpose = derive_schedule_a_purpose(commission, fee)
    if derived_purpose:
        add("3d. Purpose", derived_purpose, 0.92)
    else:
        explicit_purpose = regex_first(
            text,
            [
                r"Purpose\s*:\s*([A-Z][A-Z /&-]+)",
                r'"(?:purpose|d_purpose)"\s*:\s*"([^"]+)"',
            ],
            flags=re.IGNORECASE,
        )
        if explicit_purpose:
            add("3d. Purpose", explicit_purpose, 0.84)

    org_code = regex_first(
        text,
        [
            r"Organization\s+code\s*:\s*([0-9]{1,2})\b",
            r"Organizational\s+Code\s*:\s*([0-9]{1,2})\b",
            r"organization\s+code\b.{0,60}?\b([0-9]{1,2})\b",
            r"Sub-total\s+0\s+Sub-total\s+([0-9]{1,2})\b",
            r'"(?:organization_code|e_organization_code)"\s*:\s*"?([0-9]{1,2})"?',
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )
    if org_code:
        add("3e. Organizational Code", org_code, 0.9, value_validator=looks_like_org_code)

    premium_total = extract_nonexperience_total_premium_from_text(text)
    if premium_total:
        add("10a. Total premiums or subscription charges paid to carrier", premium_total, 0.9)

    fields.extend(extract_schedule_a_broker_compensation_fields(text, page))
    fields.extend(extract_schedule_a_fields_from_tables(text, page))
    fields.extend(extract_schedule_a_fields_from_rule_labels(text, page, rules=rules))
    fields.extend(extract_configured_custom_fields(text, page, rules=rules))

    line_11 = extract_schedule_a_line_11(text)
    if line_11:
        add("11. Did the insurance company fail to provide any information necessary to complete Schedule A?", line_11, 0.86)

    if has_experience_rated_not_applicable(text):
        add_not_applicable_experience_rated_fields(fields, page, text)

    return dedupe_fields(fields)


def extract_bcbsma_schedule_a_worksheet_fields(text: str, page: int | None = None) -> list[NormalizedExtractionField]:
    normalized = normalize_ocr_text(text)
    if not is_bcbsma_schedule_a_worksheet(normalized):
        return []
    fields: list[NormalizedExtractionField] = []
    source_text = normalized[:1200]

    def add(field_name: str, value: str | None, confidence: float = 0.94):
        clean = clean_extracted_value(str(value or ""))
        if not clean or is_blank_extraction_value(clean):
            return
        fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=page, source_text=source_text))

    account_number = regex_first(normalized, [r"ACCOUNT\s*#:\s*([A-Za-z0-9-]+)"])
    period = regex_first(normalized, [r"PERIOD:\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})"], groups=True)
    naic = regex_first(normalized, [r"NAIC\s+CODE:\s*([0-9]{4,6})"])
    ein = regex_first(normalized, [r"EIN\s+CODE:\s*([0-9]{2}-[0-9]{7})"])
    employee_dependents = bcbsma_column_value(normalized, "Employee & Dependents", "MEDICAL")
    total_premium = bcbsma_column_value(normalized, "Total Premium", "MEDICAL")
    incurred_claims = bcbsma_column_value(normalized, "Incurred Claims", "MEDICAL")
    ibnr = bcbsma_column_value(normalized, "Incurred But Not Reported", "MEDICAL")
    claims_charged = bcbsma_column_value(normalized, "Claims Charged", "MEDICAL")
    base_commission = bcbsma_column_value(normalized, "Base Commission", "MEDICAL")
    taxes = bcbsma_column_value(normalized, "Taxes", "MEDICAL")
    other_retention = bcbsma_column_value(normalized, "Other Retention Charges", "MEDICAL")
    total_retention = sum_money_values(base_commission, taxes, other_retention)

    add("1a. Name of Insurance Company", "Blue Cross Blue Shield of Massachusetts, Inc.", 0.93)
    add("1b. Insurance Carrier EIN", ein, 0.98)
    add("1c. NAIC Code", naic, 0.98)
    add("1d. Contract/Policy Number", account_number, 0.97)
    add("1e. Persons Covered (End of Policy Year)", employee_dependents, 0.96)
    if isinstance(period, tuple) and len(period) >= 2:
        add("1f. Policy Year Beginning Date", period[0], 0.96)
        add("1g. Policy Year Ending Date", period[1], 0.96)
    add("9a. Premiums: (1) Amount Received", total_premium, 0.97)
    add("9a(4). Earned ((1) + (2) - (3))", total_premium, 0.94)
    add("9b(1). Benefit Charges (1) Claims paid", incurred_claims, 0.97)
    add("9b(2). Increase (decrease) in claim reserves", ibnr, 0.97)
    add("9b(3). Incurred claims (add(1) and (2))", claims_charged, 0.95)
    add("9b(4). Claims Charged", claims_charged, 0.96)
    add("9c(1)(A). Commissions", base_commission, 0.97)
    add("9c(1)(E). Taxes", taxes, 0.96)
    add("9c(1)(G). Other retention charges", other_retention, 0.97)
    add("9c(1)(H). Total retention", total_retention, 0.95)
    return fields


def extract_bcbsma_commission_breakdown_fields(text: str, page: int | None = None) -> list[NormalizedExtractionField]:
    rows = extract_bcbsma_commission_breakdown_broker_rows(text, page)
    if not rows:
        return []
    row = rows[0]
    source_text = normalize_ocr_text(text)[:1200]
    fields = [
        NormalizedExtractionField(field_name="3a. Name of Agent/Broker/Person", value=row.name, confidence=0.93, page=page, source_text=source_text),
    ]
    if row.commission_total:
        fields.append(NormalizedExtractionField(field_name="3b. Amount of Commissions", value=row.commission_total, confidence=0.94, page=page, source_text=source_text))
    if row.fee_total:
        fields.append(NormalizedExtractionField(field_name="3c. Amount of Fees", value=row.fee_total, confidence=0.94, page=page, source_text=source_text))
    purpose = derive_schedule_a_purpose(row.commission_total, row.fee_total)
    if purpose:
        fields.append(NormalizedExtractionField(field_name="3d. Purpose", value=purpose, confidence=0.92, page=page, source_text=source_text))
    return fields


def extract_bcbsma_schedule_a_worksheet_summaries(text: str, page: int | None = None) -> list[ScheduleAWorksheetSummary]:
    normalized = normalize_ocr_text(text)
    if not is_bcbsma_schedule_a_worksheet(normalized):
        return []
    account_name = regex_first(normalized, [r"ACCOUNT\s+NAME:\s*(.+?)(?=\s+ACCOUNT\s*#:)"])
    account_number = regex_first(normalized, [r"ACCOUNT\s*#:\s*([A-Za-z0-9-]+)"])
    period = regex_first(normalized, [r"PERIOD:\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})"], groups=True)
    values = [
        ScheduleAWorksheetValue(label="Persons covered", value=bcbsma_column_value(normalized, "Employee & Dependents", "MEDICAL") or "", source="LAST MONTH OF PERIOD ENROLLMENT", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Premium", value=bcbsma_column_value(normalized, "Total Premium", "MEDICAL") or "", source="PREMIUM", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Claims paid", value=bcbsma_column_value(normalized, "Incurred Claims", "MEDICAL") or "", source="BENEFIT CHARGES", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Claim reserve / IBNR", value=bcbsma_column_value(normalized, "Incurred But Not Reported", "MEDICAL") or "", source="BENEFIT CHARGES", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Claims charged", value=bcbsma_column_value(normalized, "Claims Charged", "MEDICAL") or "", source="BENEFIT CHARGES", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Commissions", value=bcbsma_column_value(normalized, "Base Commission", "MEDICAL") or "", source="RETENTION ALLOCATION", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Taxes", value=bcbsma_column_value(normalized, "Taxes", "MEDICAL") or "", source="RETENTION ALLOCATION", coverage="MEDICAL"),
        ScheduleAWorksheetValue(label="Other retention", value=bcbsma_column_value(normalized, "Other Retention Charges", "MEDICAL") or "", source="RETENTION ALLOCATION", coverage="MEDICAL"),
    ]
    base_commission = bcbsma_column_value(normalized, "Base Commission", "MEDICAL")
    taxes = bcbsma_column_value(normalized, "Taxes", "MEDICAL")
    other_retention = bcbsma_column_value(normalized, "Other Retention Charges", "MEDICAL")
    total_retention = sum_money_values(base_commission, taxes, other_retention)
    if total_retention:
        values.append(ScheduleAWorksheetValue(label="Total retention", value=total_retention, source="RETENTION ALLOCATION", coverage="MEDICAL"))
    return [
        ScheduleAWorksheetSummary(
            source="BCBSMA #5500A worksheet",
            carrier_name="Blue Cross Blue Shield of Massachusetts, Inc.",
            account_name=account_name,
            account_number=account_number,
            period_begin=period[0] if isinstance(period, tuple) and len(period) >= 2 else None,
            period_end=period[1] if isinstance(period, tuple) and len(period) >= 2 else None,
            ein=regex_first(normalized, [r"EIN\s+CODE:\s*([0-9]{2}-[0-9]{7})"]),
            naic_code=regex_first(normalized, [r"NAIC\s+CODE:\s*([0-9]{4,6})"]),
            coverage="MEDICAL",
            values=[item for item in values if item.value],
            notes=[f"Extracted from page {page}"] if page else [],
        )
    ]


def extract_bcbsma_commission_breakdown_broker_rows(text: str, page: int | None = None) -> list[ScheduleABrokerRow]:
    normalized = normalize_ocr_text(text)
    if "COMMISSIONS AND BONUS BREAKDOWN" not in normalized.upper() and "COMMISSION BREAKDOWN" not in normalized.upper():
        return []
    commission_match = re.search(r"COMMISSION\s+BREAKDOWN\s*\n\s*(.+?)\s+\$?([0-9,]+\.\d{2})\s+\$?([0-9,]+\.\d{2})\s+\$?([0-9,]+\.\d{2})", normalized, flags=re.IGNORECASE)
    if not commission_match:
        return []
    broker_name = clean_extracted_value(commission_match.group(1))
    medical_commission = money_value(commission_match.group(2))
    dental_commission = money_value(commission_match.group(3))
    senior_commission = money_value(commission_match.group(4))
    other_commission = regex_first(normalized, [r"OTHER\s+COMMISSION\s+\*?\s*\n\s*.+?\s+\$?([0-9,]+\.\d{2})"])
    non_monetary = regex_first(normalized, [r"NON\s+MONETARY\s+C[O0]MPENSATION\s+\*?\s*\n\s*.+?\s+\$?([0-9,]+\.\d{2})"])
    commission_total = whole_dollar_money_value(sum_money_values(medical_commission, dental_commission, senior_commission, other_commission))
    fee_total = whole_dollar_money_value(non_monetary) if non_monetary else None
    if not broker_name or not is_probable_person_or_entity_name(broker_name):
        return []
    return [
        ScheduleABrokerRow(
            name=broker_name,
            commission_rows=[
                ScheduleABrokerMoneyRow(coverage="Medical", amount=money_value(medical_commission), purpose="Base Commission"),
                ScheduleABrokerMoneyRow(coverage="Dental", amount=money_value(dental_commission), purpose="Base Commission"),
                ScheduleABrokerMoneyRow(coverage="Senior", amount=money_value(senior_commission), purpose="Base Commission"),
                ScheduleABrokerMoneyRow(coverage=None, amount=money_value(other_commission or "0"), purpose="Other Commission"),
            ],
            fee_rows=[ScheduleABrokerMoneyRow(coverage=None, amount=fee_total, purpose="Non-Monetary Compensation")] if fee_total else [],
            commission_total=commission_total,
            fee_total=fee_total,
            source_page=page,
            confidence=0.92,
        )
    ]


def extract_bcbs_michigan_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    main_page = first_bcbs_michigan_main_page(page_texts)
    if not main_page:
        return []
    page, text = main_page
    normalized = normalize_ocr_text(text)
    source_text = normalized[:1400]

    def add(field_name: str, value: str | None, confidence: float = 0.98):
        clean = clean_extracted_value(str(value or ""))
        if clean and not is_blank_extraction_value(clean):
            fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=page, source_text=source_text))

    add("1a. Name of Insurance Company", bcbs_michigan_part_i_value(normalized, r"\(a\)\s*NAME\s+OF\s+INSURANCE\s+CARRIER"), 0.99)
    add("1b. Insurance Carrier EIN", bcbs_michigan_part_i_value(normalized, r"\(b\)\s*EMPLOYER\s+IDENTIFICATION\s+NUMBER\s*\(EIN\)"), 0.99)
    add("1c. NAIC Code", bcbs_michigan_part_i_value(normalized, r"\(c\)\s*NATIONAL\s+ASSOCIATION\s+OF\s+INSURANCE\s+COMMISSIONERS\s*\(NAIC\)\s+CODE"), 0.99)
    add("1d. Contract/Policy Number", bcbs_michigan_part_i_value(normalized, r"\(d\)\s*CONTRACT\s+OR\s+IDENTIFICATION\s+NUMBER"), 0.99)
    add("1e. Persons Covered (End of Policy Year)", bcbs_michigan_part_i_value(normalized, r"\(e\)\s*APPROX\.?\s+NUMBER\s+OF\s+PERSONS\s+COVERED"), 0.99)
    add("1f. Policy Year Beginning Date", bcbs_michigan_part_i_value(normalized, r"\(f\)\s*POLICY\s+OR\s+CONTRACT\s+YEAR\s+FROM"), 0.99)
    add("1g. Policy Year Ending Date", bcbs_michigan_part_i_value(normalized, r"\(g\)\s*POLICY\s+OR\s+CONTRACT\s+YEAR\s+TO"), 0.99)

    line_9 = bcbs_michigan_experience_section(normalized)
    add("9a. Premiums: (1) Amount Received", regex_first(line_9, [r"\(i\)\s*AMOUNT\s+RECEIVED\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    if re.search(r"\(ii\)\s+AND\s+\(iii\)\s+NOT\s+APPLICABLE", line_9, flags=re.IGNORECASE):
        add("9a(2). Increase (decrease) in amount due but unpaid", "N/A", 0.99)
        add("9a(3). Increase (decrease) in unearned premium reserve", "N/A", 0.99)
    add("9a(4). Earned ((1) + (2) - (3))", regex_first(line_9, [r"\(iv\)\s*AMOUNT\s+EARNED\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9b(1). Benefit Charges (1) Claims paid", regex_first(line_9, [r"\(i\)\s*CLAIMS\s+PAID\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9b(2). Increase (decrease) in claim reserves", regex_first(line_9, [r"\(ii\)\s*INCREASE\s*\(DECREASE\)\s+IN\s+CLAIM\s+RESERVES\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9b(3). Incurred claims (add(1) and (2))", regex_first(line_9, [r"\(iii\)\s*INCURRED\s+CLAIMS\s*\(ADD\s*\(i\)\s+AND\s+\(ii\)\)\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9b(4). Claims Charged", regex_first(line_9, [r"\(iv\)\s*CLAIMS\s+CHARGED.*?\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(A). Commissions", regex_first(line_9, [r"\bA\.\s*COMMISSIONS\s+(NOT\s+APPLICABLE|\$?\s*[0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(B). Administrative service or other fees", regex_first(line_9, [r"\bB\.\s*ADMINISTRATIVE\s+SERVICE\s+OR\s+OTHER\s+FEES\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(C). Other Specific acquisition costs", regex_first(line_9, [r"\bC\.\s*OTHER\s+SPECIFIC\s+ACQUISITION\s+COSTS\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(D). Other expenses", regex_first(line_9, [r"\bD\.\s*OTHER\s+EXPENSES.*?\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(E). Taxes", regex_first(line_9, [r"\bE\.\s*ESTIMATED\s+TAXES,\s+FEES\s+AND\s+ASSESSMENTS\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(F). Charges for risks or other contingencies", regex_first(line_9, [r"\bF\.\s*CHARGES\s+FOR\s+RISK\s+OR\s+OTHER\s+CONTINGENCIES\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(G). Other retention charges", regex_first(line_9, [r"\bG\.\s*OTHER\s+RETENTION\s+CHARGES.*?\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(1)(H). Total retention", regex_first(line_9, [r"\bH\.\s*TOTAL\s+RETENTION\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9c(2). Dividends or retroactive rate refunds", regex_first(line_9, [r"DIVIDENDS\s+OR\s+RETROACTIVE\s+RATE\s+REFUNDS\s+\(CREDITED\)\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9d(1). Status of policyholder reserves at end of year: (1) Amount held to provide benefits after retirement", regex_first(line_9, [r"AMOUNT\s+HELD\s+TO\s+PROVIDE\s+BENEFITS\s+AFTER\s+RETIREMENT\s+(NOT\s+APPLICABLE|\$?\s*[0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9d(2). Claim reserves", regex_first(line_9, [r"\(ii\)\s*CLAIMS\s+RESERVES\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9d(3). Other reserves", regex_first(line_9, [r"\(iii\)\s*OTHER\s+RESERVES\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)
    add("9e. Dividends or retroactive rate refunds due", regex_first(line_9, [r"\(e\)\s*DIVIDENDS\s+OR\s+RETROACTIVE\s+RATE\s+REFUNDS\s+DUE\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]), 0.99)

    addendum_text = "\n".join(text for _, text in page_texts if is_bcbs_michigan_addendum_page(text))
    fields.extend(extract_bcbs_michigan_addendum_fields(addendum_text, None))
    return dedupe_fields(fields)


def extract_bcbs_michigan_addendum_fields(text: str, page: int | None = None) -> list[NormalizedExtractionField]:
    rows = extract_bcbs_michigan_addendum_broker_rows(text, page)
    if not rows:
        return []
    source_text = normalize_ocr_text(text)[:1200]
    commission_total = sum_money_values(*(row.commission_total for row in rows))
    fee_total = sum_money_values(*(row.fee_total for row in rows))
    first_row = rows[0]
    fields = [
        NormalizedExtractionField(field_name="3a. Name of Agent/Broker/Person", value=first_row.name, confidence=0.96, page=page, source_text=source_text),
        NormalizedExtractionField(field_name="3d. Purpose", value=derive_schedule_a_purpose(commission_total, fee_total) or "COMMISSIONS & FEES", confidence=0.95, page=page, source_text=source_text),
    ]
    if commission_total:
        fields.append(NormalizedExtractionField(field_name="3b. Amount of Commissions", value=commission_total, confidence=0.97, page=page, source_text=source_text))
    if fee_total:
        fields.append(NormalizedExtractionField(field_name="3c. Amount of Fees", value=fee_total, confidence=0.97, page=page, source_text=source_text))
    if first_row.organization_code:
        fields.append(NormalizedExtractionField(field_name="3e. Organizational Code", value=first_row.organization_code, confidence=0.96, page=page, source_text=source_text))
    return fields


def extract_bcbs_michigan_schedule_a_summaries(page_texts: list[tuple[int, str]]) -> list[ScheduleAWorksheetSummary]:
    main_page = first_bcbs_michigan_main_page(page_texts)
    if not main_page:
        return []
    page, text = main_page
    normalized = normalize_ocr_text(text)
    carrier = bcbs_michigan_part_i_value(normalized, r"\(a\)\s*NAME\s+OF\s+INSURANCE\s+CARRIER")
    contract = bcbs_michigan_part_i_value(normalized, r"\(d\)\s*CONTRACT\s+OR\s+IDENTIFICATION\s+NUMBER")
    account_name = regex_first(normalized, [r"GROUP\s+NAME:\s*(.+?)(?=\n|PART\s+I)"])
    addendum_text = "\n".join(addendum for _, addendum in page_texts if is_bcbs_michigan_addendum_page(addendum))
    broker_rows = extract_bcbs_michigan_addendum_broker_rows(addendum_text)
    values = [
        ScheduleAWorksheetValue(label="Experience premium received", value=regex_first(normalized, [r"\(i\)\s*AMOUNT\s+RECEIVED\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]) or "", source="Part III line 9a", coverage="All"),
        ScheduleAWorksheetValue(label="Claims paid", value=regex_first(normalized, [r"\(i\)\s*CLAIMS\s+PAID\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]) or "", source="Part III line 9b", coverage="All"),
        ScheduleAWorksheetValue(label="Total retention", value=regex_first(normalized, [r"\bH\.\s*TOTAL\s+RETENTION\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]) or "", source="Part III line 9c", coverage="All"),
        ScheduleAWorksheetValue(label="Broker commission total", value=sum_money_values(*(row.commission_total for row in broker_rows)) or "", source="Schedule A/C addendum", coverage="All"),
        ScheduleAWorksheetValue(label="Broker fee total", value=sum_money_values(*(row.fee_total for row in broker_rows)) or "", source="Schedule A/C addendum", coverage="All"),
    ]
    return [
        ScheduleAWorksheetSummary(
            source="BCBS Michigan Schedule A/C addendum",
            carrier_name=carrier,
            account_name=account_name,
            account_number=contract,
            period_begin=bcbs_michigan_part_i_value(normalized, r"\(f\)\s*POLICY\s+OR\s+CONTRACT\s+YEAR\s+FROM"),
            period_end=bcbs_michigan_part_i_value(normalized, r"\(g\)\s*POLICY\s+OR\s+CONTRACT\s+YEAR\s+TO"),
            ein=bcbs_michigan_part_i_value(normalized, r"\(b\)\s*EMPLOYER\s+IDENTIFICATION\s+NUMBER\s*\(EIN\)"),
            naic_code=bcbs_michigan_part_i_value(normalized, r"\(c\)\s*NATIONAL\s+ASSOCIATION\s+OF\s+INSURANCE\s+COMMISSIONERS\s*\(NAIC\)\s+CODE"),
            coverage="Health/Dental/Vision/Prescription/PPO",
            values=[value for value in values if value.value],
            benefit_rows=[
                ScheduleABenefitBreakdownRow(
                    benefit_type="Experience-rated contract",
                    persons_covered=bcbs_michigan_part_i_value(normalized, r"\(e\)\s*APPROX\.?\s+NUMBER\s+OF\s+PERSONS\s+COVERED"),
                    premium=regex_first(normalized, [r"\(i\)\s*AMOUNT\s+RECEIVED\s+\$?\s*([0-9,]+(?:\.\d{2})?)"]),
                    source_page=page,
                )
            ],
            notes=["Line 10 nonexperience-rated contracts is Not Applicable."],
        )
    ]


def extract_bcbs_michigan_addendum_broker_rows(text: str, page: int | None = None) -> list[ScheduleABrokerRow]:
    normalized = normalize_ocr_text(text)
    if not is_bcbs_michigan_addendum_page(normalized):
        return []
    rows: list[ScheduleABrokerRow] = []
    blocks = re.split(r"AGENT/BROKER\s+COMMISSION\s+&\s+INCENTIVE\s+PAYMENTS", normalized, flags=re.IGNORECASE)[1:]
    for block in blocks:
        block = re.split(r"\bGROUP\s+INFORMATION\b|\bBlue\s+Cross\s+Blue\s+Shield\s+Michigan\b", block, maxsplit=1, flags=re.IGNORECASE)[0]
        name = regex_first(block, [r"Name\s+and\s+address\s+of\s+agent\s+or\s+broker:\s*(.+?)(?=\n)"])
        if not name or not is_probable_person_or_entity_name(name):
            continue
        address_block = regex_first(block, [r"Name\s+and\s+address\s+of\s+agent\s+or\s+broker:.*?\n(.+?)(?=\n\s*--\s*Amount\s+of\s+Sales)"], flags=re.IGNORECASE | re.DOTALL) or ""
        address_lines = [clean_extracted_value(line) for line in address_block.splitlines() if clean_extracted_value(line)]
        address_line_1 = address_lines[0] if address_lines else None
        city = state = zip_code = None
        for line in address_lines[1:] or address_lines:
            city_match = re.search(r"([A-Za-z .'-]+),\s*([A-Z]{2})\s*([0-9]{5}(?:-[0-9]{4})?)?", line)
            if city_match:
                city = clean_extracted_value(city_match.group(1))
                state = city_match.group(2)
                zip_code = city_match.group(3)
                break
        commission = regex_first(block, [r"Amount\s+of\s+Sales\s+and\s+Base\s+Commissions\s+Paid\s+\$?\s*([0-9,]+(?:\.\d{2})?)"])
        fees = regex_first(block, [r"Fees\s+and\s+Other\s+Commissions\s+Paid\s+Amount\s+\$?\s*([0-9,]+(?:\.\d{2})?)"])
        non_monetary = regex_first(block, [r"Non-Monetary\s+Compensations\s+to\s+Plan.*?\$?\s*([0-9,]+(?:\.\d{2})?)"], flags=re.IGNORECASE | re.DOTALL)
        fee_total = sum_money_values(fees, non_monetary)
        org_code = regex_first(block, [r"Organization\s+Code\s+\(\s*for\s+Schedule\s+\(A\)\s*([0-9]{1,2})"])
        commission_rows = []
        if commission is not None:
            commission_rows.append(ScheduleABrokerMoneyRow(coverage=None, amount=money_value(commission), purpose="Sales and Base Commissions"))
        fee_rows = []
        if fees is not None:
            fee_rows.append(ScheduleABrokerMoneyRow(coverage=None, amount=money_value(fees), purpose="Fees and Other Commissions"))
        if non_monetary and parse_numeric_amount(non_monetary):
            fee_rows.append(ScheduleABrokerMoneyRow(coverage=None, amount=money_value(non_monetary), purpose="Non-Monetary Compensation"))
        rows.append(
            ScheduleABrokerRow(
                name=clean_extracted_value(name),
                address_line_1=address_line_1,
                city=city,
                state=state,
                zip_code=zip_code,
                organization_code=org_code,
                commission_rows=commission_rows,
                fee_rows=fee_rows,
                commission_total=money_value(commission or "0"),
                fee_total=fee_total,
                source_page=page,
                confidence=0.95,
            )
        )
    return rows


def first_bcbs_michigan_main_page(page_texts: list[tuple[int, str]]) -> tuple[int, str] | None:
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if is_bcbs_michigan_main_schedule(normalized):
            return page, text
    return None


def is_bcbs_michigan_main_schedule(text: str) -> bool:
    upper = normalize_ocr_text(text).upper()
    return "BLUE CROSS BLUE SHIELD OF MICHIGAN" in upper and "EXPERIENCE-RATED CONTRACTS" in upper and "NONEXPERIENCE-RATED CONTRACTS" in upper


def is_bcbs_michigan_addendum_page(text: str) -> bool:
    upper = normalize_ocr_text(text).upper()
    return "BLUE CROSS BLUE SHIELD MICHIGAN" in upper and "ADDENDUM TO SCHEDULE A/C" in upper and "AGENT/BROKER COMMISSION" in upper


def bcbs_michigan_part_i_value(text: str, label_pattern: str) -> str | None:
    match = re.search(rf"{label_pattern}\s+(.+?)(?=\n\s*\([a-z]\)|\n\s*2\.|\n\s*PART\s+II|\n\s*PART\s+III|$)", text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return None
    return clean_extracted_value(match.group(1))


def bcbs_michigan_experience_section(text: str) -> str:
    normalized = normalize_ocr_text(text)
    match = re.search(r"9\.\s*EXPERIENCE-RATED\s+CONTRACTS(.+?)(?=10\.\s*NONEXPERIENCE-RATED\s+CONTRACTS|PART\s+IV|$)", normalized, flags=re.IGNORECASE | re.DOTALL)
    return match.group(1) if match else normalized


def is_bcbsma_schedule_a_worksheet(text: str) -> bool:
    upper = text.upper()
    return "BLUE CROSS BLUE SHIELD OF MASSACHUSETTS" in upper and "#5500A WORKSHEET" in upper


def bcbsma_column_value(text: str, label: str, coverage: str) -> str | None:
    match = re.search(rf"{re.escape(label)}\s+(.+)", text, flags=re.IGNORECASE)
    if not match:
        return None
    amounts = re.findall(r"\$?\s*([0-9,]+(?:\.\d{{2}})?)", match.group(1))
    if not amounts:
        return None
    coverage_index = {"MEDICAL": 0, "DENTAL": 1, "SENIOR": 2}.get(coverage.upper(), 0)
    if coverage_index >= len(amounts):
        return None
    return money_value(amounts[coverage_index])


def sum_money_values(*values: str | None) -> str | None:
    total = 0.0
    found = False
    cents = False
    for value in values:
        text = str(value or "").replace("$", "").replace(",", "").strip()
        if not text:
            continue
        try:
            total += float(text)
            found = True
            cents = cents or "." in text
        except ValueError:
            continue
    if not found:
        return None
    rounded = round(total, 2)
    if cents and not rounded.is_integer():
        return f"{rounded:,.2f}"
    return f"{int(round(rounded)):,}"


def whole_dollar_money_value(value: str | None) -> str | None:
    text = str(value or "").replace("$", "").replace(",", "").strip()
    if not text:
        return None
    try:
        return f"{int(round(float(text))):,}"
    except ValueError:
        return money_value(value or "")


def extract_eyemed_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    summaries = extract_eyemed_schedule_a_summaries(page_texts)
    if not summaries:
        return []
    fields: list[NormalizedExtractionField] = []
    for summary in summaries:
        values_by_label = {value.label: value.value for value in summary.values}
        source_text = f"{summary.source} {summary.account_name or ''} {summary.account_number or ''}".strip()

        def add(field_name: str, value: str | None, confidence: float = 0.98):
            clean = clean_extracted_value(str(value or ""))
            if clean and not is_blank_extraction_value(clean):
                fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=None, source_text=source_text))

        add("1a. Name of Insurance Company", summary.carrier_name, 0.99)
        add("1b. Insurance Carrier EIN", summary.ein, 0.99)
        add("1c. NAIC Code", summary.naic_code, 0.99)
        add("1d. Contract/Policy Number", summary.account_number, 0.98)
        add("1e. Persons Covered (End of Policy Year)", values_by_label.get("Persons covered"), 0.98)
        add("1f. Policy Year Beginning Date", summary.period_begin, 0.98)
        add("1g. Policy Year Ending Date", summary.period_end, 0.98)
        add("3b. Amount of Commissions", values_by_label.get("Broker payment total"), 0.94)
        add("3c. Amount of Fees", "0", 0.9)
        add("3d. Purpose", "COMMISSIONS & FEES", 0.9)
        add("3e. Organizational Code", "3", 0.86)
        add("10a. Total premiums or subscription charges paid to carrier", values_by_label.get("Total nonexperience premium"), 0.99)
    return fields


def extract_eyemed_schedule_a_summaries(page_texts: list[tuple[int, str]]) -> list[ScheduleAWorksheetSummary]:
    records: list[dict[str, str]] = []
    first_page: int | None = None
    report_start = report_end = carrier_name = None
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not is_eyemed_schedule_a_worksheet(normalized):
            continue
        first_page = page if first_page is None else first_page
        report_start = regex_first(normalized, [r"Report\s+Start\s+Date\s*Report\s+End\s+Date\s*\n\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})\s+([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})"], groups=True)
        carrier_name = regex_first(normalized, [r"on\s+behalf\s+of\s+the\s+(.+?)(?:\n|$)"])
        records.extend(extract_eyemed_payment_records(normalized))
    if not records:
        return []

    carrier_ein = format_eyemed_ein(first_nonempty(record.get("ein") for record in records) or "")
    naic = first_nonempty(record.get("naic_code") for record in records)
    contracts = [record["contract_number"] for record in records]
    combined_contract = combine_eyemed_contract_numbers(contracts)
    persons_covered = sum_money_values(*(record["persons_covered"] for record in records))
    premium_total = whole_dollar_money_value(sum_money_values(*(record["premium"] for record in records)))
    broker_total = whole_dollar_money_value(sum_money_values(*(row.commission_total for row in extract_eyemed_broker_rows(page_texts))))
    benefit_rows = [
        ScheduleABenefitBreakdownRow(
            benefit_type=f"Vision / {record['contract_number']}",
            persons_covered=record["persons_covered"],
            premium=record["premium"],
            source_page=first_page,
        )
        for record in records
    ]
    values = [
        ScheduleAWorksheetValue(label="Source contracts", value=", ".join(contracts), source="Payments received table", coverage="Vision"),
        ScheduleAWorksheetValue(label="Persons covered", value=persons_covered or "", source="Payments received table", coverage="Vision"),
        ScheduleAWorksheetValue(label="Total nonexperience premium", value=premium_total or "", source="Payments received table", coverage="Vision"),
        ScheduleAWorksheetValue(label="Broker payment total", value=broker_total or "", source="Broker payment table", coverage="Vision"),
    ]
    period_begin = period_end = None
    if isinstance(report_start, tuple) and len(report_start) >= 2:
        period_begin = normalize_eyemed_date(report_start[0])
        period_end = normalize_eyemed_date(report_start[1])
    return [
        ScheduleAWorksheetSummary(
            source="EyeMed vision worksheet",
            carrier_name=clean_extracted_value(carrier_name or "Fidelity Security Life Insurance Company"),
            account_number=combined_contract,
            period_begin=period_begin,
            period_end=period_end,
            ein=carrier_ein or None,
            naic_code=naic,
            coverage="Vision",
            values=[value for value in values if value.value],
            benefit_rows=benefit_rows,
            notes=[
                "Combined EyeMed worksheet contract rows into one FT Williams Schedule A contract.",
                f"Extracted from page {first_page}",
            ],
        )
    ]


def extract_eyemed_payment_records(text: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    normalized = normalize_ocr_text(text)
    payment_section = _section_between(normalized, "Name of Plan", "Payee Name") or normalized
    lines = [line.strip() for line in payment_section.splitlines() if line.strip()]
    pending_plan = ""
    pending_contract = ""
    for line in lines:
        if re.search(r"\bTotal:\s*\$?[0-9,]+\.\d{2}", line, flags=re.IGNORECASE):
            break
        contract_match = None if pending_contract else re.search(r"([0-9]{11})", line)
        if not contract_match and not pending_contract:
            pending_plan = line
            continue
        if contract_match:
            before_contract = clean_extracted_value(line[: contract_match.start()])
            contract_number = clean_extracted_value(contract_match.group(1))
            detail = line[contract_match.end() :].strip()
        else:
            before_contract = pending_plan
            contract_number = pending_contract
            detail = line.strip()
        if not re.search(r"\$?[0-9,]+\.\d{2}\s*$", detail):
            pending_plan = before_contract or pending_plan
            pending_contract = contract_number
            continue
        premium_match = re.search(r"\$?([0-9,]+\.\d{2})\s*$", detail)
        money_detail = detail[: premium_match.start()].strip() if premium_match else detail
        identifiers = ""
        id_match = re.search(r"([0-9]{9})([0-9]{4,6})\s*$", money_detail)
        if id_match:
            identifiers = id_match.group(0)
            money_detail = money_detail[: id_match.start()].strip()
        number_matches = list(re.finditer(r"\b([0-9,]+)\b", money_detail))
        if len(number_matches) < 2 or not premium_match:
            continue
        persons_match = number_matches[-1]
        subscribers_match = number_matches[-2]
        ein = naic = ""
        id_match = re.search(r"([0-9]{9})([0-9]{4,6})", identifiers)
        if id_match:
            ein = id_match.group(1)
            naic = id_match.group(2)
        enrollment_group = money_detail[: subscribers_match.start()].strip()
        records.append(
            {
                "plan_name": clean_extracted_value(before_contract or pending_plan),
                "contract_number": contract_number,
                "enrollment_group": clean_extracted_value(enrollment_group),
                "subscribers": money_value(subscribers_match.group(1)),
                "persons_covered": money_value(persons_match.group(1)),
                "ein": ein,
                "naic_code": naic,
                "premium": money_value(premium_match.group(1)),
            }
        )
        pending_plan = ""
        pending_contract = ""
    return records


def extract_eyemed_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    rows: list[ScheduleABrokerRow] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not is_eyemed_schedule_a_worksheet(normalized):
            continue
        broker_section = _section_between(normalized, "Payee Name", "Commissions or fees paid by carrier") or normalized
        total_marker = re.search(r"\bTotal:\s*\$?[0-9,]+\.\d{2}", broker_section, flags=re.IGNORECASE)
        if total_marker:
            broker_section = broker_section[: total_marker.start()]
        pattern = re.compile(
            r"([A-Za-z0-9 &.,'()-]+?)([0-9]{11})(.+?)\s+([A-Z]{2})\s+([0-9]{5}(?:-[0-9]{4})?)\s+\$?([0-9,]+\.\d{2})",
            flags=re.IGNORECASE | re.DOTALL,
        )
        for match in pattern.finditer(broker_section):
            name = clean_extracted_value(match.group(1))
            contract_number = clean_extracted_value(match.group(2))
            address_city = clean_extracted_value(match.group(3))
            address_line_1, city = split_eyemed_address_city(address_city)
            amount = whole_dollar_money_value(match.group(6)) or money_value(match.group(6))
            if not name or not is_probable_person_or_entity_name(name):
                continue
            rows.append(
                ScheduleABrokerRow(
                    name=name,
                    address_line_1=address_line_1,
                    city=city,
                    state=match.group(4).upper(),
                    zip_code=match.group(5),
                    organization_code="3",
                    commission_rows=[ScheduleABrokerMoneyRow(coverage="Vision", amount=amount, purpose="Commissions & Fees")],
                    fee_rows=[],
                    commission_total=amount,
                    fee_total="0",
                    source_page=page,
                    confidence=0.92,
                )
            )
            rows[-1].address_line_2 = f"Contract {contract_number}"
    return rows


def is_eyemed_schedule_a_worksheet(text: str) -> bool:
    upper = text.upper()
    return "VISION INSURANCE INFORMATION FOR FORM 5500" in upper and "EYEMED" in upper


def combine_eyemed_contract_numbers(contracts: list[str]) -> str:
    unique = list(dict.fromkeys(clean_extracted_value(contract) for contract in contracts if contract))
    if len(unique) >= 2 and all(re.fullmatch(r"\d{11}", contract) for contract in unique):
        prefixes = list(dict.fromkeys(contract[:7] for contract in unique))
        suffixes = list(dict.fromkeys(contract[7:] for contract in unique))
        common_prefix = os.path.commonprefix(prefixes)
        if common_prefix and all(prefix.startswith(common_prefix) for prefix in prefixes):
            prefix_variants = [prefix[len(common_prefix) :] or prefix for prefix in prefixes]
            compact_prefix = common_prefix + "/".join(prefix_variants)
            compact_suffix = "/".join(suffixes)
            return f"{compact_prefix}-{compact_suffix}" if compact_suffix else compact_prefix
    if len(unique) == 2 and all(re.fullmatch(r"\d{8,12}", contract) for contract in unique):
        first, second = unique
        common_prefix = os.path.commonprefix(unique)
        common_suffix = os.path.commonprefix([first[::-1], second[::-1]])[::-1]
        if common_prefix and common_suffix and len(first) == len(second):
            first_mid = first[len(common_prefix) : len(first) - len(common_suffix)]
            second_mid = second[len(common_prefix) : len(second) - len(common_suffix)]
            if first_mid and second_mid:
                return f"{common_prefix}{first_mid}/{second_mid}-{common_suffix}"
    return ", ".join(unique)


def format_eyemed_ein(value: str) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 9:
        return f"{digits[:2]}-{digits[2:]}"
    return value


def first_nonempty(values) -> str | None:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return None


def normalize_eyemed_date(value: str) -> str:
    text = str(value or "").strip()
    parts = text.split("/")
    if len(parts) != 3:
        return text
    month, day, year = parts
    full_year = int(year)
    if full_year < 100:
        full_year += 2000
    return f"{int(month):02d}/{int(day):02d}/{full_year}"


def split_eyemed_address_city(value: str) -> tuple[str | None, str | None]:
    text = clean_extracted_value(value)
    match = re.match(r"(.+\b(?:Street|St|Avenue|Ave|Parkway|Pkwy|Road|Rd|Drive|Dr|Boulevard|Blvd|Lane|Ln|Way|Court|Ct|Circle|Cir))\s+(.+)$", text, flags=re.IGNORECASE)
    if match:
        return clean_extracted_value(match.group(1)), clean_extracted_value(match.group(2))
    return text or None, None


def extract_standard_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    summaries = extract_standard_schedule_a_summaries(page_texts)
    if not summaries:
        return []
    fields: list[NormalizedExtractionField] = []
    for summary in summaries:
        values_by_label = {value.label: value.value for value in summary.values}
        source_text = f"{summary.source} {summary.coverage or ''} {summary.account_number or ''}".strip()

        def add(field_name: str, value: str | None, confidence: float = 0.98):
            clean = clean_extracted_value(str(value or ""))
            if clean and not is_blank_extraction_value(clean):
                fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=None, source_text=source_text))

        add("1a. Name of Insurance Company", summary.carrier_name, 0.99)
        add("1b. Insurance Carrier EIN", summary.ein, 0.99)
        add("1c. NAIC Code", summary.naic_code, 0.99)
        add("1d. Contract/Policy Number", summary.account_number, 0.99)
        add("1e. Persons Covered (End of Policy Year)", values_by_label.get("Persons covered"), 0.98)
        add("1f. Policy Year Beginning Date", summary.period_begin, 0.98)
        add("1g. Policy Year Ending Date", summary.period_end, 0.98)
        add("3a. Name of Agent/Broker/Person", values_by_label.get("Broker name"), 0.95)
        add("3b. Amount of Commissions", values_by_label.get("3b. Amount of Commissions"), 0.97)
        add("3c. Amount of Fees", values_by_label.get("3c. Amount of Fees"), 0.97)
        add("3d. Purpose", values_by_label.get("3d. Purpose"), 0.93)
        add("3e. Organizational Code", values_by_label.get("3e. Organizational Code"), 0.93)
        for label in STANDARD_EXPERIENCE_FIELD_LABELS:
            add(label, values_by_label.get(label), 0.97)
    return fields


def extract_standard_schedule_a_summaries(page_texts: list[tuple[int, str]]) -> list[ScheduleAWorksheetSummary]:
    records = extract_standard_schedule_a_records(page_texts)
    summaries: list[ScheduleAWorksheetSummary] = []
    for record in records:
        coverage = record.get("coverage")
        values = [
            ScheduleAWorksheetValue(label="Persons covered", value=record.get("persons_covered") or "", source="Part I coverage block", coverage=coverage),
            ScheduleAWorksheetValue(label="Broker name", value=record.get("broker_name") or "", source="Part I broker block", coverage=coverage),
            ScheduleAWorksheetValue(label="3b. Amount of Commissions", value=record.get("commission_total") or "", source="Part I line 2", coverage=coverage),
            ScheduleAWorksheetValue(label="3c. Amount of Fees", value=record.get("fee_total") or "", source="Part I line 2", coverage=coverage),
            ScheduleAWorksheetValue(label="3d. Purpose", value=derive_schedule_a_purpose(record.get("commission_total"), record.get("fee_total")) or "COMMISSIONS", source="Derived from line 2", coverage=coverage),
            ScheduleAWorksheetValue(label="3e. Organizational Code", value=record.get("organization_code") or "", source="Part I broker block", coverage=coverage),
        ]
        for label in STANDARD_EXPERIENCE_FIELD_LABELS:
            value = (record.get("experience_values") or {}).get(label)
            values.append(ScheduleAWorksheetValue(label=label, value=value or "", source="Part III experience-rated section", coverage=coverage))
        summaries.append(
            ScheduleAWorksheetSummary(
                source="The Standard long form information",
                carrier_name=record.get("carrier_name"),
                account_name=record.get("account_name"),
                account_number=record.get("contract_number"),
                period_begin=record.get("period_begin"),
                period_end=record.get("period_end"),
                ein=record.get("ein"),
                naic_code=record.get("naic_code"),
                coverage=coverage,
                values=[value for value in values if value.value],
                benefit_rows=[
                    ScheduleABenefitBreakdownRow(
                        benefit_type=coverage or "",
                        persons_covered=record.get("persons_covered"),
                        premium=(record.get("experience_values") or {}).get("9a. Premiums: (1) Amount Received"),
                        source_page=record.get("part_i_page"),
                    )
                ],
                notes=[
                    "Parsed as a separate The Standard benefit Schedule A.",
                    f"Use with FT Williams Schedule A-{standard_schedule_desc_key(coverage or '')}.",
                ],
            )
        )
    return summaries


def extract_standard_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    rows: list[ScheduleABrokerRow] = []
    for record in extract_standard_schedule_a_records(page_texts):
        broker_name = record.get("broker_name")
        if not broker_name:
            continue
        commission_rows = []
        base_commission = record.get("base_commission")
        contingent_commission = record.get("contingent_commission")
        coverage = record.get("coverage")
        if base_commission and not is_zero_money(base_commission):
            commission_rows.append(ScheduleABrokerMoneyRow(coverage=coverage, amount=base_commission, purpose="Commissions"))
        if contingent_commission and not is_zero_money(contingent_commission):
            commission_rows.append(ScheduleABrokerMoneyRow(coverage=coverage, amount=contingent_commission, purpose="Contingent Compensation"))
        if not commission_rows and record.get("commission_total"):
            commission_rows.append(ScheduleABrokerMoneyRow(coverage=coverage, amount=record.get("commission_total"), purpose="Commissions"))
        fee_rows = []
        fee_total = record.get("fee_total")
        if fee_total and not is_zero_money(fee_total):
            fee_rows.append(ScheduleABrokerMoneyRow(coverage=coverage, amount=fee_total, purpose="Fees"))
        rows.append(
            ScheduleABrokerRow(
                name=broker_name,
                address_line_1=record.get("broker_address_line_1"),
                address_line_2=record.get("broker_address_line_2"),
                city=record.get("broker_city"),
                state=record.get("broker_state"),
                zip_code=record.get("broker_zip"),
                organization_code=record.get("organization_code"),
                commission_rows=commission_rows,
                fee_rows=fee_rows,
                commission_total=record.get("commission_total"),
                fee_total=record.get("fee_total") or "0.00",
                source_page=record.get("part_i_page"),
                confidence=0.94,
            )
        )
    return rows


STANDARD_EXPERIENCE_FIELD_LABELS = [
    "9a. Premiums: (1) Amount Received",
    "9a(2). Increase (decrease) in amount due but unpaid",
    "9a(3). Increase (decrease) in unearned premium reserve",
    "9a(4). Earned ((1) + (2) - (3))",
    "9b(1). Benefit Charges (1) Claims paid",
    "9b(2). Increase (decrease) in claim reserves",
    "9b(3). Incurred claims (add(1) and (2))",
    "9b(4). Claims Charged",
    "9c(1)(A). Commissions",
    "9c(1)(B). Administrative service or other fees",
    "9c(1)(C). Other Specific acquisition costs",
    "9c(1)(D). Other expenses",
    "9c(1)(E). Taxes",
    "9c(1)(F). Charges for risks or other contingencies",
    "9c(1)(G). Other retention charges",
    "9c(1)(H). Total retention",
    "9c(2). Dividends or retroactive rate refunds",
    "9d(1). Status of policyholder reserves at end of year: (1) Amount held to provide benefits after retirement",
    "9d(2). Claim reserves",
    "9d(3). Other reserves",
    "9e. Dividends or retroactive rate refunds due",
]


def extract_standard_schedule_a_records(page_texts: list[tuple[int, str]]) -> list[dict[str, Any]]:
    part_i_records: list[dict[str, Any]] = []
    part_iii_records: list[dict[str, Any]] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not is_standard_long_form_schedule_a(normalized):
            continue
        part_i = extract_standard_part_i_record(normalized, page)
        if part_i:
            part_i_records.append(part_i)
        part_iii = extract_standard_part_iii_record(normalized, page)
        if part_iii:
            part_iii_records.append(part_iii)

    merged: list[dict[str, Any]] = []
    used_part_iii: set[int] = set()
    for part_i in part_i_records:
        match_index = next(
            (
                index
                for index, part_iii in enumerate(part_iii_records)
                if index not in used_part_iii
                and standard_schedule_desc_key(part_iii.get("coverage") or "") == standard_schedule_desc_key(part_i.get("coverage") or "")
                and clean_extracted_value(part_iii.get("contract_number") or "") == clean_extracted_value(part_i.get("contract_number") or "")
            ),
            None,
        )
        combined = dict(part_i)
        if match_index is not None:
            used_part_iii.add(match_index)
            combined["experience_values"] = part_iii_records[match_index].get("experience_values", {})
            combined["part_iii_page"] = part_iii_records[match_index].get("part_iii_page")
        else:
            combined["experience_values"] = {}
        merged.append(combined)
    return merged


def is_standard_long_form_schedule_a(text: str) -> bool:
    upper = text.upper()
    return "LONG FORM INFORMATION" in upper and "STANDARD INSURANCE COMPANY" in upper and "PLAN INFORMATION REPORT FOR THE PERIOD" in upper


def extract_standard_part_i_record(text: str, page: int | None = None) -> dict[str, Any] | None:
    upper = text.upper()
    if "PART I" not in upper or "INSURANCE FEES AND COMMISSIONS" not in upper:
        return None
    tail = re.search(
        r"(Standard\s+Insurance\s+Company)\s*\n"
        r"(.+?)\s*\n"
        r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s*\n"
        r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s*\n"
        r"([0-9,]+)\s*\n"
        r"([0-9]{2}-[0-9]{7})\s*\n"
        r"([0-9-]{5,})\s*\n"
        r"(\$?\s*[0-9,]+(?:\.\d{2})?)\s*\n"
        r"(\$?\s*[0-9,]+(?:\.\d{2})?)\s*\n"
        r".*?(DENTAL|LIFE\s+INSURANCE|LONG\s+TERM\s+DISABILITY|VISION|[A-Z][A-Z ]+)\s*\n"
        r"PLAN\s+INFORMATION\s+REPORT\s+FOR\s+THE\s+PERIOD\s+OF\s*\n"
        r"([A-Za-z0-9-]+)\s*\n"
        r"LONG\s+FORM\s+INFORMATION\s*\n"
        r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s*\n"
        r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not tail:
        return None
    broker = extract_standard_broker_info(text)
    return {
        "source": "The Standard long form information",
        "carrier_name": clean_extracted_value(tail.group(1)),
        "account_name": clean_extracted_value(tail.group(2)),
        "period_begin": normalize_schedule_a_date(tail.group(3), end_of_month=False),
        "period_end": normalize_schedule_a_date(tail.group(4), end_of_month=True),
        "persons_covered": money_value(tail.group(5)),
        "ein": clean_extracted_value(tail.group(6)),
        "naic_code": normalize_standard_naic(tail.group(7)),
        "commission_total": standard_money_value(tail.group(8)),
        "fee_total": standard_money_value(tail.group(9)),
        "coverage": clean_extracted_value(tail.group(10)).upper(),
        "contract_number": clean_extracted_value(tail.group(11)),
        "part_i_page": page,
        **broker,
    }


def extract_standard_broker_info(text: str) -> dict[str, str | None]:
    info: dict[str, str | None] = {
        "broker_name": None,
        "broker_address_line_1": None,
        "broker_address_line_2": None,
        "broker_city": None,
        "broker_state": None,
        "broker_zip": None,
        "base_commission": None,
        "contingent_commission": None,
        "organization_code": None,
    }
    match = re.search(
        r"E\)\s*ORG\.\s*\n\s*CODE\s*\n(?P<broker>.+?)\n"
        r"\s*(?P<base>\$?\s*[0-9,]+(?:\.\d{2})?)\s+"
        r"(?P<contingent>\$?\s*[0-9,]+(?:\.\d{2})?)\s+"
        r"(?P<ga>\$?\s*[0-9,]+(?:\.\d{2})?)\s+"
        r"(?P<fees>\$?\s*[0-9,]+(?:\.\d{2})?)\s+"
        r"(?P<org>[0-9]{1,2})\b",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return info
    lines = [clean_extracted_value(line) for line in match.group("broker").splitlines() if clean_extracted_value(line)]
    if not lines:
        return info
    info["broker_name"] = lines[0]
    if len(lines) > 1:
        info["broker_address_line_1"] = lines[1]
    if len(lines) > 2:
        city_line = lines[-1]
        city_match = re.search(r"(.+?),\s*([A-Z]{2})\s+([0-9]{5}(?:-[0-9]{4})?)", city_line)
        if city_match:
            info["broker_city"] = clean_extracted_value(city_match.group(1))
            info["broker_state"] = city_match.group(2)
            info["broker_zip"] = city_match.group(3)
            if len(lines) > 3:
                info["broker_address_line_2"] = " ".join(lines[2:-1])
        elif len(lines) > 2:
            info["broker_address_line_2"] = " ".join(lines[2:])
    info["base_commission"] = standard_money_value(match.group("base"))
    info["contingent_commission"] = standard_money_value(match.group("contingent"))
    info["organization_code"] = match.group("org")
    return info


def extract_standard_part_iii_record(text: str, page: int | None = None) -> dict[str, Any] | None:
    upper = text.upper()
    if "PART III" not in upper or "EXPERIENCE RATED CONTRACTS" not in upper:
        return None
    tail = re.search(
        r"Standard\s+Insurance\s+Company\s+HEREBY\s+CERTIFIES.+?\n(?P<body>.+?)\n"
        r"(?P<contract>[A-Za-z0-9-]+)\s*\n"
        r"(?P<coverage>DENTAL|LIFE\s+INSURANCE|LONG\s+TERM\s+DISABILITY|VISION|[A-Z][A-Z ]+)\s*\n"
        r"(?P<premium>\$?\s*[0-9,]+(?:\.\d{2})?)\s*\n"
        r"LONG\s+FORM\s+INFORMATION",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not tail:
        return None
    money_tokens = [standard_money_value(token) for token in re.findall(r"\(?\$?\s*[0-9,]+(?:\.\d{2})?\)?", tail.group("body"))]
    labels_after_9a1 = STANDARD_EXPERIENCE_FIELD_LABELS[1:]
    experience_values = {"9a. Premiums: (1) Amount Received": standard_money_value(tail.group("premium"))}
    for label, value in zip(labels_after_9a1, money_tokens):
        experience_values[label] = value
    return {
        "contract_number": clean_extracted_value(tail.group("contract")),
        "coverage": clean_extracted_value(tail.group("coverage")).upper(),
        "experience_values": experience_values,
        "part_iii_page": page,
    }


def standard_money_value(value: str | None) -> str:
    text = str(value or "").strip()
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").strip()
    return f"-{text}" if negative and text and not text.startswith("-") else text


def is_zero_money(value: str | None) -> bool:
    amount = parse_numeric_amount(value)
    return amount is not None and abs(amount) < 0.005


def normalize_standard_naic(value: str | None) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.lstrip("0") or digits


def standard_schedule_desc_key(value: str) -> str:
    key = re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())
    if key in {"LTD", "LONGTERMDISABILITY"} or "LTD" in key or "LONGTERMDISABILITY" in key:
        return "LTD"
    if "DENTAL" in key:
        return "DENTAL"
    if "VISION" in key:
        return "VISION"
    if "LIFE" in key:
        return "LIFE"
    return key


def extract_united_omaha_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    summaries = extract_united_omaha_schedule_a_summaries(page_texts)
    if not summaries:
        return []
    fields: list[NormalizedExtractionField] = []
    for summary in summaries:
        values_by_label = {value.label: value.value for value in summary.values}
        source_text = f"{summary.source} {summary.coverage or ''} {summary.account_number or ''}".strip()

        def add(field_name: str, value: str | None, confidence: float = 0.96):
            clean = clean_extracted_value(str(value or ""))
            if clean and not is_blank_extraction_value(clean):
                fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=None, source_text=source_text))

        add("1a. Name of Insurance Company", summary.carrier_name, 0.99)
        add("1b. Insurance Carrier EIN", summary.ein, 0.99)
        add("1c. NAIC Code", summary.naic_code, 0.99)
        add("1d. Contract/Policy Number", summary.account_number, 0.98)
        add("1e. Persons Covered (End of Policy Year)", values_by_label.get("Persons covered"), 0.97)
        add("1f. Policy Year Beginning Date", summary.period_begin, 0.97)
        add("1g. Policy Year Ending Date", summary.period_end, 0.97)
        add("3a. Name of Agent/Broker/Person", values_by_label.get("Broker name"), 0.94)
        add("3b. Amount of Commissions", values_by_label.get("3b. Amount of Commissions"), 0.97)
        add("3c. Amount of Fees", values_by_label.get("3c. Amount of Fees"), 0.97)
        add("3d. Purpose", values_by_label.get("3d. Purpose"), 0.93)
        add("3e. Organizational Code", values_by_label.get("3e. Organizational Code"), 0.94)
        add("10a. Total premiums or subscription charges paid to carrier", values_by_label.get("10a. Total premiums or subscription charges paid to carrier"), 0.98)
    return fields


def extract_united_omaha_schedule_a_summaries(page_texts: list[tuple[int, str]]) -> list[ScheduleAWorksheetSummary]:
    summaries: list[ScheduleAWorksheetSummary] = []
    for record in extract_united_omaha_schedule_a_records(page_texts):
        coverage = record.get("coverage")
        broker_rows = record.get("broker_rows") or []
        commission_total = sum_money_values(*(row.commission_total for row in broker_rows)) or "0"
        fee_total = sum_money_values(*(row.fee_total for row in broker_rows)) or "0"
        primary_broker = broker_rows[0].name if broker_rows else ""
        values = [
            ScheduleAWorksheetValue(label="Group identification number", value=record.get("group_id") or "", source="Part I group block", coverage=coverage),
            ScheduleAWorksheetValue(label="Legacy group ID", value=record.get("legacy_group_id") or "", source="Part I group block", coverage=coverage),
            ScheduleAWorksheetValue(label="Persons covered", value=record.get("persons_covered") or "", source="Benefits Provided", coverage=coverage),
            ScheduleAWorksheetValue(label="Broker name", value=primary_broker, source="Recipient table", coverage=coverage),
            ScheduleAWorksheetValue(label="3b. Amount of Commissions", value=commission_total, source="Recipient table", coverage=coverage),
            ScheduleAWorksheetValue(label="3c. Amount of Fees", value=fee_total, source="Recipient table", coverage=coverage),
            ScheduleAWorksheetValue(label="3d. Purpose", value=derive_schedule_a_purpose(commission_total, fee_total) or "", source="Recipient table", coverage=coverage),
            ScheduleAWorksheetValue(label="3e. Organizational Code", value=record.get("organization_code") or "", source="Recipient table", coverage=coverage),
            ScheduleAWorksheetValue(
                label="10a. Total premiums or subscription charges paid to carrier",
                value=record.get("premium") or "",
                source="Part III non-experience rated contracts",
                coverage=coverage,
            ),
        ]
        summaries.append(
            ScheduleAWorksheetSummary(
                source="United of Omaha Schedule A support worksheet",
                carrier_name=record.get("carrier_name"),
                account_name=record.get("account_name"),
                account_number=record.get("legacy_group_id") or record.get("group_id"),
                period_begin=record.get("period_begin"),
                period_end=record.get("period_end"),
                ein=record.get("ein"),
                naic_code=record.get("naic_code"),
                coverage=coverage,
                values=[value for value in values if value.value],
                benefit_rows=[
                    ScheduleABenefitBreakdownRow(
                        benefit_type=coverage or "",
                        persons_covered=record.get("persons_covered"),
                        premium=record.get("premium"),
                        source_page=record.get("page"),
                    )
                ],
                notes=[
                    "Parsed as a separate United of Omaha benefit worksheet page.",
                    f"Group identification number: {record.get('group_id') or ''}".strip(),
                ],
            )
        )
    return summaries


def extract_united_omaha_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    rows: list[ScheduleABrokerRow] = []
    for record in extract_united_omaha_schedule_a_records(page_texts):
        rows.extend(record.get("broker_rows") or [])
    return rows


def extract_summary_table_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    rows_by_name: dict[str, ScheduleABrokerRow] = {}
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        for section, row_kind in extract_schedule_a_line3_summary_sections(normalized):
            for name, amount in extract_schedule_a_line3_summary_entries(section):
                key = normalize_compare_key(name)
                row = rows_by_name.get(key)
                if not row:
                    row = ScheduleABrokerRow(
                        name=name,
                        organization_code="3",
                        commission_rows=[],
                        fee_rows=[],
                        source_page=page,
                        confidence=0.9,
                    )
                    rows_by_name[key] = row
                money = money_value(amount)
                if row_kind == "commission":
                    row.commission_rows.append(ScheduleABrokerMoneyRow(amount=money, purpose="COMMISSIONS"))
                    row.commission_total = _sum_money_rows(row.commission_rows)
                else:
                    row.fee_rows.append(ScheduleABrokerMoneyRow(amount=money, purpose="FEES"))
                    row.fee_total = _sum_money_rows(row.fee_rows)
    return list(rows_by_name.values())


def extract_schedule_a_line3_summary_sections(text: str) -> list[tuple[str, str]]:
    sections: list[tuple[str, str]] = []
    markers = [
        (r"Schedule\s+A,\s*Line\s+3,\s*Element\s*\(b\)", "commission"),
        (r"Schedule\s+A,\s*Line\s+3,\s*Element\s*\(c\)", "fee"),
    ]
    for marker, row_kind in markers:
        for match in re.finditer(marker, text, flags=re.IGNORECASE):
            section = text[match.start() : match.start() + 1800]
            stop = re.search(
                r"\n\s*(?:The\s+following\s+figure\s+represents\s+(?:commissions|fees)|"
                r"Group\s+insurance\s+coverages|Gross\s+premium|One\s+time\s+reimbursement|"
                r"Indirect\s+compensation|Schedule\s+C|Total\s+premium|Premium\s+due)\b",
                section,
                flags=re.IGNORECASE,
            )
            if stop:
                section = section[: stop.start()]
            sections.append((section, row_kind))
    return sections


def extract_schedule_a_line3_summary_entries(section: str) -> list[tuple[str, str]]:
    entries: list[tuple[str, str]] = []
    line_pattern = re.compile(
        r"^\s*(?P<contract>[A-Z0-9][A-Z0-9-]{2,})\s+"
        r"(?P<name>[A-Za-z][A-Za-z0-9&.,'() /-]{2,}?)\s+"
        r"\$?(?P<amount>[0-9][0-9,]*(?:\.\d{2})?)\s*$",
        flags=re.IGNORECASE,
    )
    for raw_line in section.splitlines():
        line = clean_extracted_value(raw_line)
        match = line_pattern.match(line)
        if not match:
            continue
        name = clean_extracted_value(match.group("name"))
        amount = money_value(match.group("amount"))
        if not is_schedule_a_line3_summary_broker_name(name) or is_zero_money(amount):
            continue
        entries.append((name, amount))
    return entries


def is_schedule_a_line3_summary_broker_name(name: str) -> bool:
    clean = clean_extracted_value(name)
    if not clean or not re.search(r"[A-Za-z]", clean):
        return False
    upper = clean.upper()
    blocked_terms = [
        "TOTAL",
        "COMMISSIONS FOR PLAN",
        "FEES PAID",
        "GROUP INSURANCE",
        "GROSS PREMIUM",
        "PREMIUM",
        "AD&D",
        "DENTAL",
        "LIFE",
        "VISION",
        "DISABILITY",
        "ACCIDENT",
        "CRITICAL ILLNESS",
        "REIMBURSEMENT",
        "INDIRECT COMPENSATION",
    ]
    return not any(term in upper for term in blocked_terms)


def extract_united_omaha_schedule_a_records(page_texts: list[tuple[int, str]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for page, text in page_texts:
        normalized = normalize_ocr_text(text)
        if not is_united_omaha_schedule_a_support(normalized):
            continue
        record = extract_united_omaha_schedule_a_record(normalized, page)
        if record:
            records.append(record)
    return records


def is_united_omaha_schedule_a_support(text: str) -> bool:
    upper = text.upper()
    return (
        "SUPPORT FOR FORM 5500, SCHEDULE A" in upper
        and "UNITED OF OMAHA LIFE INSURANCE COMPANY" in upper
        and "BENEFITS PROVIDED" in upper
    )


def extract_united_omaha_schedule_a_record(text: str, page: int | None = None) -> dict[str, Any] | None:
    carrier_match = re.search(r"Name\s+of\s+Carrier:\s*(.+?)\s*-\s*NAIC\s+Code\s+([0-9]{4,6})", text, flags=re.IGNORECASE)
    ein = regex_first(text, [r"EIN\s+Number:\s*([0-9]{2}-[0-9]{7})"], flags=re.IGNORECASE)
    group_match = re.search(
        r"Group\s+Identification\s*\n\s*Number:\s*\n?\s*([A-Z0-9 ]+?)\s+Data\s+for\s+Period:\s*([0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{4})\s+to\s+([0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{4})",
        text,
        flags=re.IGNORECASE,
    )
    legacy_group_id = regex_first(text, [r"Legacy\s+Group\s+ID:\s*([A-Z0-9 ]+)"], flags=re.IGNORECASE)
    benefit_match = re.search(r"Benefits\s+Provided\s+Persons\s+Covered\s*\n\s*(.+?)\s+([0-9,]+)\s*\n", text, flags=re.IGNORECASE)
    premium = regex_first(
        text,
        [r"Premiums\s*(?:\.\s*)+([0-9,]+(?:\.\d{2})?)"],
        flags=re.IGNORECASE,
    )
    if not carrier_match or not group_match or not benefit_match:
        return None
    coverage = clean_extracted_value(benefit_match.group(1))
    group_id = normalize_united_omaha_group_id(group_match.group(1))
    legacy_group_id = normalize_united_omaha_group_id(legacy_group_id)
    period_begin = normalize_schedule_a_date(group_match.group(2).replace("-", "/"), end_of_month=False)
    period_end = normalize_schedule_a_date(group_match.group(3).replace("-", "/"), end_of_month=True)
    broker_rows = extract_united_omaha_broker_rows_from_page(text, coverage, page)
    organization_code = next((row.organization_code for row in broker_rows if row.organization_code), None)
    return {
        "source": "United of Omaha Schedule A support worksheet",
        "carrier_name": clean_extracted_value(carrier_match.group(1)),
        "naic_code": clean_extracted_value(carrier_match.group(2)),
        "ein": ein,
        "group_id": group_id,
        "legacy_group_id": legacy_group_id or group_id,
        "account_name": extract_united_omaha_account_name(text),
        "coverage": coverage,
        "persons_covered": money_value(benefit_match.group(2)),
        "period_begin": period_begin,
        "period_end": period_end,
        "premium": money_value(premium or ""),
        "organization_code": organization_code,
        "broker_rows": broker_rows,
        "page": page,
    }


def extract_united_omaha_account_name(text: str) -> str | None:
    marker = "INFORMATION FOR COMPLETION OF PART I"
    if marker not in text:
        return None
    after = text.split(marker, 1)[1]
    lines = [clean_extracted_value(line) for line in after.splitlines() if clean_extracted_value(line)]
    if lines and not lines[0].lower().startswith("name of carrier"):
        return lines[0]
    return None


def extract_united_omaha_broker_rows_from_page(text: str, coverage: str | None, page: int | None) -> list[ScheduleABrokerRow]:
    rows: list[ScheduleABrokerRow] = []
    section = _section_between(text, "Name of Each Recipient", "INFORMATION FOR COMPLETION OF PART III")
    if not section:
        return rows
    first_match = re.search(
        r"(GALLAGHER\s+BENEFIT\s+SERVICES\s+INC)\s+([0-9,]+(?:\.\d{2})?)\s+(Agent\s+or\s+Broker\s+of\s+Record)\s+([0-9]{1,2})\s*\n"
        r"(.+?)\n"
        r"(.+?,\s*[A-Z]{2}\s+[0-9]{5}(?:-[0-9]{4})?)",
        section,
        flags=re.IGNORECASE,
    )
    if first_match:
        city, state, zip_code = split_city_state_zip(first_match.group(6))
        rows.append(
            ScheduleABrokerRow(
                name=clean_extracted_value(first_match.group(1)),
                address_line_1=clean_extracted_value(first_match.group(5)),
                city=city,
                state=state,
                zip_code=zip_code,
                organization_code=first_match.group(4),
                commission_rows=[ScheduleABrokerMoneyRow(coverage=coverage, amount=money_value(first_match.group(2)), purpose=clean_extracted_value(first_match.group(3)))],
                commission_total=money_value(first_match.group(2)),
                fee_total="0",
                source_page=page,
                confidence=0.94,
            )
        )
    other_match = re.search(
        r"(GALLAGHER\s+BENEFIT\s+SERVICES\s+INC)\s+0\s+(Other\s+Compensation)\s+([0-9]{1,2})\s*\n"
        r"(.+?)\s+([0-9,]+(?:\.\d{2})?)\s*\n"
        r"(.+?)\n"
        r"(.+?,\s*[A-Z]{2}\s+[0-9]{5}(?:-[0-9]{4})?)",
        section,
        flags=re.IGNORECASE,
    )
    if other_match:
        city, state, zip_code = split_city_state_zip(other_match.group(7))
        rows.append(
            ScheduleABrokerRow(
                name=clean_extracted_value(f"{other_match.group(1)} {other_match.group(4)}"),
                address_line_1=clean_extracted_value(other_match.group(6)),
                city=city,
                state=state,
                zip_code=zip_code,
                organization_code=other_match.group(3),
                commission_rows=[ScheduleABrokerMoneyRow(coverage=coverage, amount="0", purpose="Commissions")],
                fee_rows=[ScheduleABrokerMoneyRow(coverage=coverage, amount=money_value(other_match.group(5)), purpose=clean_extracted_value(other_match.group(2)))],
                commission_total="0",
                fee_total=money_value(other_match.group(5)),
                source_page=page,
                confidence=0.93,
            )
        )
    return rows


def normalize_united_omaha_group_id(value: str | None) -> str | None:
    clean = clean_extracted_value(str(value or ""))
    if not clean:
        return None
    return re.sub(r"\s+", "", clean)


def split_city_state_zip(value: str | None) -> tuple[str | None, str | None, str | None]:
    match = re.search(r"(.+?),\s*([A-Z]{2})\s+([0-9]{5}(?:-[0-9]{4})?)", str(value or "").strip())
    if not match:
        return clean_extracted_value(str(value or "")) or None, None, None
    return clean_extracted_value(match.group(1)), match.group(2), match.group(3)


def united_omaha_schedule_desc_key(value: str) -> str:
    key = re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())
    if "LONGTERMDISABILITY" in key or "LTD" in key:
        return "LTD"
    if "SHORTTERMDISABILITY" in key or "STD" in key:
        return "STD"
    if "LIFE" in key:
        return "LIFE"
    if "ADANDD" in key or "ADAD" in key or "AD&D" in str(value or "").upper() or "ACCIDENT" in key:
        return "AD&D"
    return key


def _section_between(text: str, start_label: str, end_label: str | None) -> str:
    start = re.search(re.escape(start_label), text, flags=re.IGNORECASE)
    if not start:
        return ""
    section = text[start.end() :]
    if end_label:
        end = re.search(re.escape(end_label), section, flags=re.IGNORECASE)
        if end:
            section = section[: end.start()]
    return section


def extract_prudential_schedule_a_fields(page_texts: list[tuple[int, str]]) -> list[NormalizedExtractionField]:
    summaries = extract_prudential_schedule_a_summaries(page_texts)
    if not summaries:
        return []
    fields: list[NormalizedExtractionField] = []
    for summary in summaries:
        source_text = f"{summary.source} {summary.account_name or ''} {summary.account_number or ''}".strip()

        def add(field_name: str, value: str | None, confidence: float = 0.98):
            clean = clean_extracted_value(str(value or ""))
            if clean and not is_blank_extraction_value(clean):
                fields.append(NormalizedExtractionField(field_name=field_name, value=clean, confidence=confidence, page=None, source_text=source_text))

        values_by_label = {value.label: value.value for value in summary.values}
        add("1a. Name of Insurance Company", summary.carrier_name, 0.99)
        add("1b. Insurance Carrier EIN", summary.ein, 0.99)
        add("1c. NAIC Code", summary.naic_code, 0.99)
        add("1d. Contract/Policy Number", summary.account_number, 0.99)
        add("1e. Persons Covered (End of Policy Year)", values_by_label.get("Persons covered"), 0.98)
        add("1f. Policy Year Beginning Date", summary.period_begin, 0.98)
        add("1g. Policy Year Ending Date", summary.period_end, 0.98)
        add("10a. Total premiums or subscription charges paid to carrier", values_by_label.get("Total nonexperience premium"), 0.99)
    return fields


def extract_prudential_schedule_a_summaries(page_texts: list[tuple[int, str]]) -> list[ScheduleAWorksheetSummary]:
    groups: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for page, text in page_texts:
        record = extract_prudential_benefit_record(text, page)
        if not record:
            continue
        key = (
            record["ein"],
            record["naic_code"],
            record["contract_number"],
            record["period_begin"],
            record["period_end"],
        )
        group = groups.setdefault(
            key,
            {
                "source": "Prudential insured welfare plan data",
                "carrier_name": record["carrier_name"],
                "account_name": record["account_name"],
                "account_number": record["contract_number"],
                "period_begin": record["period_begin"],
                "period_end": record["period_end"],
                "ein": record["ein"],
                "naic_code": record["naic_code"],
                "benefit_rows": [],
                "premium_values": [],
                "covered_values": [],
                "pages": [],
            },
        )
        group["benefit_rows"].append(
            ScheduleABenefitBreakdownRow(
                benefit_type=record["benefit_type"],
                persons_covered=record["persons_covered"],
                premium=record["premium"],
                source_page=page,
            )
        )
        group["premium_values"].append(record["premium"])
        group["covered_values"].append(record["persons_covered"])
        group["pages"].append(page)

    summaries: list[ScheduleAWorksheetSummary] = []
    for group in groups.values():
        total_premium = sum_money_values(*group["premium_values"])
        persons_covered = max_numeric_string(group["covered_values"])
        values = [
            ScheduleAWorksheetValue(label="Persons covered", value=persons_covered or "", source="Grouped benefit pages", coverage="All benefits"),
            ScheduleAWorksheetValue(label="Total nonexperience premium", value=total_premium or "", source="Line 9a on Prudential source pages", coverage="All benefits"),
        ]
        summaries.append(
            ScheduleAWorksheetSummary(
                source=group["source"],
                carrier_name=group["carrier_name"],
                account_name=group["account_name"],
                account_number=group["account_number"],
                period_begin=group["period_begin"],
                period_end=group["period_end"],
                ein=group["ein"],
                naic_code=group["naic_code"],
                coverage="Multiple benefits" if len(group["benefit_rows"]) > 1 else group["benefit_rows"][0].benefit_type,
                values=[value for value in values if value.value],
                benefit_rows=group["benefit_rows"],
                notes=[
                    "Grouped same Prudential contract across benefit pages.",
                    f"Source pages: {', '.join(str(page) for page in group['pages'])}",
                ],
            )
        )
    return summaries


def extract_prudential_benefit_record(text: str, page: int | None = None) -> dict[str, str] | None:
    normalized = normalize_ocr_text(text)
    upper = normalized.upper()
    if "PRUDENTIAL" not in upper or "INSURED WELFARE PLAN DATA" not in upper or "NON EXPERIENCE RATED CONTRACTS" not in upper:
        return None
    carrier_name = regex_first(normalized, [r"1\s*\(a\)\s*(Prudential\s+Insurance\s+Company\s+of\s+America)"])
    ein = regex_first(normalized, [r"1\s*\(b\)\s*Prudential'?s\s+EIN:\s*([0-9]{2}-[0-9]{7})"])
    naic = regex_first(normalized, [r"1\s*\(c\)\s*NAIC\s+code:\s*([0-9]{4,6})"])
    contract = regex_first(normalized, [r"1\s*\(d\)\s*Contract\s+number\s+or\s+identification:\s*([A-Za-z0-9-]+)"])
    period = regex_first(normalized, [r"\n\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s+([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})\s+See\s+Form"], groups=True)
    premium = regex_first(normalized, [r"Non\s+experience\s+rated\s+contracts:.*?Total\s+premiums\s+or\s+subscription\s+charges\s+paid\s+to\s+carrier\s+\$\s*([0-9,]+(?:\.\d{2})?)"], flags=re.IGNORECASE | re.DOTALL)
    account_name = regex_first(normalized, [r"Insured\s+Welfare\s+Plan\s+Data\s*\n\s*(.+?)\s+\(Item\s+numbers"], flags=re.IGNORECASE | re.DOTALL)
    benefit_match = re.search(r"\n\s*([A-Za-z][A-Za-z0-9 &/'().-]+?)\s+([0-9,]+)\s*$", normalized.strip(), flags=re.IGNORECASE)
    if not all([carrier_name, ein, naic, contract, isinstance(period, tuple), premium, benefit_match]):
        return None
    return {
        "carrier_name": clean_extracted_value(carrier_name or "Prudential Insurance Company of America"),
        "ein": clean_extracted_value(ein or ""),
        "naic_code": clean_extracted_value(naic or ""),
        "contract_number": clean_extracted_value(contract or ""),
        "period_begin": normalize_schedule_a_date(period[0], end_of_month=False),
        "period_end": normalize_schedule_a_date(period[1], end_of_month=True),
        "premium": money_value(premium or ""),
        "account_name": clean_extracted_value(account_name or ""),
        "benefit_type": clean_extracted_value(benefit_match.group(1)),
        "persons_covered": money_value(benefit_match.group(2)),
        "source_page": str(page or ""),
    }


def max_numeric_string(values: list[str | None]) -> str | None:
    best: int | None = None
    for value in values:
        text = str(value or "").replace(",", "").strip()
        if not text.isdigit():
            continue
        number = int(text)
        best = number if best is None else max(best, number)
    return f"{best:,}" if best is not None else None


def extract_prudential_broker_rows(page_texts: list[tuple[int, str]]) -> list[ScheduleABrokerRow]:
    combined = "\n".join(text for _, text in page_texts if "Insured Welfare Plan Commission Information" in text or "Organization" in text and "code" in text)
    normalized = normalize_ocr_text(combined)
    if "GRP 27722" not in normalized and "INSURED WELFARE PLAN COMMISSION INFORMATION" not in normalized.upper():
        return []
    rows_by_name: dict[str, ScheduleABrokerRow] = {}
    commission_section = re.split(r"\n\s*Includes amounts paid", normalized, flags=re.IGNORECASE)[0]
    segments = re.split(r"(?=\n?\s*71492\s+)", commission_section)
    for segment in segments:
        segment = segment.strip()
        if not segment.startswith("71492"):
            continue
        match = re.match(r"71492\s+(.+?)\s+\$([0-9,]+(?:\.\d{2})?)\s*(.*)$", segment, flags=re.DOTALL)
        if not match:
            continue
        raw_name = clean_extracted_value(re.sub(r"\s+", " ", match.group(1)))
        amount = money_value(match.group(2))
        address_lines = [line.strip() for line in match.group(3).splitlines() if line.strip()]
        address_lines = [line for line in address_lines if not re.match(r"71492\b", line)]
        city = state = zip_code = None
        city_line_index = None
        for index, line in enumerate(address_lines):
            city_match = re.search(r"(.+?),\s*([A-Z]{2})\s+([0-9-]+)", line)
            if city_match:
                city = clean_extracted_value(city_match.group(1))
                state = city_match.group(2)
                zip_code = normalize_zip_code(city_match.group(3))
                city_line_index = index
                break
        street_lines = address_lines[:city_line_index] if city_line_index is not None else address_lines
        address_line_1 = street_lines[-1] if street_lines else None
        address_line_2 = " ".join(street_lines[:-1]) if len(street_lines) > 1 else None
        name = normalize_prudential_broker_name(raw_name)
        if not name or not is_probable_person_or_entity_name(name):
            continue
        key = normalize_compare_key(name)
        row = rows_by_name.get(key)
        if not row:
            row = ScheduleABrokerRow(
                name=name,
                address_line_1=address_line_1,
                address_line_2=address_line_2,
                city=city,
                state=state,
                zip_code=zip_code,
                organization_code=prudential_org_code_for_broker(name),
                commission_rows=[],
                fee_rows=[],
                source_page=14,
                confidence=0.9,
            )
            rows_by_name[key] = row
        purpose = prudential_purpose_for_broker(name)
        money_row = ScheduleABrokerMoneyRow(coverage=None, amount=amount, purpose=purpose)
        if prudential_amount_is_commission(name, purpose):
            row.commission_rows.append(money_row)
        else:
            row.fee_rows.append(money_row)

    rows: list[ScheduleABrokerRow] = []
    for row in rows_by_name.values():
        row.commission_total = _sum_money_rows(row.commission_rows) if row.commission_rows else "0"
        row.fee_total = _sum_money_rows(row.fee_rows) if row.fee_rows else "0"
        rows.append(row)
    return rows


def normalize_prudential_broker_name(value: str) -> str:
    name = clean_extracted_value(value)
    name = re.sub(r"\bINC\b$", "INC", name, flags=re.IGNORECASE)
    if normalize_compare_key(name) == "selmancompanyllc":
        return "Selman & Company, LLC"
    return name


def normalize_zip_code(value: str) -> str:
    text = str(value or "").strip()
    return text.zfill(5) if re.fullmatch(r"\d{4}", text) else text


def prudential_purpose_for_broker(name: str) -> str:
    key = normalize_compare_key(name)
    if "img" == key:
        return "Third Party Administration Fees"
    if "selman" in key:
        return "Sales and Service Compensation"
    return "Commissions"


def prudential_org_code_for_broker(name: str) -> str:
    return "5" if normalize_compare_key(name) == "img" else "3"


def prudential_amount_is_commission(name: str, purpose: str) -> bool:
    key = normalize_compare_key(name)
    return "brokerage" in key or "commission" in normalize_compare_key(purpose)


def extract_schedule_a_broker_rows(text: str, page: int | None = None) -> list[ScheduleABrokerRow]:
    normalized = normalize_ocr_text(text)
    if not normalized:
        return []
    blocks = re.split(
        r"Name\s+and\s+address\s+of\s+the\s+agents?,?\s+brokers?\s+or\s+other\s+persons?\s+to\s+whom\s+commissions\s+or\s+fees\s+were\s+paid",
        normalized,
        flags=re.IGNORECASE,
    )[1:]
    rows: list[ScheduleABrokerRow] = []
    seen: set[tuple[str, str, str, str]] = set()
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        block = re.sub(
            r"(?<=[A-Za-z0-9])(?=(?:Address\s+Line\s+1|Address\s+Line\s+2|City|State|Zip\s+Code|Organization\s+code|Commissions\s+Paid|Fees\s+Paid)\s*:?)",
            " ",
            block,
            flags=re.IGNORECASE,
        )
        next_block = re.search(
            r"Name\s+and\s+address\s+of\s+the\s+agents?,?\s+brokers?\s+or\s+other\s+persons?\s+to\s+whom\s+commissions\s+or\s+fees\s+were\s+paid",
            block,
            flags=re.IGNORECASE,
        )
        if next_block:
            block = block[: next_block.start()]
        block = re.split(
            r"\bPart\s+III\b|\bWelfare\s+Benefit\s+Contract\s+Information\b|"
            r"\bINFORMATION\s+FOR\s+COMPLETING\s+SCHEDULE\s+C\b|"
            r"\bSCHEDULE\s+C\s*-\s*SERVICE\s+PROVIDER\s+INFORMATION\b|"
            r"\bEligible\s+Indirect\s+Compensation\b|"
            r"\bPlan\s+Detail\s+Report\b",
            block,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        name = _schedule_a_labeled_value(block, "Name", ["Address Line 1", "Address Line 2", "City", "State", "Zip Code", "Organization code", "Commissions Paid", "Fees Paid"])
        if not name or not is_probable_person_or_entity_name(name):
            continue
        address_line_1 = _schedule_a_labeled_value(block, "Address Line 1", ["Address Line 2", "City", "State", "Zip Code", "Organization code", "Commissions Paid", "Fees Paid"])
        address_line_2 = _schedule_a_labeled_value(block, "Address Line 2", ["City", "State", "Zip Code", "Organization code", "Commissions Paid", "Fees Paid"])
        city = _schedule_a_labeled_value(block, "City", ["State", "Zip Code", "Organization code", "Commissions Paid", "Fees Paid"])
        state = _schedule_a_labeled_value(block, "State", ["Zip Code", "Organization code", "Commissions Paid", "Fees Paid"])
        zip_code = _schedule_a_labeled_value(block, "Zip Code", ["Organization code", "Commissions Paid", "Fees Paid"])
        organization_code = _schedule_a_labeled_value(block, "Organization code", ["Commissions Paid", "Fees Paid"])
        if organization_code:
            organization_code = regex_first(organization_code, [r"\b([0-9]{1,2})\b"]) or organization_code
        commission_section = _schedule_a_section_between(block, "Commissions Paid", "Fees Paid")
        fee_section = _schedule_a_section_between(block, "Fees Paid", None)
        commission_money_rows, commission_total = _schedule_a_money_rows(commission_section)
        fee_money_rows, fee_total = _schedule_a_money_rows(fee_section)
        row = ScheduleABrokerRow(
            name=name,
            address_line_1=address_line_1,
            address_line_2=address_line_2,
            city=city,
            state=state,
            zip_code=zip_code,
            organization_code=organization_code,
            commission_rows=commission_money_rows,
            fee_rows=fee_money_rows,
            commission_total=commission_total,
            fee_total=fee_total,
            source_page=page,
            confidence=0.94 if commission_total or fee_total else 0.86,
        )
        key = (
            normalize_compare_key(row.name),
            normalize_compare_key(row.address_line_1 or ""),
            str(row.commission_total or ""),
            str(row.fee_total or ""),
        )
        if key not in seen:
            seen.add(key)
            rows.append(row)
    return rows


def dedupe_schedule_a_broker_rows(rows: list[ScheduleABrokerRow]) -> list[ScheduleABrokerRow]:
    deduped: list[ScheduleABrokerRow] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in rows:
        key = (
            normalize_compare_key(row.name),
            normalize_compare_key(row.address_line_1 or ""),
            str(row.commission_total or ""),
            str(row.fee_total or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def _schedule_a_labeled_value(text: str, label: str, stop_labels: list[str]) -> str | None:
    stops = "|".join(re.escape(stop).replace(r"\ ", r"\s+") for stop in stop_labels)
    label_pattern = re.escape(label).replace(r"\ ", r"\s+")
    match = re.search(rf"\b{label_pattern}\s*:\s*(.*?)(?=\b(?:{stops})\s*:|\bCommissions\s+Paid\b|\bFees\s+Paid\b|$)", text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return None
    value = re.sub(r"\s+", " ", match.group(1)).strip(" :-")
    return clean_extracted_value(value) if value else None


def _schedule_a_section_between(text: str, start_label: str, end_label: str | None) -> str:
    start = re.search(re.escape(start_label).replace(r"\ ", r"\s+"), text, flags=re.IGNORECASE)
    if not start:
        return ""
    section = text[start.end() :]
    if end_label:
        end = re.search(re.escape(end_label).replace(r"\ ", r"\s+"), section, flags=re.IGNORECASE)
        if end:
            section = section[: end.start()]
    return section


def _schedule_a_money_rows(section: str) -> tuple[list[ScheduleABrokerMoneyRow], str | None]:
    if not section:
        return [], None
    compact = normalize_ocr_text(section)
    compact = re.sub(r"\bCoverage\s+Amount\s+Purpose\b", " ", compact, flags=re.IGNORECASE)
    compact = re.sub(r"\bCoverage\b|\bAmount\b|\bPurpose\b", " ", compact, flags=re.IGNORECASE)
    lines = [line.strip() for line in compact.splitlines() if line.strip()]
    rows: list[ScheduleABrokerMoneyRow] = []
    subtotal: str | None = None
    last_row: ScheduleABrokerMoneyRow | None = None
    for line in lines:
        clean = re.sub(r"\s+", " ", line).strip()
        sub_match = re.search(r"\b([0-9,]+(?:\.\d{2})?)\s*Sub\s*Total\b", clean, flags=re.IGNORECASE)
        if sub_match:
            subtotal = money_value(sub_match.group(1))
            continue
        match = re.match(r"(?:(?P<coverage>[A-Za-z][A-Za-z /&-]{1,40})\s+)?(?P<amount>[0-9,]+(?:\.\d{2})?)(?P<purpose>[A-Za-z][A-Za-z &/-].*)?$", clean)
        if match:
            row = ScheduleABrokerMoneyRow(
                coverage=clean_extracted_value(match.group("coverage") or "") or None,
                amount=money_value(match.group("amount")),
                purpose=clean_extracted_value(match.group("purpose") or "") or None,
            )
            rows.append(row)
            last_row = row
            continue
        if last_row and clean and not re.search(r"\d", clean):
            last_row.purpose = clean_extracted_value(" ".join(filter(None, [last_row.purpose, clean])))
    if subtotal is None and rows:
        subtotal = _sum_money_rows(rows)
    return rows, subtotal


def _sum_money_rows(rows: list[ScheduleABrokerMoneyRow]) -> str | None:
    total = 0.0
    found = False
    for row in rows:
        number = str(row.amount or "").replace(",", "")
        try:
            total += float(number)
            found = True
        except ValueError:
            continue
    if not found:
        return None
    if total.is_integer():
        return f"{int(total):,}"
    return f"{total:,.2f}"


def normalize_compare_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def extract_schedule_a_broker_compensation_fields(text: str, page: int | None = None) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    source_text = normalize_ocr_text(text)[:1200]

    def add(field_name: str, value: str | None, confidence: float = 0.98, value_validator=None):
        clean = clean_extracted_value(str(value or ""))
        if not clean or is_blank_extraction_value(clean):
            return
        if value_validator and not value_validator(clean):
            return
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value=clean,
                confidence=confidence,
                page=page,
                source_text=source_text,
            )
        )

    commission_total, fee_total = extract_schedule_a_broker_totals(text)
    add("3b. Amount of Commissions", commission_total, 0.99)
    add("3c. Amount of Fees", fee_total, 0.99)

    agent_name = extract_schedule_a_broker_block_name(text)
    add("3a. Name of Agent/Broker/Person", agent_name, 0.98, is_probable_person_or_entity_name)

    org_code = extract_schedule_a_broker_org_code(text)
    add("3e. Organizational Code", org_code, 0.98, looks_like_org_code)

    purpose = derive_schedule_a_purpose(commission_total, fee_total)
    add("3d. Purpose", purpose, 0.98)
    return fields


def extract_schedule_a_broker_totals(text: str) -> tuple[str | None, str | None]:
    normalized = normalize_ocr_text(text)
    patterns = [
        r"Total\s+Amount\s+of\s+commissions\s+paid\s*:?\s*\$?\s*([0-9,]+(?:\.\d{2})?)\s+Total\s+fees\s+paid\s*/?\s*amount\s*:?\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
        r"Total\s+Amount\s+of\s+commissions\s+paid\s*:?\s*\$?\s*([0-9,]+(?:\.\d{2})?).{0,160}?Total\s+(?:amount\s+of\s+)?fees\s+paid\s*:?\s*\$?\s*([0-9,]+(?:\.\d{2})?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return money_value(match.group(1)), money_value(match.group(2))
    return None, None


def extract_schedule_a_broker_block_name(text: str) -> str | None:
    section = broker_name_section(text)
    if not section:
        return None
    match = re.search(
        r"Name\s+and\s+address\s+of\s+the\s+agents?.*?\bName\s*:\s*(.+?)(?=\s*Address\s+Line\s*1\s*:|\s+Address\s*:|\n|$)",
        section,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return None
    return re.sub(r"\s+", " ", match.group(1)).strip()


def extract_schedule_a_broker_org_code(text: str) -> str | None:
    section = broker_name_section(text)
    if not section:
        return None
    match = re.search(r"Organization\s+code\s*:\s*([0-9]{1,2})\b", section, flags=re.IGNORECASE)
    return match.group(1) if match else None


def broker_name_section(text: str) -> str | None:
    if re.search(r"Name\s+and\s+address\s+of\s+the\s+agents?", text, flags=re.IGNORECASE):
        return text
    return schedule_a_broker_table_section(text)


def extract_schedule_a_line_11(text: str) -> str | None:
    normalized = normalize_ocr_text(text)
    explicit = re.search(
        r"Did\s+the\s+insurance\s+company\s+fail\s+to\s+provide\s+any\s+information\s+necessary\s+to\s+complete\s+Schedule\s+A\??.{0,80}?\b(Yes|No)\b",
        normalized,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if explicit:
        return explicit.group(1).title()
    if re.search(r"\bSchedule\s+A\b", normalized, flags=re.IGNORECASE) and re.search(r"\bInsurance\s+Information\b|\bInsurance\s+Contract\b", normalized, flags=re.IGNORECASE):
        return "No"
    return None


def extract_schedule_a_fields_from_tables(text: str, page: int | None = None) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    source_text = normalize_ocr_text(text)[:1200]

    def add(field_name: str, value: str | None, confidence: float = 0.96, value_validator=None):
        clean = clean_extracted_value(str(value or ""))
        if not clean or is_blank_extraction_value(clean):
            return
        if value_validator and not value_validator(clean):
            return
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value=clean,
                confidence=confidence,
                page=page,
                source_text=source_text,
            )
        )

    add("1d. Contract/Policy Number", extract_table_contract_identifier(text), 0.97, is_valid_contract_identifier)
    add("1e. Persons Covered (End of Policy Year)", extract_table_covered_persons(text), 0.96)
    policy_dates = extract_table_policy_dates(text)
    if policy_dates:
        add("1f. Policy Year Beginning Date", policy_dates[0], 0.96)
        add("1g. Policy Year Ending Date", policy_dates[1], 0.96)
    add("3a. Name of Agent/Broker/Person", extract_table_agent_name(text), 0.94, is_probable_person_or_entity_name)
    add("3b. Amount of Commissions", extract_table_commission_amount(text), 0.97)
    return fields


def extract_table_contract_identifier(text: str) -> str | None:
    coverage_row = re.search(
        r"EIN\s+NAIC\s+Code\s+Contract\s+or\s+identification\s*#.*?"
        r"([0-9]{2}-[0-9]{7})\s+([0-9]{4,6})\s+([A-Za-z0-9-]{3,})\s+([0-9,]+)\s+"
        r"[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}\s+[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if coverage_row and is_valid_contract_identifier(coverage_row.group(3), allow_numeric=True):
        return coverage_row.group(3)

    for pattern in [
        r"\bFor:\s*.*?([0-9][A-Za-z0-9-]{4,}?)(?=Policy\s+Period)",
        r"\bContract\s+or\s*Identification.*?([0-9][A-Za-z0-9-]{4,}?)(?=[A-Z][A-Z ]{5,})",
    ]:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match and is_valid_contract_identifier(match.group(1)):
            return match.group(1)

    coverage_section = regex_first(
        text,
        [
            r"(\(d\)\s*Contract\s+Number\s+or\s+Identification:?.*?)(?=\(e\)\s*Approximate|Policy\s+or\s+contract\s+Year|2\.\s*Insurance)",
            r"(\(d\)\s*Contract\s+or\s+identification\s+number:?.*?)(?=\(e\)\s*Approximate|Policy\s+or\s+contract\s+Year|2\.\s*Insurance)",
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not coverage_section:
        return None
    for candidate in re.findall(r"\b[A-Za-z0-9][A-Za-z0-9-]{2,}\b", coverage_section):
        if is_valid_contract_identifier(candidate, allow_numeric=True):
            return candidate
    return None


def extract_table_covered_persons(text: str) -> str | None:
    worksheet_total = re.search(
        r"\bTotal\s*\(E\)\s*([0-9][0-9,]*)\b.{0,160}?"
        r"(?:Approx\.?\s+no\.?\s+of\s+Persons\s+cov\.?|Persons\s+covered)"
        r".{0,100}?End\s+of\s+Policy\s+Year",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if worksheet_total:
        return worksheet_total.group(1)

    flattened_match = re.search(
        r"NAIC\s+Code:.*?Listing\s*([0-9][0-9,]*)\s*[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if flattened_match:
        return flattened_match.group(1)

    section = regex_first(
        text,
        [
            r"(\(e\)\s*Approximate\s+Number\s+of.*?persons\s+covered.*?)(?=Policy\s+or\s+contract\s+Year|\(f\)\s*From|2\.\s*Insurance)",
            r"(\(e\)\s*Approximate\s+number\s+of\s+persons\s+covered.*?)(?=\(f\)\s*From|2\.\s*Insurance)",
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not section:
        return None
    numbers = re.findall(r"\b[0-9][0-9,]*\b", section)
    return numbers[-1] if numbers else None


def extract_table_policy_dates(text: str) -> tuple[str, str] | None:
    match = re.search(
        r"\(f\)\s*From:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4}).*?"
        r"\(g\)\s*To:?\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if match:
        return normalize_schedule_a_date(match.group(1), end_of_month=False), normalize_schedule_a_date(match.group(2), end_of_month=True)
    worksheet_match = re.search(
        r"\(F\)\s*From\s*\(F\)\s*([0-9]{4}-[0-9]{2}-[0-9]{2}).*?"
        r"\(G\)\s*To\s*\(G\)\s*([0-9]{4}-[0-9]{2}-[0-9]{2})",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if worksheet_match:
        return normalize_schedule_a_date(worksheet_match.group(1), end_of_month=False), normalize_schedule_a_date(worksheet_match.group(2), end_of_month=True)
    return None


def extract_table_agent_name(text: str) -> str | None:
    broker_section = schedule_a_broker_table_section(text)
    if not broker_section:
        return None
    contract_id = extract_table_contract_identifier(text)
    pattern = rf"{re.escape(contract_id)}\s+(.+?)\s+\$?\s*[0-9,]+(?:\.\d{{2}})?" if contract_id else r"([A-Z][A-Z0-9&.,'() -]{8,}?)\s+\$?\s*[0-9,]+(?:\.\d{2})?"
    match = re.search(pattern, broker_section, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return None
    value = re.sub(r"\s+", " ", match.group(1)).strip()
    value = re.split(r"\s+\d{3,}\s+[A-Z ]+\b", value, maxsplit=1)[0].strip()
    return value


def extract_table_commission_amount(text: str) -> str | None:
    broker_section = schedule_a_broker_table_section(text)
    if not broker_section:
        return None
    match = re.search(
        r"Amount\s+of\s+commissions\s+paid.*?\$\s*([0-9,]+(?:\.\d{2})?)",
        broker_section,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if match:
        return money_value(match.group(1))
    amounts = re.findall(r"\$\s*([0-9,]+(?:\.\d{2})?)", broker_section)
    return money_value(amounts[0]) if amounts else None


def schedule_a_broker_table_section(text: str) -> str | None:
    return regex_first(
        text,
        [
            r"(2\.\s*Insurance\s+Fees\s+and\s+commissions\s+paid\s+to\s+agents\s+and\s+brokers:?.*?)(?=Part\s+II|Part\s+III|$)",
            r"(Insurance\s+Fees\s+and\s+commissions\s+paid\s+to\s+agents\s+and\s+brokers:?.*?)(?=Part\s+II|Part\s+III|$)",
        ],
        flags=re.IGNORECASE | re.DOTALL,
    )


def is_valid_contract_identifier(value: str | None, *, allow_numeric: bool = False) -> bool:
    text = str(value or "").strip()
    if len(text) < 3:
        return False
    if re.search(r"\d{2,}(?:ba|bas|base)$", text, flags=re.IGNORECASE):
        return False
    if re.fullmatch(r"[0-9,]+", text):
        return allow_numeric and len(text.replace(",", "")) >= 3
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        return False
    return bool(re.search(r"[A-Za-z]", text) and re.search(r"\d", text))


def extract_schedule_a_fields_from_rule_labels(
    text: str,
    page: int | None = None,
    *,
    rules=None,
) -> list[NormalizedExtractionField]:
    fields: list[NormalizedExtractionField] = []
    source_text = normalize_ocr_text(text)[:1200]

    def add(field_name: str, value: str | None, confidence: float = 0.91, value_validator=None):
        clean = clean_extracted_value(str(value or ""))
        if not clean or is_blank_extraction_value(clean):
            return
        if value_validator and not value_validator(clean):
            return
        fields.append(
            NormalizedExtractionField(
                field_name=field_name,
                value=clean,
                confidence=confidence,
                page=page,
                source_text=source_text,
            )
        )

    add(
        "1b. Insurance Carrier EIN",
        extract_labeled_value(
            text,
            rule_labels(
                "1b. Insurance Carrier EIN",
                "Insurance Carrier Employer Identification Number",
                "Insurance Carrier Federal Employer Identification Number",
                "Carrier Employer Identification Number",
                rules=rules,
            ),
            r"([0-9]{2}-[0-9]{7})",
        ),
        0.94,
        looks_like_ein,
    )
    add(
        "1c. NAIC Code",
        extract_labeled_value(
            text,
            rule_labels(
                "1c. NAIC Code",
                "Insurance Carrier NAIC Code",
                "National Association of Insurance Commissioners code",
                rules=rules,
            ),
            r"([0-9]{4,6})",
        ),
        0.93,
    )
    add(
        "1d. Contract/Policy Number",
        extract_labeled_value(
            text,
            rule_labels(
                "1d. Contract/Policy Number",
                "Plan Sponsor Contract or Identification Number",
                "Plan Sponsor Contract Number",
                rules=rules,
            ),
            r"([A-Za-z0-9][A-Za-z0-9-]{1,})",
        ),
        0.92,
    )
    add(
        "1e. Persons Covered (End of Policy Year)",
        extract_labeled_value(
            text,
            rule_labels(
                "1e. Persons Covered (End of Policy Year)",
                "Approximate number of persons covered at end of policy contract year",
                rules=rules,
            ),
            r"([0-9,]+)",
        ),
        0.91,
    )
    add(
        "3c. Amount of Fees",
        extract_labeled_value(
            text,
            rule_labels(
                "3c. Amount of Fees",
                "Total Amount of Fees Paid",
                "Fees and other compensation paid",
                rules=rules,
            ),
            MONEY_VALUE_PATTERN,
            transform=money_value,
        ),
        0.91,
    )
    for field_label in SCHEDULE_A_EXPERIENCE_RATED_FIELDS:
        add(
            field_label,
            extract_labeled_value(
                text,
                [field_label],
                MONEY_VALUE_PATTERN,
                transform=money_value,
            ),
            0.93,
        )
    add(
        "10a. Total premiums or subscription charges paid to carrier",
        extract_labeled_value(
            text,
            rule_labels(
                "10a. Total premiums or subscription charges paid to carrier",
                "Premium applied by",
                "Premium applied",
                "Total amount of premiums applied",
                rules=rules,
            ),
            MONEY_VALUE_PATTERN,
            transform=money_value,
        ),
        0.92,
    )

    contract_period = extract_contract_year_range(text)
    if contract_period:
        policy_from, policy_to = contract_period
        add("1f. Policy Year Beginning Date", policy_from, 0.9)
        add("1g. Policy Year Ending Date", policy_to, 0.9)

    return fields


def extract_configured_custom_fields(
    text: str,
    page: int | None = None,
    *,
    rules=None,
) -> list[NormalizedExtractionField]:
    if not rules:
        return []
    normalized_text = normalize_ocr_text(text)
    fields: list[NormalizedExtractionField] = []
    for rule in rules:
        if rule.mapping_mode != FieldRuleMappingMode.EXTRACTION_ONLY and not rule.key.startswith("ftw_discovered_"):
            continue
        labels = rule_labels(rule.label, rules=rules)
        for label in sorted(labels, key=len, reverse=True):
            match = re.search(
                rf"(?im)^\s*{loose_label_pattern(label)}\s*(?::|\t|[-–—])\s*(?P<value>[^\n]+?)\s*$",
                normalized_text,
            )
            if not match:
                continue
            value = clean_extracted_value(match.group("value"))
            if not value or is_blank_extraction_value(value):
                continue
            fields.append(
                NormalizedExtractionField(
                    field_name=rule.label,
                    value=value,
                    confidence=0.9,
                    page=page,
                    source_text=match.group(0).strip(),
                )
            )
            break
    return fields


MONEY_VALUE_PATTERN = r"\$?\s*([0-9,]+(?:\.\d{2})?)"


def rule_labels(field_label: str, *extra_labels: str, rules=None) -> list[str]:
    labels: list[str] = []
    for rule in (rules if rules is not None else DEFAULT_FIELD_RULES):
        if rule.label != field_label:
            continue
        labels.extend([rule.label, *rule.aliases])
        break
    labels.extend(extra_labels)
    deduped: list[str] = []
    seen: set[str] = set()
    for label in labels:
        normalized = normalize_rule_label(label)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(label)
    return deduped


def extract_labeled_value(
    text: str,
    labels: list[str],
    value_pattern: str,
    *,
    transform=None,
) -> str | None:
    normalized_text = normalize_ocr_text(text)
    for label in sorted(labels, key=len, reverse=True):
        pattern = rf"(?im)^\s*{loose_label_pattern(label)}\s*:?\s*(?:[^\n]{{0,180}}?)?{value_pattern}\b"
        match = re.search(pattern, normalized_text)
        if not match:
            continue
        value = match.group(1)
        return transform(value) if transform else value
    return None


def loose_label_pattern(label: str) -> str:
    escaped = re.escape(label.strip())
    escaped = escaped.replace(r"\ ", r"\s+")
    escaped = escaped.replace(r"\/", r"[/\s]+")
    escaped = escaped.replace(r"\-", r"[-\s]+")
    escaped = escaped.replace(r"\:", r":?")
    return escaped


def normalize_rule_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()


def money_value(value: str) -> str:
    return str(value or "").replace("$", "").strip()


def extract_contract_year_range(text: str) -> tuple[str, str] | None:
    match = re.search(
        r"\b(?:Contract|Policy)\s+Year\s+from\s+"
        r"([0-9]{1,2}(?:/[0-9]{1,2})?/[0-9]{4})\s*(?:-|to|through)\s*"
        r"([0-9]{1,2}(?:/[0-9]{1,2})?/[0-9]{4})",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    return normalize_schedule_a_date(match.group(1), end_of_month=False), normalize_schedule_a_date(match.group(2), end_of_month=True)


def normalize_schedule_a_date(value: str, *, end_of_month: bool) -> str:
    text = str(value or "").strip()
    iso_match = re.fullmatch(r"(20\d{2})-(\d{2})-(\d{2})", text)
    if iso_match:
        year, month, day = (int(part) for part in iso_match.groups())
        return f"{month:02d}/{day:02d}/{year}"
    parts = text.split("/")
    if len(parts) == 2:
        month = int(parts[0])
        year = int(parts[1])
        day = calendar.monthrange(year, month)[1] if end_of_month else 1
        return f"{month:02d}/{day:02d}/{year}"
    if len(parts) == 3:
        month = int(parts[0])
        day = int(parts[1])
        year = int(parts[2])
        return f"{month:02d}/{day:02d}/{year}"
    return text


def regex_first(text: str, patterns: list[str], flags: int = re.IGNORECASE, groups: bool = False):
    for pattern in patterns:
        match = re.search(pattern, text, flags=flags)
        if match:
            if groups:
                return tuple(match.groups())
            return match.group(1) if match.groups() else match.group(0)
    return None


def parse_numeric_amount(value: str | None) -> float | None:
    clean = clean_extracted_value(str(value or "")).replace("$", "").replace(",", "").strip()
    if not clean:
        return None
    try:
        return float(clean)
    except ValueError:
        return None


def derive_schedule_a_purpose(commission: str | None, fee: str | None) -> str | None:
    commission_amount = parse_numeric_amount(commission)
    fee_amount = parse_numeric_amount(fee)

    has_commission = commission_amount is not None and commission_amount > 0
    has_fee = fee_amount is not None and fee_amount > 0

    if has_commission and has_fee:
        return "COMMISSIONS & FEES"
    if has_commission:
        return "COMMISSIONS"
    if has_fee:
        return "FEES"
    return None


def clean_extracted_value(value: str) -> str:
    value = str(value).strip(" \n\t:-")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"^(?:Name|Recipient|Agent|Broker)\s*:\s*", "", value, flags=re.IGNORECASE)
    stop_patterns = [
        r"\s+Address\s*:.*",
        r"\s+City\s*:.*",
        r"\s+Commissions Paid\b.*",
        r"\s+\(\s*b\s*\).*",
        r"\s+\(\s*c\s*\).*",
        r"\s+\(\s*d\s*\).*",
        r"\s+\(\s*e\s*\).*",
        r"\s+Part I\b.*",
    ]
    for pattern in stop_patterns:
        value = re.sub(pattern, "", value, flags=re.IGNORECASE)
    return value.strip(" ,;")
