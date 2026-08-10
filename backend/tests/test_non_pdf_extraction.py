"""A Schedule A that arrives as a spreadsheet must extract like one.

Once intake accepted Excel, CSV and email attachments, the first live test
showed the documents reaching the dashboard with a single field populated.
The reason was that the label-driven parsing - which is the strongest part of
Schedule A extraction - only ever ran on PDFs, so a spreadsheet was left with
whatever the AI extractor happened to return.
"""
import unittest
from io import BytesIO
import zipfile

from app.services.extractor import (
    extract_document_text_pages,
    extract_fields_from_document_text,
    extract_tabular_broker_rows,
)

CSV_EXPORT = b"""Field,Value
Name of Insurance Carrier,Guardian Life Insurance Company of America
EIN,13-5123390
NAIC Code,64246
Contract/Policy Number,CSV-455102
Policy Year Beginning Date,01/01/2025
Policy Year Ending Date,12/31/2025
Persons Covered End of Policy Year,88
Total premiums paid,45230.10
Name of Agent/Broker,CSV Test Insurance Partners LLC
Amount of Commissions,3120.45
Amount of Fees,0.00
Purpose,Standard Commissions
"""


def _workbook(rows, sheet_name="Schedule A"):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _docx_table(rows):
    namespace = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    table_rows = []
    for row in rows:
        cells = "".join(
            f'<w:tc><w:p><w:r><w:t>{value}</w:t></w:r></w:p></w:tc>'
            for value in row
        )
        table_rows.append(f"<w:tr>{cells}</w:tr>")
    document = (
        f'<w:document xmlns:w="{namespace}"><w:body><w:tbl>'
        + "".join(table_rows)
        + "</w:tbl></w:body></w:document>"
    )
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("word/document.xml", document)
    return buffer.getvalue()


CARRIER_SHEET = [
    ["SCHEDULE A INFORMATION FOR FORM 5500"],
    ["Name of Insurance Carrier", "American Life Insurance Company"],
    ["EIN", "51-0104066"],
    ["NAIC Code", "60488"],
    ["Contract/Policy Number", "999001"],
    ["Persons covered at end of policy year", 412],
    ["Total premiums paid", 218450.75],
    [],
    [
        "Agent Number",
        "Name and Address of Each Recipient of Fees and/or Commissions",
        "Amount of Commissions Paid",
        "Amount of Fees Paid",
        "Purpose for Which Paid",
    ],
    ["TIG-778120", "Format Test Brokerage LLC, 500 Market Street, Boston MA 02110", 7412.33, 615.00, "Standard Commissions"],
]


class SpreadsheetExtractionTests(unittest.TestCase):
    def test_a_carrier_workbook_extracts_the_same_fields_as_a_statement(self):
        fields = {f.field_name: f.value for f in extract_fields_from_document_text(_workbook(CARRIER_SHEET), "carrier.xlsx")}
        self.assertEqual(fields["1a. Name of Insurance Company"], "American Life Insurance Company")
        self.assertEqual(fields["1b. Insurance Carrier EIN"], "51-0104066")
        self.assertEqual(fields["1c. NAIC Code"], "60488")
        self.assertEqual(fields["1d. Contract/Policy Number"], "999001")
        self.assertEqual(fields["1e. Persons Covered (End of Policy Year)"], "412")

    def test_the_broker_row_of_a_workbook_is_read_without_currency_signs(self):
        fields = {f.field_name: f.value for f in extract_fields_from_document_text(_workbook(CARRIER_SHEET), "carrier.xlsx")}
        # The agent number is a separate cell and must not end up in the name.
        self.assertEqual(fields["3a. Name of Agent/Broker/Person"], "Format Test Brokerage LLC")
        self.assertEqual(fields["3b. Amount of Commissions"], "7,412.33")
        self.assertEqual(fields["3d. Purpose"], "Standard Commissions")

    def test_rows_of_numbers_without_a_commission_column_are_not_broker_rows(self):
        sheet = _workbook(
            [
                ["Participant counts by month"],
                ["January", 120, 118],
                ["February", 121, 119],
            ]
        )
        pages = extract_document_text_pages(sheet, "counts.xlsx")
        self.assertEqual(extract_tabular_broker_rows(pages), [])

    def test_all_workbook_sheets_are_extracted(self):
        from openpyxl import Workbook

        workbook = Workbook()
        workbook.active.title = "Cover"
        workbook.active.append(["Document", "Schedule A"])
        details = workbook.create_sheet("Contract Details")
        details.append(["Name of Insurance Carrier", "Second Sheet Carrier LLC"])
        details.append(["Contract/Policy Number", "MULTI-2025"])
        buffer = BytesIO()
        workbook.save(buffer)

        fields = {
            field.field_name: field.value
            for field in extract_fields_from_document_text(buffer.getvalue(), "multi-sheet.xlsx")
        }

        self.assertEqual(fields["1a. Name of Insurance Company"], "Second Sheet Carrier LLC")
        self.assertEqual(fields["1d. Contract/Policy Number"], "MULTI-2025")


class WordExtractionTests(unittest.TestCase):
    def test_label_value_rows_in_a_word_table_are_extracted(self):
        document = _docx_table(
            [
                ["Name of Insurance Carrier", "Word Table Carrier, Inc."],
                ["EIN", "12-3456789"],
                ["Amount of Fees", "2,450.00"],
            ]
        )

        fields = {field.field_name: field.value for field in extract_fields_from_document_text(document, "table.docx")}

        self.assertEqual(fields["1a. Name of Insurance Company"], "Word Table Carrier, Inc.")
        self.assertEqual(fields["1b. Insurance Carrier EIN"], "12-3456789")
        self.assertEqual(fields["3c. Amount of Fees"], "2,450.00")


class DelimitedExportTests(unittest.TestCase):
    def test_quoted_commas_in_csv_values_are_preserved(self):
        csv_bytes = b'Field,Value\n"Name of Insurance Carrier","Health Advocate Solutions, Inc."\n"Amount of Fees","5,134.75"\n'

        fields = {f.field_name: f.value for f in extract_fields_from_document_text(csv_bytes, "quoted.csv")}

        self.assertEqual(fields["1a. Name of Insurance Company"], "Health Advocate Solutions, Inc.")
        self.assertEqual(fields["3c. Amount of Fees"], "5,134.75")

    def test_a_csv_export_extracts_the_full_field_set(self):
        fields = {f.field_name: f.value for f in extract_fields_from_document_text(CSV_EXPORT, "export.csv")}
        for name, expected in (
            ("1a. Name of Insurance Company", "Guardian Life Insurance Company of America"),
            ("1c. NAIC Code", "64246"),
            ("1d. Contract/Policy Number", "CSV-455102"),
            ("3a. Name of Agent/Broker/Person", "CSV Test Insurance Partners LLC"),
            ("3b. Amount of Commissions", "3120.45"),
            ("3d. Purpose", "Standard Commissions"),
        ):
            with self.subTest(field=name):
                self.assertEqual(fields[name], expected)

    def test_an_unreadable_workbook_returns_nothing_rather_than_raising(self):
        self.assertEqual(extract_document_text_pages(b"not a workbook", "broken.xlsx"), [])
        self.assertEqual(extract_fields_from_document_text(b"not a workbook", "broken.xlsx"), [])


class EmailBodyExtractionTests(unittest.TestCase):
    def test_carrier_values_written_in_an_email_body_are_extracted(self):
        body = b"""Good afternoon,
Legal Name: Health Advocate Solutions, Inc.
Address: 3043 Walton Road, Plymouth Meeting, PA 19642
EIN: 23-3080019
PEPM Fees Paid (January 2025 through December 2025): $5,134.75
Approximate employee lives covered at end of calendar year (December 2025):
EAP 490
There was no indirect compensation for the stated period.
"""

        fields = {f.field_name: f.value for f in extract_fields_from_document_text(body, "Health Advocate email body.txt")}

        self.assertEqual(fields["1a. Name of Insurance Company"], "Health Advocate Solutions, Inc.")
        self.assertEqual(fields["1b. Insurance Carrier EIN"], "23-3080019")
        self.assertEqual(fields["1e. Persons Covered (End of Policy Year)"], "490")
        self.assertEqual(fields["1f. Policy Year Beginning Date"], "01/01/2025")
        self.assertEqual(fields["1g. Policy Year Ending Date"], "12/31/2025")
        self.assertEqual(fields["3c. Amount of Fees"], "5,134.75")


if __name__ == "__main__":
    unittest.main()
