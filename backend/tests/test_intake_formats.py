"""Clients do not only send PDFs.

A Schedule A arrives as a spreadsheet, a scan, or an Outlook email with the
real document attached. Before this, anything that was not a PDF (or a Word
file recognised by its contents) was silently ignored at intake - the file
never appeared on the dashboard and nobody was told why.
"""
import email.message
import unittest
from io import BytesIO
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from app.services.intake_formats import (
    SUPPORTED_INTAKE_EXTENSIONS,
    is_supported_intake_file,
    normalize_intake_document,
    normalize_intake_documents,
)


def _eml(attachments: list[tuple[str, bytes, str, str]], body: str = "Please see attached.") -> bytes:
    """Build a real email with attachments, the way a carrier would send one."""
    message = email.message.EmailMessage()
    message["Subject"] = "RE: [EXT] AlphaSights Schedule A - Health Advocate"
    message["From"] = "carrier@example.com"
    message["To"] = "filings@erisapros.com"
    message.set_content(body)
    for name, data, maintype, subtype in attachments:
        message.add_attachment(data, maintype=maintype, subtype=subtype, filename=name)
    return message.as_bytes()


def _xls(rows: list[list], sheet_name: str = "Schedule A") -> bytes:
    """A pre-2007 .xls workbook, the format some carriers still produce."""
    import xlwt

    book = xlwt.Workbook()
    sheet = book.add_sheet(sheet_name)
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _docx_bytes(text: str = "Schedule A") -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/word/document.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            "</Types>",
        )
        archive.writestr(
            "word/document.xml",
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            f"<w:body><w:p><w:r><w:t>{text}</w:t></w:r></w:p></w:body></w:document>",
        )
    return buffer.getvalue()


class PassThroughTests(unittest.TestCase):
    def test_ordinary_formats_are_passed_through_untouched(self):
        for name in ("carrier.pdf", "form.docx", "data.xlsx", "rows.csv", "notes.txt", "scan.png"):
            with self.subTest(name=name):
                result = normalize_intake_document(name, b"payload-" + name.encode())
                self.assertEqual(result.file_name, name)
                self.assertEqual(result.file_bytes, b"payload-" + name.encode())
                self.assertFalse(result.converted)
                self.assertIsNone(result.note)

    def test_the_formats_clients_actually_send_are_accepted(self):
        for name in (
            "a.pdf",
            "a.docx",
            "a.doc",
            "a.xlsx",
            "a.xls",
            "a.csv",
            "a.txt",
            "a.png",
            "a.jpg",
            "a.jpeg",
            "a.tiff",
            "a.msg",
            "a.eml",
        ):
            with self.subTest(name=name):
                self.assertTrue(is_supported_intake_file(name))

    def test_files_that_are_not_documents_are_still_rejected(self):
        for name in ("archive.zip", "database.db", "video.mp4", "noextension"):
            with self.subTest(name=name):
                self.assertFalse(is_supported_intake_file(name))

    def test_extension_matching_ignores_case(self):
        self.assertTrue(is_supported_intake_file("CHLIC_09995A.PDF"))
        self.assertIn(".pdf", SUPPORTED_INTAKE_EXTENSIONS)

    def test_a_docx_with_a_pdf_file_name_is_routed_by_its_real_format(self):
        raw = _docx_bytes("Carrier EIN: 12-3456789")

        result = normalize_intake_document("carrier-schedule-a.pdf", raw)

        self.assertEqual(result.file_name, "carrier-schedule-a.docx")
        self.assertEqual(result.original_file_name, "carrier-schedule-a.pdf")
        self.assertEqual(result.file_bytes, raw)
        self.assertIn("detected as DOCX", result.conversion or "")


class EmailUnwrappingTests(unittest.TestCase):
    def test_outlook_html_body_is_used_when_the_only_attachment_is_a_logo(self):
        storage = "__attach_version1.0_#00000000"
        streams = {
            "__substg1.0_10130102": b"<html><body><p>Legal Name: Health Advocate Solutions, Inc.</p><p>EIN: 23-3080019</p></body></html>",
            f"{storage}/__substg1.0_37010102": b"logo bytes",
            f"{storage}/__substg1.0_3707001F": "image001.png".encode("utf-16-le"),
        }

        class FakeOle:
            def __enter__(self): return self
            def __exit__(self, *_args): return None
            def listdir(self, **_kwargs):
                return [[storage], *[key.split("/") for key in streams if "/" in key]]
            def exists(self, path): return str(path).replace("\\", "/") in streams
            def openstream(self, path):
                key = "/".join(path) if isinstance(path, list) else str(path)
                return BytesIO(streams[key])

        with patch("olefile.OleFileIO", return_value=FakeOle()):
            results = normalize_intake_documents("Health Advocate.msg", b"fake msg")

        self.assertEqual([item.file_name for item in results], ["Health Advocate email body.txt"])
        self.assertIn("Health Advocate Solutions, Inc.", results[0].file_bytes.decode("utf-8"))

    def test_email_body_is_kept_and_signature_logo_is_ignored(self):
        raw = _eml(
            [("image001.png", b"not-the-schedule-a" * 2000, "image", "png")],
            body="""Good afternoon,
Legal Name: Health Advocate Solutions, Inc.
Address: 3043 Walton Road, Plymouth Meeting, PA 19642
EIN: 23-3080019
PEPM Fees Paid (January 2025 through December 2025): $5,134.75
Approximate employee lives covered at end of calendar year: 490
There was no indirect compensation for the stated period.
""",
        )

        results = normalize_intake_documents("Health Advocate.eml", raw)

        self.assertEqual([item.file_name for item in results], ["Health Advocate email body.txt"])
        body = results[0].file_bytes.decode("utf-8")
        self.assertIn("Health Advocate Solutions, Inc.", body)
        self.assertIn("23-3080019", body)
        self.assertIn("$5,134.75", body)
        self.assertNotIn("image001.png", [item.file_name for item in results])

    def test_the_attachment_is_filed_not_the_email(self):
        raw = _eml([("AlphaSights Schedule A.pdf", b"%PDF-1.4 schedule a", "application", "pdf")])
        result = normalize_intake_document("4. RE_ [EXT]AlphaSights Schedule A.eml", raw)

        self.assertEqual(result.file_name, "AlphaSights Schedule A.pdf")
        self.assertEqual(result.file_bytes, b"%PDF-1.4 schedule a")
        self.assertTrue(result.converted)
        self.assertEqual(result.original_file_name, "4. RE_ [EXT]AlphaSights Schedule A.eml")

    def test_a_schedule_a_attachment_wins_over_a_signature_logo(self):
        raw = _eml(
            [
                ("logo.png", b"x" * 5000, "image", "png"),
                ("Schedule A 2025.pdf", b"%PDF-1.4 small but right", "application", "pdf"),
            ]
        )
        result = normalize_intake_document("carrier.eml", raw)
        self.assertEqual(result.file_name, "Schedule A 2025.pdf")

    def test_the_largest_usable_attachment_wins_when_no_name_matches(self):
        raw = _eml(
            [
                ("tiny.pdf", b"%PDF-1.4 tiny", "application", "pdf"),
                ("bigger.pdf", b"%PDF-1.4 " + b"x" * 2000, "application", "pdf"),
            ]
        )
        result = normalize_intake_document("carrier.eml", raw)
        self.assertEqual(result.file_name, "bigger.pdf")

    def test_an_email_with_no_attachment_is_kept_and_flagged(self):
        raw = _eml([], body="The Schedule A is pasted below, not attached.")
        result = normalize_intake_document("carrier.eml", raw)

        self.assertEqual(result.file_name, "carrier.eml")
        self.assertFalse(result.converted)
        self.assertIn("no attachment", (result.note or "").lower())

    def test_an_email_whose_only_attachment_is_unusable_is_flagged(self):
        raw = _eml([("archive.zip", b"PK\x03\x04 zip", "application", "zip")])
        result = normalize_intake_document("carrier.eml", raw)
        self.assertEqual(result.file_name, "carrier.eml")
        self.assertIsNotNone(result.note)

    def test_a_signature_logo_is_never_treated_as_a_schedule_a(self):
        raw = _eml([("image001.png", b"company-logo" * 2000, "image", "png")])

        result = normalize_intake_document("carrier.eml", raw)

        self.assertEqual(result.file_name, "carrier.eml")
        self.assertIn("manual", (result.note or "").lower())

    def test_a_corrupt_email_does_not_raise(self):
        result = normalize_intake_document("broken.msg", b"this is not an outlook file at all")
        self.assertEqual(result.file_name, "broken.msg")
        self.assertIsNotNone(result.note)

    def test_attachment_paths_are_stripped_to_a_file_name(self):
        raw = _eml([("C:\\Users\\carrier\\Schedule A.pdf", b"%PDF-1.4 x", "application", "pdf")])
        result = normalize_intake_document("carrier.eml", raw)
        self.assertEqual(result.file_name, "Schedule A.pdf")


class LegacyExcelTests(unittest.TestCase):
    def test_a_legacy_xls_is_converted_to_xlsx(self):
        try:
            import xlwt  # noqa: F401
        except ImportError:
            self.skipTest("xlwt is only needed to author the test fixture")

        raw = _xls(
            [
                ["Schedule A - The International Group, Inc.", "", ""],
                ["Contract Number", "187319", ""],
                ["Total Premium", 125000.5, ""],
                ["Commission Paid", 4200, ""],
            ]
        )
        result = normalize_intake_document(
            "1. The Interational Group Inc _187319_01-01-25 thru 12-31-25_Schedule A ALIC Form.xls",
            raw,
        )

        self.assertTrue(result.file_name.endswith(".xlsx"))
        self.assertTrue(result.converted)
        self.assertIsNone(result.note)

        # The converted workbook must actually contain the values, otherwise
        # extraction would run against an empty sheet.
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(result.file_bytes))
        sheet = workbook["Schedule A"]
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn("187319", [str(value) for value in values])
        self.assertIn(125000.5, values)
        self.assertIn("Commission Paid", values)

    def test_a_corrupt_xls_is_kept_and_flagged_rather_than_lost(self):
        result = normalize_intake_document("broken.xls", b"not really a workbook")
        self.assertEqual(result.file_name, "broken.xls")
        self.assertFalse(result.converted)
        self.assertIn("manual", (result.note or "").lower())


if __name__ == "__main__":
    unittest.main()
